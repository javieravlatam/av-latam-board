"""
parse_ventas_grupo_cl.py — Enriquecimiento SIC Chile: FECHA VENCIMIENTO + RUT por folio
=========================================================================================
Lee el archivo "Ventas*GRUPO AV LATAM*.xlsx" del inbox y genera
apps/sic_av/data/vencimientos_cl.json con metadatos de vencimiento por documento.

NO produce monto cobrado (no existe esa información en la fuente).
NO modifica AVBOARD, dashboards ni datos históricos.
NO duplica facturas ya presentes en TX_CL (ese pipeline sigue siendo update_avboard.py).
Este script SOLO extrae: folio → fecha_vencimiento, rut, razón social, región, comentarios.

Ejecutar desde la raíz del repo:
    python3 scripts/parse_ventas_grupo_cl.py

O como parte del pipeline principal (llamado desde update_avboard.py si se integra).

Schema esperado del archivo Excel (detección automática):
    Columnas requeridas: MES, Rut, Razón Social, Fecha, Vendedor, Documento,
                         Folio, PAÍS, FECHA VENCIMIENTO, Negocio
    Headers en fila 2 (fila 1 vacía).
    Datos desde fila 3.

Salida: apps/sic_av/data/vencimientos_cl.json
"""

import os
import json
import glob
import logging
from datetime import datetime

log = logging.getLogger(__name__)

# Vendedores excluidos del SIC (mismo listado que sic_data_adapter.js NO_COMERCIAL)
NO_COMERCIAL = {"OFICINA", "LABORATORIO", "EN TERRENO 1", "JAVIER ALMEIDA"}

COLUMNAS_REQUERIDAS = {
    "MES", "Rut", "Razón Social", "Fecha", "Vendedor",
    "Documento", "Folio", "PAÍS", "FECHA VENCIMIENTO", "Negocio"
}

SCHEMA_VERSION = "ventas_grupo_av_latam_v1"


def fmt_date(v):
    """Normaliza fecha a 'YYYY-MM-DD' o None."""
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s or s == "None":
        return None
    # dd/mm/yyyy (formato Excel chile)
    if "/" in s:
        try:
            return datetime.strptime(s, "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s[:10]


def normalizar_folio(folio_raw):
    """Convierte folio Excel (int, float o str) a string entero limpio."""
    if folio_raw is None:
        return None
    try:
        return str(int(float(str(folio_raw))))
    except (ValueError, TypeError):
        return str(folio_raw).strip()


def detectar_archivo_inbox(inbox_dir):
    """
    Busca el archivo más reciente que coincida con 'Ventas*GRUPO AV LATAM*.xlsx'.
    Retorna la ruta o None si no se encuentra.
    """
    patron = os.path.join(inbox_dir, "Ventas*GRUPO AV LATAM*.xlsx")
    archivos = sorted(glob.glob(patron), key=os.path.getmtime, reverse=True)
    return archivos[0] if archivos else None


def validar_schema(header_row):
    """Verifica que las columnas requeridas estén presentes."""
    headers_set = set(h.strip() for h in header_row if h)
    faltantes = COLUMNAS_REQUERIDAS - headers_set
    return faltantes


def procesar_archivo(ruta_excel):
    """
    Parsea el archivo Excel y retorna (docs_dict, advertencias, excluidos).

    docs_dict: {folio_str: {...datos de enriquecimiento...}}
    advertencias: lista de strings
    excluidos: lista de dicts con razón de exclusión
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl no instalado. Ejecutar: pip install openpyxl --break-system-packages")

    wb = openpyxl.load_workbook(ruta_excel, data_only=True)

    if "Ventas" not in wb.sheetnames:
        raise ValueError(f"Hoja 'Ventas' no encontrada en {ruta_excel}. Hojas disponibles: {wb.sheetnames}")

    ws = wb["Ventas"]

    # Headers en fila 2 (fila 1 vacía en este formato)
    header_row = [str(c.value).strip() if c.value else "" for c in ws[2]]

    faltantes = validar_schema(header_row)
    if faltantes:
        raise ValueError(
            f"Schema no reconocido en {ruta_excel}. "
            f"Columnas requeridas ausentes: {faltantes}. "
            f"Esperado: {COLUMNAS_REQUERIDAS}"
        )

    docs = {}
    advertencias = []
    excluidos = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        if all(v is None for v in row):
            continue

        r = dict(zip(header_row, row))

        # Filtros básicos
        pais = str(r.get("PAÍS") or "").strip()
        if pais != "CHILE":
            continue

        negocio = str(r.get("Negocio") or "").strip()
        if negocio != "AV":
            log.debug("Fila excluida: negocio=%s (se esperaba 'AV')", negocio)
            continue

        folio_str = normalizar_folio(r.get("Folio"))
        if not folio_str:
            advertencias.append("Fila sin folio detectada — omitida")
            continue

        vend = str(r.get("Vendedor") or "").strip().upper()
        if vend in NO_COMERCIAL:
            excluidos.append({
                "folio": folio_str,
                "vendedor": vend,
                "razon": "NO_COMERCIAL",
                "descripcion": f"Vendedor '{vend}' no es personal comercial (mismo criterio que SICAdapter.NO_COMERCIAL)"
            })
            continue

        # Datos de enriquecimiento
        venc_str      = fmt_date(r.get("FECHA VENCIMIENTO"))
        fecha_emis    = fmt_date(r.get("Fecha"))
        fecha_guia    = fmt_date(r.get("Fecha guía"))
        rut           = str(r.get("Rut") or "").strip()
        razon         = str(r.get("Razón Social") or "").strip()
        doc_tipo      = str(r.get("Documento") or "").strip()
        region        = str(r.get("Región") or "").strip()
        comentarios   = str(r.get("Comentarios") or "").strip()
        total         = 0.0
        try:
            total = float(r.get("Total") or 0)
        except (ValueError, TypeError):
            pass

        if folio_str not in docs:
            docs[folio_str] = {
                "folio":               folio_str,
                "tipo_doc":            doc_tipo,
                "rut":                 rut,
                "razon_social":        razon,
                "vendedor":            vend,
                "fecha_emision":       fecha_emis,
                "fecha_vencimiento":   venc_str,
                "region":              region,
                "comentarios":         comentarios,
                "fecha_guia_despacho": fecha_guia,
                "total_fuente":        total
            }
        else:
            # Líneas adicionales del mismo documento (multi-línea por producto):
            # mantener la primera fecha_vencimiento encontrada.
            if venc_str and not docs[folio_str]["fecha_vencimiento"]:
                docs[folio_str]["fecha_vencimiento"] = venc_str

    # Detectar pares Guía → Factura (mismo cliente + mismo monto > 0)
    # Se marca la guía con reemplazada_por_folio para que la UI pueda advertirlo.
    guias    = {f: d for f, d in docs.items() if "Guía" in d["tipo_doc"]}
    facturas = {f: d for f, d in docs.items() if "Factura" in d["tipo_doc"]}

    for gf, gd in guias.items():
        if gd["total_fuente"] <= 0:
            continue
        for ff, fd in facturas.items():
            if (gd["razon_social"] == fd["razon_social"] and
                    abs(gd["total_fuente"] - fd["total_fuente"]) < 1.0):
                docs[gf]["reemplazada_por_folio"] = ff
                advertencias.append(
                    f"Guía {gf} ({gd['razon_social']}) convertida a Factura {ff} "
                    f"— mismo cliente y monto ({gd['total_fuente']:,.0f} CLP). "
                    "Ambos registros existen en TX_CL; no se deben sumar (doble conteo)."
                )
                break

    return docs, advertencias, excluidos


def generar_vencimientos_cl(inbox_dir, output_path):
    """
    Punto de entrada principal.
    Genera output_path (vencimientos_cl.json) a partir del archivo Excel más
    reciente del inbox. Si no existe el archivo, limpia el JSON anterior.
    """
    ruta_excel = detectar_archivo_inbox(inbox_dir)

    if not ruta_excel:
        log.warning("parse_ventas_grupo_cl: no se encontró 'Ventas*GRUPO AV LATAM*.xlsx' en %s", inbox_dir)
        # Limpiar JSON anterior para no dejar datos stale de un archivo ya removido
        if os.path.exists(output_path):
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump({
                    "schema_version": SCHEMA_VERSION,
                    "fuente": None,
                    "fecha_proceso": datetime.now().strftime("%Y-%m-%d"),
                    "advertencia": "Archivo 'Ventas*GRUPO AV LATAM*.xlsx' no encontrado en inbox — JSON limpiado",
                    "documentos": [],
                    "advertencias": [],
                    "excluidos": []
                }, f, ensure_ascii=False, indent=2)
        return False

    log.info("parse_ventas_grupo_cl: procesando %s", ruta_excel)

    docs, advertencias, excluidos = procesar_archivo(ruta_excel)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    salida = {
        "schema_version": SCHEMA_VERSION,
        "fuente":         os.path.basename(ruta_excel),
        "fecha_proceso":  datetime.now().strftime("%Y-%m-%d"),
        "total_documentos": len(docs),
        "documentos":     list(docs.values()),
        "advertencias":   advertencias,
        "excluidos":      excluidos
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(salida, f, ensure_ascii=False, indent=2)

    log.info(
        "parse_ventas_grupo_cl: %d documentos generados → %s",
        len(docs), output_path
    )
    return True


# ---------------------------------------------------------------------------
# Ejecución directa (standalone)
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    REPO_ROOT  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    INBOX_DIR  = os.path.join(REPO_ROOT, "inbox")
    OUTPUT     = os.path.join(REPO_ROOT, "apps", "sic_av", "data", "vencimientos_cl.json")

    ok = generar_vencimientos_cl(INBOX_DIR, OUTPUT)

    if ok:
        with open(OUTPUT, encoding="utf-8") as f:
            resultado = json.load(f)

        print(f"\n✅  vencimientos_cl.json generado ({resultado['total_documentos']} documentos)")
        print(f"   Fuente: {resultado['fuente']}")
        print(f"   Fecha:  {resultado['fecha_proceso']}")

        if resultado["advertencias"]:
            print(f"\n⚠️  Advertencias ({len(resultado['advertencias'])}):")
            for a in resultado["advertencias"]:
                print(f"   · {a}")

        if resultado["excluidos"]:
            print(f"\nℹ️  Excluidos ({len(resultado['excluidos'])}):")
            for e in resultado["excluidos"]:
                print(f"   · Folio {e['folio']} [{e['vendedor']}]: {e['razon']}")

        print(f"\nDocumentos (folio | tipo | vencimiento | razón social):")
        for d in sorted(resultado["documentos"], key=lambda x: int(x["folio"])):
            conv = f"  → reemplazada por {d.get('reemplazada_por_folio','')}" if d.get("reemplazada_por_folio") else ""
            print(f"  {d['folio']:6} | {d['tipo_doc'][:22]:22} | {d['fecha_vencimiento']} | {d['razon_social'][:28]}{conv}")
    else:
        print("⚠️  No se encontró archivo de ventas en inbox. Sin cambios.")
