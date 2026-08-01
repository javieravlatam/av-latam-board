#!/usr/bin/env python3
"""
build_master_dataset.py
=======================
SSOT — Genera data/master_prices.json desde el Libro Base oficial.

Este script es la única fuente de verdad para:
  - Productos
  - Presentaciones (formatos)
  - SKUs
  - Precios Piso
  - País
  - Costo de fábrica
  - Margen objetivo

Todos los módulos del ecosistema deben leer de data/master_prices.json,
NUNCA directamente del Excel. Los únicos que leen el Excel son:
  - Este script (build_master_dataset.py)
  - update_avboard.py (durante compute_iec_*) — recibe piso_path desde master

Pipeline oficial:
  nuevo libro base AV 2026.xlsx
        ↓
  build_master_dataset.py (este script)
        ↓
  data/master_prices.json  ← SSOT canónico
        ↓
  ┌─────────────────────────────────────────┐
  │  update_avboard.py  (IEC, avboard_data) │
  │  gen_cotizador_json.py (cotizador JSON) │
  │  Panel_IEC (lee avboard_data.js)        │
  │  Executive Board (lee avboard_data.js)  │
  │  SIC (lee avboard_data.js)             │
  └─────────────────────────────────────────┘

Ejecutar: python3 scripts/build_master_dataset.py
Output:   data/master_prices.json
"""
import json
import pathlib
import datetime
import openpyxl

ROOT  = pathlib.Path(__file__).parent.parent
INBOX = ROOT / 'inbox'
DATA  = ROOT / 'data'
DATA.mkdir(exist_ok=True)

LIBRO_BASE_GLOB = [
    'nuevo libro base AV 2026*.xlsx',
    'Libro Base AV*.xlsx',
]

SHEET_CL = 'Pricing Piso Chile'
SHEET_PE = 'Pricing Piso Peru'
HEADER_ROW = 5   # 1-indexed; data starts at row 6

VERSION = datetime.date.today().strftime('%Y-%m-%d')


def _find_libro_base():
    """Retorna el Libro Base más reciente del inbox."""
    candidates = []
    for pattern in LIBRO_BASE_GLOB:
        candidates += list(INBOX.glob(pattern))
    if not candidates:
        raise FileNotFoundError(
            f"No se encontró Libro Base en {INBOX}. "
            f"Patrones buscados: {LIBRO_BASE_GLOB}"
        )
    return sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def _parse_sheet(wb, sheet_name, pais):
    """
    Lee una hoja de precios piso y retorna lista de dicts canónicos.
    Lógica de selección de precio piso:
      - Preferir 'NUEVO PRECIO PISO PROPUESTO' (col J) si existe y es numérico
      - Fallback a 'PRECIO PISO (CALCULADO)' (col G)
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Hoja '{sheet_name}' no encontrada en el Libro Base.")

    ws = wb[sheet_name]
    headers_raw = list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))[0]
    headers = [str(h).replace('\n', ' ').strip().upper() if h else '' for h in headers_raw]

    def _col(keyword_list):
        for kw in keyword_list:
            for i, h in enumerate(headers):
                if all(k.upper() in h for k in kw.split('|')):
                    return i
        return None

    idx_sku     = _col(['SKU'])
    idx_prod    = _col(['PRODUCTO'])
    idx_fmt     = _col(['FORMATO'])
    idx_costo_pkg = _col(['PRECIO COMPRA|LO MIRANDA']) or _col(['COSTO FÁBRICA|PKG']) or _col(['COSTO|PKG'])
    idx_costo_u = _col(['COSTO|UNIDAD'])
    idx_pp_calc = _col(['PRECIO PISO|CALCULADO'])
    idx_pp_prop = _col(['NUEVO PRECIO|PISO'])
    idx_marg_c  = _col(['MARGEN|CALC'])
    idx_marg_p  = _col(['MARGEN|PROPUESTO'])
    idx_clasif  = _col(['CLASIF'])
    idx_orden   = _col(['ORDEN|FMT'])

    moneda = 'CLP' if pais == 'CL' else 'USD'

    entries = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        sku  = str(row[idx_sku]).strip()  if idx_sku  is not None and row[idx_sku]  else ''
        prod = str(row[idx_prod]).strip() if idx_prod is not None and row[idx_prod] else ''
        fmt  = str(row[idx_fmt]).strip()  if idx_fmt  is not None and row[idx_fmt]  else ''

        if not sku or sku == 'None' or not prod or prod == 'None':
            continue

        def _num(idx):
            if idx is None: return None
            v = row[idx]
            try:
                fv = float(v)
                return fv if fv > 0 else None
            except (TypeError, ValueError):
                return None

        # Precio piso: preferir propuesto, fallback calculado
        pp_prop = _num(idx_pp_prop)
        pp_calc = _num(idx_pp_calc)
        pp      = pp_prop if pp_prop is not None else pp_calc

        costo_pkg = _num(idx_costo_pkg)
        costo_u   = _num(idx_costo_u)
        marg_calc = _num(idx_marg_c)
        marg_prop = _num(idx_marg_p)
        clasif    = str(row[idx_clasif]).strip() if idx_clasif is not None and row[idx_clasif] else ''
        # Quitar emojis de clasificación para texto limpio
        clasif_clean = clasif.replace('🟢', '').replace('🟡', '').replace('🔴', '').strip()
        orden = int(row[idx_orden]) if idx_orden is not None and row[idx_orden] and str(row[idx_orden]).isdigit() else None

        entry = {
            'sku':              sku,
            'producto':         prod,
            'formato':          fmt,
            'pais':             pais,
            'moneda':           moneda,
            'precio_piso':      round(pp, 2) if pp is not None else None,
            'precio_piso_calc': round(pp_calc, 2) if pp_calc is not None else None,
            'precio_piso_prop': round(pp_prop, 2) if pp_prop is not None else None,
            'costo_paquete':    round(costo_pkg, 2) if costo_pkg is not None else None,
            'costo_unitario':   round(costo_u, 4) if costo_u is not None else None,
            'margen_calc':      round(marg_calc, 4) if marg_calc is not None else None,
            'margen_propuesto': round(marg_prop, 4) if marg_prop is not None else None,
            'clasificacion':    clasif_clean,
            'estado':           'ACTIVO',
            'orden_formato':    orden,
        }
        entries.append(entry)

    return entries


def build(libro_base_path=None):
    """
    Construye el master dataset y lo guarda en data/master_prices.json.
    Retorna el dict del dataset para uso programático.
    """
    if libro_base_path is None:
        libro_base_path = _find_libro_base()

    print(f"📖 Leyendo Libro Base: {libro_base_path.name}")
    wb = openpyxl.load_workbook(libro_base_path, data_only=True)

    entries_cl = _parse_sheet(wb, SHEET_CL, 'CL')
    entries_pe = _parse_sheet(wb, SHEET_PE, 'PE')

    all_entries = entries_cl + entries_pe

    # Estadísticas
    cl_con_pp   = sum(1 for e in entries_cl if e['precio_piso'] is not None)
    pe_con_pp   = sum(1 for e in entries_pe if e['precio_piso'] is not None)
    cl_sin_pp   = len(entries_cl) - cl_con_pp
    pe_sin_pp   = len(entries_pe) - pe_con_pp

    master = {
        '_meta': {
            'version':        VERSION,
            'fuente':         libro_base_path.name,
            'fecha_fuente':   datetime.datetime.fromtimestamp(
                libro_base_path.stat().st_mtime
            ).strftime('%Y-%m-%d %H:%M:%S'),
            'generado':       datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'generado_por':   'build_master_dataset.py',
            'total_productos': len(all_entries),
            'total_cl':       len(entries_cl),
            'total_pe':       len(entries_pe),
            'cl_con_precio_piso': cl_con_pp,
            'cl_sin_precio_piso': cl_sin_pp,
            'pe_con_precio_piso': pe_con_pp,
            'pe_sin_precio_piso': pe_sin_pp,
            'ssot': True,
            'nota': (
                'Este archivo es el ÚNICO origen de datos para precios piso, '
                'productos y presentaciones. NO editar manualmente. '
                'Regenerar ejecutando: python3 scripts/build_master_dataset.py'
            ),
        },
        'productos': all_entries,
        # Índices para lookup rápido: clave = "PROD_NORM|FMT_NORM"
        '_index_cl': {
            f"{e['producto'].upper()}|{e['formato'].upper()}": e
            for e in entries_cl
        },
        '_index_pe': {
            f"{e['producto'].upper()}|{e['formato'].upper()}": e
            for e in entries_pe
        },
        '_index_sku': {
            e['sku']: e
            for e in all_entries if e['sku']
        },
    }

    out_path = DATA / 'master_prices.json'
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(master, f, ensure_ascii=False, indent=2)

    print(f"✅ master_prices.json generado: {len(all_entries)} SKUs "
          f"({len(entries_cl)} Chile · {len(entries_pe)} Perú)")
    print(f"   Chile: {cl_con_pp} con PP · {cl_sin_pp} sin PP")
    print(f"   Perú:  {pe_con_pp} con PP · {pe_sin_pp} sin PP")
    print(f"   Guardado: {out_path}")

    return master


if __name__ == '__main__':
    build()
