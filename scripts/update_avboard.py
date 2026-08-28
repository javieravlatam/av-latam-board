#!/usr/bin/env python3
"""
update_avboard.py — Motor de actualización automática AVBOARD
Agroveca Grupo LATAM 2026

Uso:
    python3 scripts/update_avboard.py

Detecta automáticamente los archivos más recientes en /inbox y actualiza:
  - repo/avboard_data.js
  - repo/avboard_clientes.js
  - repo/Panel_IEC_Auditoria_2026.html  (TX_CL)
  - logs/update_log.txt
  - logs/resumen_actualizacion.md
  - logs/alertas.md

Reglas:
  - NO modifica diseño visual de paneles
  - NO requiere intervención del usuario
  - Proceso incremental: detecta y procesa solo archivos nuevos
  - Valida sintaxis JS antes de escribir
"""

import os, sys, re, json, glob, subprocess, textwrap, math, unicodedata
from datetime import datetime, date
from pathlib import Path

import pandas as pd
import numpy as np

# P0 Sprint: catálogo canónico de vendedores (reemplaza rtc_map/vend_map_pe/VEND_ORDER hardcoded)
_scripts_dir = Path(__file__).parent
if str(_scripts_dir) not in sys.path:
    sys.path.insert(0, str(_scripts_dir))
try:
    import vendors_catalog as _vc
    _VC_AVAILABLE = True
except ImportError:
    _VC_AVAILABLE = False
    print("⚠ vendors_catalog.py no encontrado — usando maps LEGACY")

# ── JSON encoder robusto (maneja numpy types) ──────────────────────────────────
class _NpEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (np.integer,)):   return int(obj)
        if isinstance(obj, (np.floating,)):  return float(obj)
        if isinstance(obj, (np.bool_,)):     return bool(obj)
        if isinstance(obj, np.ndarray):      return obj.tolist()
        return super().default(obj)

def _jdumps(obj, **kw):
    return json.dumps(obj, cls=_NpEncoder, ensure_ascii=False, **kw)

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT   = Path(__file__).parent.parent          # /avboard/
INBOX  = ROOT / 'inbox'
REPO = ROOT
LOGS   = REPO / 'logs'
SCRIPTS = ROOT / 'scripts'

AVBOARD_DATA    = REPO / 'avboard_data.js'
AVBOARD_CLI     = REPO / 'avboard_clientes.js'
PANEL_IEC       = REPO / 'Panel_IEC_Auditoria_2026.html'

TODAY = date.today().strftime('%d/%m/%Y')
NOW   = datetime.now().strftime('%Y-%m-%d %H:%M')
CACHE_V = datetime.now().strftime('%Y%m%d%H%M')  # cache-busting ?v= — incluye hora+minuto para forzar re-descarga en cada corrida

# ── Master Data Layer — Tier 0 ────────────────────────────────────────────────
_MD_DIR = Path(__file__).parent.parent / "pipeline" / "master_data"

def _load_fx_rate(moneda_origen: str, moneda_destino: str, fallback: float = 950.0) -> float:
    """Lee tipo de cambio desde fx_rates.json (Master Data canónico).
    Fallback solo si el archivo no existe o no contiene el par activo."""
    _path = _MD_DIR / "fx_rates.json"
    try:
        data = json.loads(_path.read_text())
        for rate in data.get("rates", []):
            if (rate.get("moneda_origen") == moneda_origen
                    and rate.get("moneda_destino") == moneda_destino
                    and rate.get("active", False)):
                val = float(rate["valor"])
                print(f"   💱 TC {moneda_origen}/{moneda_destino} = {val:,.0f} [fuente: fx_rates.json · {rate.get('periodo','?')}]")
                return val
        print(f"   ⚠ TC {moneda_origen}/{moneda_destino} no encontrado en fx_rates.json — usando fallback {fallback}")
    except Exception as _e:
        print(f"   ⚠ fx_rates.json no accesible ({_e}) — usando fallback {fallback}")
    return fallback

def _load_gg_decisions() -> list:
    """Lee decisiones GG activas desde gg_decisions.json (Master Data canónico).
    Retorna lista vacía si el archivo no existe (backward compatible)."""
    _path = _MD_DIR / "gg_decisions.json"
    try:
        data = json.loads(_path.read_text())
        active = [d for d in data.get("decisiones", []) if d.get("active", False)]
        print(f"   📋 GG Decisions cargadas: {len(active)} activas desde gg_decisions.json")
        return active
    except Exception as _e:
        print(f"   ⚠ gg_decisions.json no accesible ({_e}) — sin decisiones GG activas")
        return []

# ── Constantes de negocio desde Master Data ───────────────────────────────────
TC_CLP_USD   = _load_fx_rate("CLP", "USD")   # fuente canónica: pipeline/master_data/fx_rates.json
_GG_DECISIONS = _load_gg_decisions()          # reemplaza GG-001/GG-002 hardcodeados

# Presupuesto Chile y Perú desde Libro Base, con fallback LEGACY
import sys as _sys_pb
import os as _os_pb

_scripts_dir_pb = _os_pb.path.dirname(_os_pb.path.abspath(__file__))
if _scripts_dir_pb not in _sys_pb.path:
    _sys_pb.path.insert(0, _scripts_dir_pb)

try:
    from ppto_libro_base import get_ppto_all
    _ppto = get_ppto_all()
except Exception as _ppto_err:
    print(f"⚠️  No se pudo cargar presupuesto desde Libro Base: {_ppto_err}")
    print("⚠️  Usando presupuesto LEGACY")
    from ppto_libro_base import PPTO_MENSUAL_CL_LEGACY, PPTO_MENSUAL_PE_LEGACY
    _ppto = {
        "chile": {
            "mensual": list(PPTO_MENSUAL_CL_LEGACY),
            "ppto_4m": sum(PPTO_MENSUAL_CL_LEGACY[:4]),
            "ppto_5m": sum(PPTO_MENSUAL_CL_LEGACY[:5]),
            "anual": sum(PPTO_MENSUAL_CL_LEGACY),
            "source": "LEGACY",
            "warning": str(_ppto_err),
        },
        "peru": {
            "mensual": list(PPTO_MENSUAL_PE_LEGACY),
            "ppto_4m": sum(PPTO_MENSUAL_PE_LEGACY[:4]),
            "ppto_5m": sum(PPTO_MENSUAL_PE_LEGACY[:5]),
            "anual": sum(PPTO_MENSUAL_PE_LEGACY),
            "source": "LEGACY",
            "warning": str(_ppto_err),
        },
    }

PPTO_MENSUAL_CL = _ppto["chile"]["mensual"]
PPTO_ANUAL_CL   = _ppto["chile"]["anual"]
PPTO_SOURCE_CL  = _ppto["chile"]["source"]

PPTO_MENSUAL_PE = _ppto["peru"]["mensual"]
PPTO_ANUAL_PE   = _ppto["peru"]["anual"]
PPTO_SOURCE_PE  = _ppto["peru"]["source"]

print(f"  Ppto Chile [{PPTO_SOURCE_CL}]: 5m={sum(PPTO_MENSUAL_CL[:5]):,.0f} / anual={PPTO_ANUAL_CL:,.0f} CLP")
print(f"  Ppto Perú  [{PPTO_SOURCE_PE}]: 5m={sum(PPTO_MENSUAL_PE[:5]):,.1f} / anual={PPTO_ANUAL_PE:,.1f} USD")

# Presupuesto por RTC — DINÁMICO desde Libro Base (get_ppto_rtc).
# Ya no se hardcodean los valores individuales por vendedor: cualquier cambio
# en inbox/nuevo libro base AV 2026.xlsx se refleja en el siguiente proceso.
try:
    from ppto_libro_base import get_ppto_rtc as _get_ppto_rtc
    PPTO_RTC_CL = _get_ppto_rtc("CL")
    PPTO_RTC_MENSUAL_PE = _get_ppto_rtc("PE")
    if not PPTO_RTC_CL or all(sum(v) == 0 for v in PPTO_RTC_CL.values()):
        raise ValueError("get_ppto_rtc('CL') retornó vacío o todos ceros — usando LEGACY")
    if not PPTO_RTC_MENSUAL_PE or all(sum(v) == 0 for v in PPTO_RTC_MENSUAL_PE.values()):
        raise ValueError("get_ppto_rtc('PE') retornó vacío o todos ceros — usando LEGACY")
    print(f"  RTCs Chile desde Libro Base: {sorted(PPTO_RTC_CL.keys())}")
    print(f"  RTCs Perú  desde Libro Base: {sorted(PPTO_RTC_MENSUAL_PE.keys())}")
except Exception as _rtc_err:
    print(f"⚠️  No se pudo cargar RTCs desde Libro Base: {_rtc_err}. Usando LEGACY.")
    PPTO_RTC_CL = {
        'caroca':      [12500000,  6000000, 14500000,  8831000, 12500000,  8730000,  6000000, 25500000, 10800000,  8700000,  5000000, 12500000],
        'laratro':     [36000000, 10600000,  7500000, 16600000, 22500000, 10000000,  7800000, 21000000, 25000000, 20000000, 37500000, 19500000],
        'encina':      [14200000,  7500000,  7500000,  5000000,  5000000,  5000000,  5000000, 12000000, 21000000, 28000000, 31000000, 28700000],
        'velasquez':   [14858000, 11502000,  6500000, 10000000,  5000000, 20000000, 38000000, 12000000, 48000000, 50000000, 18000000, 20000000],
        'veverka':     [ 6000000,  6000000,  6000000,  6000000,  6000000,  6000000,  6000000,  6000000,  6000000,  6000000,  6000000,  6000000],
        'munoz':       [ 6025000,  5310000,  3978000,  3306000,  5661000,  8360000,         0,        0,        0,        0,        0,        0],
        'franco_riffo':[ 1769000,  1490000,  4709000,  3836000,  3065000,  4175000,         0,        0,        0,        0,        0,        0],
    }
    PPTO_RTC_MENSUAL_PE = {
        'infante':    [16260, 30164,     0, 67708, 37358, 26733, 15346,      0,      0,      0,      0,      0],
        'aguirre':    [12025, 10572, 10000, 16149, 13128,  7644, 25021,  30000, 100000,  70000,  90000,  40000],
        'atalaya':    [22123, 17722, 10139, 17028, 21306, 34049, 25000,  19000,  23000,  23000,  18000,  10000],
        'gonzales':   [ 1261,  1469,  2498,  1820,  1521,  2431,  8000,      0,   5000,      0,   5000,      0],
        'valladares': [    0,   221,  5310,  4823,  5153, 10319, 10000,  10000,  10000,  15000,  10000,  10000],
        'diaz':       [    0,     0,     0,     0,     0, 22300, 15000,  15000,  30000,  30000,  20000,  35000],
        'martha':     [    0,     0,     0,     0,     0,     0,     0,  10000,  10000,  15000,  15000,  15000],
    }

PPTO_RTC_ANUAL_PE = {k: sum(v) for k, v in PPTO_RTC_MENSUAL_PE.items()}

# NOTA: la antigua tabla estática RENTABILIDAD (alertas hardcodeadas) fue
# reemplazada 2026-06-24 por compute_productos(), que calcula rentabilidad
# real por país×producto×formato desde precios piso + ventas. Ver sección
# "MÓDULO PRODUCTOS" más abajo.

MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
MESES_FULL = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
              'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']


# ══════════════════════════════════════════════════════════════════════════════
#  1. DETECCIÓN DE ARCHIVOS INBOX
# ══════════════════════════════════════════════════════════════════════════════

def _parse_date_from_name(path):
    """Extrae objeto date del nombre del archivo para ordenamiento correcto."""
    name = path.stem
    # DD-MM-YYYY o DD.MM.YYYY
    m = re.search(r'(\d{1,2})[-.](\d{2})[-.](\d{4})', name)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    # DD-MM sin año
    m2 = re.search(r'(\d{1,2})[-](\d{2})(?!\d)', name)
    if m2:
        try:
            return date(2026, int(m2.group(2)), int(m2.group(1)))
        except ValueError:
            pass
    return date(2000, 1, 1)


def detect_inbox_files():
    """
    [P0 Sprint] Clasificador unificado: delega al inbox_detector (SSOT) y lee raw_registry.json.
    Ya no hace globs independientes — elimina la duplicación de lógica (V-02).
    """
    return load_files_from_registry()


def load_files_from_registry() -> dict:
    """
    Lee pipeline/raw_registry.json (generado por inbox_detector.py) y traduce
    los vigentes a las claves que el pipeline espera internamente.

    Tabla de traducción (registry_key → pipeline_key):
      VENTAS_CL|AGROCOMERCIAL_CL      → chile_ventas
      VENTAS_PE|AGROVECA_PE            → peru_ventas
      CXC_CL_AGROCOMERCIAL|AGROCOMERCIAL_CL → cxc_agro
      CXC_CL_AGROVECA|AGROVECA_CL     → cxc_avch
      LIBRO_BASE|GRUPO_AV_LATAM        → libro_base + piso_chile + piso_peru

    Si raw_registry.json no existe o es inválido: aborta con mensaje claro.
    """
    REGISTRY = REPO / "pipeline" / "raw_registry.json"
    if not REGISTRY.exists():
        print("❌ pipeline/raw_registry.json no encontrado.")
        print("   Ejecuta primero: python3 scripts/inbox_detector.py")
        sys.exit(1)

    with open(REGISTRY, encoding="utf-8") as f:
        reg = json.load(f)

    vigentes = reg.get("vigentes", {})
    if not vigentes:
        print("❌ raw_registry.json no tiene vigentes. Revisa el inbox.")
        sys.exit(1)

    # Tabla de traducción: clave del registry → clave(s) del pipeline
    REGISTRY_TO_PIPELINE = {
        "VENTAS_CL|AGROCOMERCIAL_CL":             ["chile_ventas"],
        "VENTAS_PE|AGROVECA_PE":                   ["peru_ventas"],
        "CXC_CL_AGROCOMERCIAL|AGROCOMERCIAL_CL":  ["cxc_agro"],
        "CXC_CL_AGROVECA|AGROVECA_CL":            ["cxc_avch"],
        "CXC_PE|AGROVECA_PE":                     ["cxc_pe"],
        "LIBRO_BASE|GRUPO_AV_LATAM":               ["libro_base", "piso_chile", "piso_peru"],
    }

    files = {}
    for reg_key, pipeline_keys in REGISTRY_TO_PIPELINE.items():
        filename = vigentes.get(reg_key)
        if filename:
            path = INBOX / filename
            if path.exists():
                for pk in pipeline_keys:
                    files[pk] = path
                if "libro_base" in pipeline_keys:
                    print(f"   📗 SSOT Libro Base: {filename}")
            else:
                print(f"   ⚠ Vigente registrado pero archivo no existe en inbox: {filename}")

    # Reporte de lo que se cargó vs lo que faltó
    for reg_key in REGISTRY_TO_PIPELINE:
        if reg_key not in vigentes:
            tipo = reg_key.split("|")[0]
            print(f"   ℹ No hay vigente para {tipo} en registry")

    return files


def extract_date_from_filename(path):
    """Extrae fecha de corte (DD/MM/YYYY) desde el nombre del archivo."""
    d = _parse_date_from_name(path)
    if d.year > 2000:
        return d.strftime('%d/%m/%Y')
    return TODAY


# ══════════════════════════════════════════════════════════════════════════════
#  2. EXTRACCIÓN DE DATOS
# ══════════════════════════════════════════════════════════════════════════════

def extract_chile_ventas(path):
    """Extrae ventas Chile del Libro de Ventas.
    Retorna dict con ytd, mensual, por_rtc, df_tx (DataFrame).
    """
    df = pd.read_excel(path, sheet_name='VENTAS', header=1)
    df.columns = [str(c).strip() for c in df.columns]
    df['fecha_dt'] = pd.to_datetime(df['Fecha'], format='%d/%m/%Y', errors='coerce')
    df['total_n']  = pd.to_numeric(df['Total'], errors='coerce').fillna(0)
    df['pv_n']     = pd.to_numeric(df['Precio Uni'], errors='coerce')

    # Mensual por mes
    mensual = {}
    for i, mes in enumerate(MESES_FULL):
        v = df[df['MES'] == mes]['total_n'].sum()
        mensual[mes] = int(v)

    # Por RTC mensual (nombre → clave) — cargado desde catálogo canónico (V-03)
    if _VC_AVAILABLE:
        rtc_map, _src = _vc.get_rtc_map_safe("CL")
        if _src == "legacy":
            print("   ⚠ extract_chile_ventas: rtc_map cargado desde LEGACY")
    else:
        rtc_map = {
            'PABLO LARATRO': 'laratro', 'FRANCISCO VELASQUEZ': 'velasquez',
            'JORGE CAROCA': 'caroca', 'RODRIGO ENCINA': 'encina',
            'IVAN VEVERKA': 'veverka', 'VALENTINA MUÑOZ': 'munoz',
            'RAYEN BERNAZAR': 'bernazar', 'JAVIER ALMEIDA': 'almeida',
        }
    # Normalizar a UPPER para comparación robusta
    rtc_map = {k.upper(): v for k, v in rtc_map.items()}

    # Normalizar columna Vendedor a UPPER para comparación robusta con rtc_map
    df['_vend_upper'] = df['Vendedor'].astype(str).str.strip().str.upper()

    rtc_mensual = {}
    for vend_upper, key in rtc_map.items():
        meses_vals = []
        for mes in MESES_FULL:
            v = df[(df['MES'] == mes) & (df['_vend_upper'] == vend_upper)]['total_n'].sum()
            meses_vals.append(int(v))
        if any(v > 0 for v in meses_vals):
            rtc_mensual[key] = meses_vals

    # VENDEDOR_NO_HOMOLOGADO: detectar vendedores en Excel que no están en el catálogo (V-03)
    _all_vendors_cl = df['_vend_upper'].dropna().unique()
    _known_upper = set(rtc_map.keys())
    _unknown_cl = [v for v in _all_vendors_cl if v and v not in _known_upper
                   and v not in ('NAN', '', 'TOTAL', 'TOTAL GENERAL')]
    if _unknown_cl:
        _totales_desconocidos = {
            v: int(df[df['_vend_upper'] == v]['total_n'].sum())
            for v in _unknown_cl
        }
        print(f"   ⚠ VENDEDOR_NO_HOMOLOGADO CL: {_unknown_cl}")
        print(f"     Montos: {_totales_desconocidos}")
        print(f"     Agregar aliases en pipeline/vendors.json si son vendedores reales.")

    # YTD (sum of all months with data)
    months_with_data = [m for m in MESES_FULL if mensual.get(m, 0) > 0]
    ytd_5m = sum(mensual.get(m, 0) for m in MESES_FULL)
    n_months = len(months_with_data)
    ytd_4m = sum(mensual.get(MESES_FULL[i], 0) for i in range(min(4, n_months)))
    # Último mes con datos reales (el nombre 'mayo_parcial' se mantiene por
    # compatibilidad con los paneles, pero apunta dinámicamente al mes más
    # reciente con ventas — antes quedaba fijo en 'MAYO' aunque el corte avanzara)
    ultimo_mes_idx = n_months - 1 if n_months > 0 else 4
    mayo_parcial_dinamico = mensual.get(MESES_FULL[ultimo_mes_idx], 0) if n_months > 0 else mensual.get('MAYO', 0)

    # T1 por RTC (Jan-Mar)
    rtc_t1 = {}
    for vend, key in rtc_map.items():
        t1 = sum(df[(df['MES'] == MESES_FULL[i]) & (df['Vendedor'] == vend)]['total_n'].sum() for i in range(3))
        if t1 > 0:
            rtc_t1[key] = int(t1)

    # Ppto cumplimiento
    n_ppto = min(4, n_months)
    ppto_4m = sum(PPTO_MENSUAL_CL[:4])
    cumpl_4m = round(ytd_4m / ppto_4m, 4) if ppto_4m > 0 else 0
    # ppto_5m / cumplimiento_5m: ANTES sumaba siempre los primeros 5 meses de
    # presupuesto aunque ytd_5m (numerador) ya sumara más meses reales (ej. al
    # llegar Junio) — esto inflaba el % de cumplimiento mostrado. Ahora el
    # presupuesto usado como denominador sigue la misma ventana dinámica
    # (n_months) que el numerador, para que ambos midan el mismo período.
    ppto_5m = sum(PPTO_MENSUAL_CL[:n_months]) if n_months > 0 else sum(PPTO_MENSUAL_CL[:5])
    cumpl_5m = round(ytd_5m / ppto_5m, 4) if ppto_5m > 0 else 0

    return {
        'ytd_5m':       ytd_5m,
        'ytd_4m':       ytd_4m,
        'mayo_parcial': mayo_parcial_dinamico,
        'ppto_4m':      int(ppto_4m),
        'ppto_5m':      int(ppto_5m),
        'cumplimiento_4m': cumpl_4m,
        'cumplimiento_5m': cumpl_5m,
        'mensual':      [mensual.get(m, 0) for m in MESES_FULL],
        'rtc_t1':       rtc_t1,
        'rtc_mensual':  rtc_mensual,
        'n_months':     n_months,
        'df':           df,
        'rtc_map':      rtc_map,
    }


def extract_peru_ventas(path):
    """Extrae ventas Perú."""
    df_raw = pd.read_excel(path, sheet_name='RESUMEN', header=None)

    # Pivot table: rows=vendedor, cols=mes
    # Find header row: must have VENDEDOR + at least one month column
    header_row = None
    for i, row in df_raw.iterrows():
        vals = [str(v).upper() for v in row if pd.notna(v)]
        has_vend = any('VENDEDOR' in v for v in vals)
        has_mes  = any('ENERO' in v or 'FEBRERO' in v or 'MARZO' in v for v in vals)
        if has_vend and has_mes:
            header_row = i
            break

    if header_row is None:
        raise ValueError(f"No se encontró fila de encabezado en {path}")

    df = pd.read_excel(path, sheet_name='RESUMEN', header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    # Find rows with vendor data — mapa cargado desde catálogo canónico (V-03)
    if _VC_AVAILABLE:
        vend_map_pe, _src_pe = _vc.get_rtc_map_safe("PE")
        if _src_pe == "legacy":
            print("   ⚠ extract_peru_ventas: vend_map_pe cargado desde LEGACY")
    else:
        vend_map_pe = {
            'OSCAR INFANTE': 'infante', 'NICOLL NAVARRO': 'navarro',
            'OMAR ATALAYA': 'atalaya', 'ANTONIO GONZALES': 'gonzales',
            'LISBETH AGUIRRE': 'aguirre', 'LIZBETH AGUIRRE': 'aguirre',
            'PATRICIA VALLADARES': 'valladares', 'SUSAN DIAZ': 'diaz',
            'SUSAN DÍAZ': 'diaz', 'MARTHA HIDALGO': 'martha',
            'MARTHA HIDALGO - KAM': 'martha',
        }
    # Normalizar a UPPER (catálogo ya los entrega en UPPER, pero doble-check)
    vend_map_pe = {k.upper(): v for k, v in vend_map_pe.items()}

    # Detectar dinámicamente TODAS las columnas de mes presentes en el archivo
    # (antes esto estaba fijo a 5 meses Ene-May vía mes_cols_pe/range(5) — cuando
    #  el corte avanzó a Junio, "Total general" ya incluía Junio pero el desglose
    #  mensual_pe no, y quedaba desincronizado del ytd_5m real. Ver MESES_FULL global.)
    available_cols = {}
    for col in df.columns:
        col_up = str(col).upper()
        for i, m_name in enumerate(MESES_FULL):
            if m_name in col_up:
                available_cols[i] = col
                break

    # Mismo principio que arriba: el índice del "mes actual/parcial" (antes
    # fijo en 4 = Mayo) ahora sigue dinámicamente al último mes con columna
    # detectada en el Excel, para que 'mayo'/'mayo_parcial' avancen solos
    # cuando el corte avanza (ej. a Junio) sin quedar pegados en Mayo.
    n_months_pe  = len(available_cols)
    last_idx_pe  = max(available_cols.keys()) if available_cols else 4

    por_vendedor = {}
    rtc_mensual_pe = {}
    mensual_pe = [0] * 12

    for _, row in df.iterrows():
        vend_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
        vend_name = str(row.get(vend_col, row.iloc[1] if len(row) > 1 else '')).strip().upper()

        matched_key = None
        for vname, key in vend_map_pe.items():
            if vname.upper() in vend_name or vend_name in vname.upper():
                matched_key = key
                break

        if matched_key is None:
            # VENDEDOR_NO_HOMOLOGADO PE: nombre con ventas no está en el catálogo
            _SKIP_NAMES = {'', 'NAN', 'NONE', 'TOTAL', 'TOTAL GENERAL', 'GRAN TOTAL', 'SUBTOTAL'}
            if vend_name and vend_name not in _SKIP_NAMES and not vend_name.startswith('TOTAL'):
                # Verificar si tiene al menos un mes con valor > 0
                _has_data = any(
                    available_cols.get(i) and pd.to_numeric(
                        row.get(available_cols[i], 0), errors='coerce') > 0
                    for i in range(12)
                )
                if _has_data:
                    print(f"   ⚠ VENDEDOR_NO_HOMOLOGADO PE: '{vend_name}' "
                          f"— agregar alias en pipeline/vendors.json")
            continue

        monthly = []
        for i in range(12):
            col = available_cols.get(i)
            if col and col in row.index:
                val = pd.to_numeric(row[col], errors='coerce')
                monthly.append(round(float(val), 1) if pd.notna(val) else 0)
            else:
                monthly.append(0)

        # Total col
        total_col = [c for c in df.columns if 'total' in str(c).lower() or 'general' in str(c).lower()]
        ytd = round(float(pd.to_numeric(row[total_col[0]], errors='coerce')), 1) if total_col else sum(monthly)

        por_vendedor[matched_key] = {
            'nombre': vend_name.title(),
            'ytd': round(ytd),
            'mayo': round(monthly[last_idx_pe]) if last_idx_pe < len(monthly) else 0,
        }
        rtc_mensual_pe[matched_key] = [round(v) for v in monthly]

        for i, v in enumerate(monthly):
            mensual_pe[i] += round(v)

    ytd_5m_pe = sum(round(v['ytd']) for v in por_vendedor.values())
    ytd_4m_pe = sum(mensual_pe[:4])

    ppto_4m_pe = sum(PPTO_MENSUAL_PE[:4])
    cumpl_4m_pe = round(ytd_4m_pe / ppto_4m_pe, 4) if ppto_4m_pe > 0 else 0
    # ppto_5m_pe: igual que en Chile, el denominador ahora sigue la misma
    # ventana dinámica (n_months_pe) que ytd_5m_pe — antes quedaba fijo en
    # 5 meses de presupuesto aunque ya hubiera 6+ meses de venta real,
    # inflando el % de cumplimiento.
    ppto_5m_pe = sum(PPTO_MENSUAL_PE[:n_months_pe]) if n_months_pe > 0 else sum(PPTO_MENSUAL_PE[:5])
    cumpl_5m_pe = round(ytd_5m_pe / ppto_5m_pe, 4) if ppto_5m_pe > 0 else 0

    return {
        'ytd_5m':          ytd_5m_pe,
        'ytd_4m':          ytd_4m_pe,
        'mayo_parcial':    mensual_pe[last_idx_pe] if n_months_pe > 0 else mensual_pe[4],
        'ppto_4m':         int(ppto_4m_pe),
        'ppto_5m':         int(ppto_5m_pe),
        'ppto_mes_label':  MESES[n_months_pe - 1] if n_months_pe > 0 else MESES[4],
        'cumplimiento_4m': cumpl_4m_pe,
        'cumplimiento_5m': cumpl_5m_pe,
        'mensual':         mensual_pe,
        'por_vendedor':    por_vendedor,
        'rtc_mensual':     rtc_mensual_pe,
    }


def _apply_peru_vendor_gg_decisions(pe_data):
    """
    Aplica decisiones de Gerencia General sobre atribución de vendedores Perú.
    Se ejecuta DESPUÉS de extract_peru_ventas() como capa de normalización.
    NO modifica la fuente original.

    Interpreta y aplica decisiones activas desde gg_decisions.json (Master Data Tier 0).
    Se ejecuta DESPUÉS de extract_peru_ventas() como capa de normalización.
    NO modifica la fuente original.

    Tipos soportados:
      - reasignacion_transaccion: mueve monto entre vendedores en mes específico (nivel agregado)
      - fusion_vendedor: fusiona source_vendor en target_vendor (suma series y elimina source)

    Para agregar GG-003+: editar gg_decisions.json únicamente — cero cambios en Python.
    """
    pv = pe_data.get('por_vendedor', {})
    rm = pe_data.get('rtc_mensual',  {})
    ms = pe_data.get('mensual',      [0] * 12)

    for dec in _GG_DECISIONS:
        dec_id  = dec.get('id', '?')
        tipo    = dec.get('tipo', '')
        criterio = dec.get('criterio', {})
        accion   = dec.get('accion', {})

        if tipo == 'reasignacion_transaccion':
            # ── Nivel agregado: mover monto_usd de source.mes_idx → target.mes_idx ──
            src = accion.get('source_vendor_id')
            tgt = accion.get('vendor_id_override')
            mes_idx = int(accion.get('mes_idx', 0))
            monto   = float(accion.get('monto_usd', 0))
            if src and tgt and src in pv and tgt in pv:
                src_m = list(rm.get(src, [0] * 12))
                tgt_m = list(rm.get(tgt, [0] * 12))
                adj = min(monto, max(0.0, src_m[mes_idx]))
                if adj > 0:
                    src_m[mes_idx] = round(src_m[mes_idx] - adj)
                    tgt_m[mes_idx] = round(tgt_m[mes_idx] + adj)
                    rm[src] = src_m
                    rm[tgt] = tgt_m
                    pv[src]['ytd'] = max(0, round(pv[src].get('ytd', 0) - adj))
                    pv[tgt]['ytd'] = round(pv[tgt].get('ytd', 0) + adj)
                    print(f"   {dec_id} ✓ USD {adj:,} {src}.mes[{mes_idx}] → {tgt}.mes[{mes_idx}]"
                          f" | {tgt}.ytd={pv[tgt]['ytd']:,}")
                else:
                    print(f"   {dec_id} ⚠ {src}.mes[{mes_idx}]={src_m[mes_idx]} — sin ajuste (adj=0)")
            elif src and src not in pv:
                print(f"   {dec_id} ℹ vendor '{src}' no encontrado en por_vendedor — sin ajuste")

        elif tipo == 'fusion_vendedor':
            # ── Fusionar source → target: suma series, elimina source ──────────
            src   = criterio.get('vendor_id')
            tgt   = accion.get('fused_into')
            nombre_display = accion.get('nombre_display', tgt)
            if src and tgt and src in pv:
                # Guard: si tgt no tiene datos activos, verificar que sea un vendor_id
                # registrado en el catálogo canónico — previene creación de vendedores fantasma
                if tgt not in pv:
                    _vcat_path = ROOT / "pipeline" / "vendors.json"
                    _known_ids: set = set()
                    try:
                        _vcat = json.loads(_vcat_path.read_text())
                        _known_ids = {v['vendor_id'] for v in _vcat.get('vendors', [])}
                    except Exception:
                        pass
                    if _known_ids and tgt not in _known_ids:
                        print(f"   {dec_id} ⚠ SKIP — vendor_to '{tgt}' no está en catálogo canónico"
                              f" — decisión ignorada (guard seguridad)")
                        continue
                nav_ytd = pv[src].get('ytd', 0)
                nav_m   = list(rm.get(src, [0] * 12))
                if tgt not in pv:
                    pv[tgt] = {'nombre': nombre_display, 'ytd': 0}
                    rm[tgt] = [0] * 12
                tgt_m = list(rm.get(tgt, [0] * 12))
                rm[tgt] = [round(tgt_m[i] + nav_m[i]) for i in range(12)]
                pv[tgt]['ytd']    = round(pv[tgt].get('ytd', 0) + nav_ytd)
                pv[tgt]['nombre'] = nombre_display
                del pv[src]
                if src in rm:
                    del rm[src]
                # recalcular mensual consolidado
                ms_new = [0] * 12
                for arr in rm.values():
                    for i, v in enumerate(arr):
                        ms_new[i] += round(v)
                ms = ms_new
                print(f"   {dec_id} ✓ {src} fusionado en {tgt} | {tgt}.ytd={pv[tgt]['ytd']:,}")
            elif src and src not in pv:
                print(f"   {dec_id} ℹ vendor '{src}' no encontrado — fusión no aplicable")

        else:
            # tipo desconocido — warning explícito, nunca silencioso
            print(f"   {dec_id} ⚠ tipo desconocido '{tipo}' — decisión ignorada")

    pe_data['por_vendedor'] = pv
    pe_data['rtc_mensual']  = rm
    pe_data['mensual']      = ms
    pe_data['ytd_5m'] = sum(v.get('ytd', 0) for v in pv.values())
    return pe_data


def extract_cxc_chile(agro_path, avch_path):
    """Extrae y consolida CxC Chile de ambas entidades."""

    def read_mora(path, sheet='Detalle Mora'):
        df = pd.read_excel(path, sheet_name=sheet, header=1)
        df.columns = [str(c).strip() for c in df.columns]
        df = df[df['Rut'].notna()].copy()
        df['Total Doc'] = pd.to_numeric(df['Total Doc'], errors='coerce').fillna(0)
        df['Días Mora'] = pd.to_numeric(df['Días Mora'], errors='coerce').fillna(0)
        return df[df['Total Doc'] != 0]

    df_agro = read_mora(agro_path)
    df_avch = read_mora(avch_path)

    def tramos(df):
        return {
            't90':   int(df[df['Tramo'] == '+90']['Total Doc'].sum()),
            't6190': int(df[df['Tramo'] == '61-90']['Total Doc'].sum()),
            't3160': int(df[df['Tramo'] == '31-60']['Total Doc'].sum()),
            't030':  int(df[df['Tramo'] == '0-30']['Total Doc'].sum()),
        }

    tr_agro = tramos(df_agro)
    tr_avch = tramos(df_avch)

    total_agro = int(df_agro['Total Doc'].sum())
    total_avch = int(df_avch['Total Doc'].sum())
    total      = total_agro + total_avch

    tr_comb = {
        't90':   tr_agro['t90']   + tr_avch['t90'],
        't6190': tr_agro['t6190'] + tr_avch['t6190'],
        't3160': tr_agro['t3160'] + tr_avch['t3160'],
        't030':  tr_agro['t030']  + tr_avch['t030'],
    }

    # Por RTC
    rtc_name_map = {
        'FRANCISCO VELASQUEZ': 'velasquez',
        'PABLO LARATRO': 'laratro',
        'RODRIGO ENCINA': 'encina',
        'JORGE CAROCA': 'caroca',
        'IVAN VEVERKA': 'veverka',
        'IVÁN VEVERKA': 'veverka',
        'VALENTINA MUÑOZ': 'munoz',
        'JUAN PABLO NEIRA': 'neira',
        'FRANCO RIFFO': 'franco_riffo',
        'RAYEN BERNAZAR': 'otros',
        'EN TERRENO 1': 'otros',
        'OFICINA': 'otros',
        'JAVIER ALMEIDA': 'almeida',
    }

    df_all = pd.concat([
        df_agro.assign(entidad='agro'),
        df_avch.assign(entidad='avch')
    ])

    def normalize_vend(name):
        n = str(name).strip().upper()
        return rtc_name_map.get(n, 'otros')

    df_all['rtc_key'] = df_all['Vendedor'].apply(normalize_vend)

    por_rtc = {}
    for key in df_all['rtc_key'].unique():
        sub = df_all[df_all['rtc_key'] == key]
        t_total  = int(sub['Total Doc'].sum())
        t_t90    = int(sub[sub['Tramo'] == '+90']['Total Doc'].sum())
        t_venc   = int(sub[sub['Días Mora'] > 0]['Total Doc'].sum())
        por_rtc[key] = {
            'total':   t_total,
            'pct':     round(t_total / total, 4) if total > 0 else 0,
            'vencida': t_venc,
            't90':     t_t90,
            'riesgo':  'CRÍTICO' if t_t90 > 0 else ('RIESGO' if t_venc > t_total * 0.3 else 'NORMAL'),
        }

    # Cuentas críticas (+90d)
    crit = df_all[df_all['Tramo'] == '+90'].copy()
    crit_by_client = crit.groupby(['Razón Social']).agg(
        monto=('Total Doc', 'sum'),
        max_mora=('Días Mora', 'max'),
        rtcs=('Vendedor', lambda x: ' / '.join(sorted(set(str(v) for v in x)))),
    ).reset_index().sort_values('monto', ascending=False)

    cuentas_criticas = []
    for _, row in crit_by_client.iterrows():
        cuentas_criticas.append({
            'cliente': str(row['Razón Social']).strip(),
            'rtc':     str(row['rtcs']),
            'dias':    int(row['max_mora']),
            'monto':   int(row['monto']),
            'estado':  'CRÍTICO',
            'alerta':  'PRIORIDAD_MAXIMA' if row['monto'] > 5000000 else 'URGENTE',
        })

    # Agregar clientes resueltos conocidos
    cuentas_criticas.append({
        'cliente': 'GOYASERVICE SPA',
        'rtc': 'Jorge Caroca',
        'dias': None,
        'monto': 4194750,
        'estado': 'RESUELTO',
        'alerta': None,
        'nota': 'PAGADO ✅ entre 17/04 y 29/04',
    })

    # Per-client CxC dict for avboard_clientes
    cxc_por_cliente = {}
    for _, row in df_all.iterrows():
        rut = str(row.get('Rut', '')).strip()
        if not rut or rut == 'nan':
            continue
        if rut not in cxc_por_cliente:
            cxc_por_cliente[rut] = {
                'saldo': 0, 'max_mora': 0, 'tramo': '0-30', 'estado': 'Normal'
            }
        cxc_por_cliente[rut]['saldo'] += int(row['Total Doc'])
        mora = int(row['Días Mora'])
        if mora > cxc_por_cliente[rut]['max_mora']:
            cxc_por_cliente[rut]['max_mora'] = mora
            if mora > 90:
                cxc_por_cliente[rut]['tramo'] = '+90'
                cxc_por_cliente[rut]['estado'] = 'Crítico'
            elif mora > 60:
                cxc_por_cliente[rut]['tramo'] = '61-90'
                cxc_por_cliente[rut]['estado'] = 'Riesgo'
            elif mora > 30:
                cxc_por_cliente[rut]['tramo'] = '31-60'
                cxc_por_cliente[rut]['estado'] = 'Alerta'

    return {
        'total':       total,
        'vencida':     tr_comb['t90'] + tr_comb['t6190'] + tr_comb['t3160'],
        'al_dia':      tr_comb['t030'],
        'tramos':      tr_comb,
        'tramos_pct':  {k: round(v/total, 4) if total else 0 for k, v in tr_comb.items()},
        'por_entidad': {
            'agrocomercial': {'nombre': 'Agrocomercial', 'total': total_agro, 'tramos': tr_agro},
            'agroveca_chile': {'nombre': 'Agroveca Chile', 'total': total_avch, 'tramos': tr_avch},
        },
        'por_rtc':           por_rtc,
        'cuentas_criticas':  cuentas_criticas,
        'cxc_por_cliente':   cxc_por_cliente,
        'df_all':            df_all,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  3. CÁLCULO IEC
# ══════════════════════════════════════════════════════════════════════════════

# ──────────────────────────────────────────────────────────────────────────────
#  3.0 NORMALIZACIÓN DE NOMBRES Y FORMATOS (compartida Chile + Perú)
# ──────────────────────────────────────────────────────────────────────────────
#  Los nombres de producto en ventas (Chile y Perú) llegan con variantes de
#  marca (VECA/AV), tildes, separadores (' - ', ' X ', sin separador), typos
#  y ruido de empaque/lote. Esta sección centraliza esa limpieza para que
#  TODO el sistema (IEC, TX_CL, módulo Productos) use la misma lógica de
#  matching contra las tablas de precio piso — evita que un mismo SKU se
#  trate distinto en dos partes del pipeline.

NAME_NORMALIZE = {
    # Aliases originales (Panel_IEC, desde antes de este módulo)
    'VECA MOVE': 'AV MOVE', 'VECA ROOT MAX': 'AV ROOT MAX', 'VECASIL FORTE': 'AV SILFORTE',
    'VECA HUMIC ROOT': 'AV HUMIC ROOT', 'CYTOPRIME': 'AV CYTO PRIME', 'CYTO PRIME': 'AV CYTO PRIME',
    'VECA PLUS POTASIO': 'AV PLUS POTASIO', 'VECA PLUS MAGNESIO': 'AV PLUS MAGNESIO',
    'VECA MICROMIX': 'AV MICROMIX', 'VECA PLUS CALCIO': 'AV PLUS CALCIO',
    'VECA PLUS ZINC': 'AV PLUS ZINC', 'VECA PLUS FOSFORO': 'AV PLUS FOSFORO',
    # Aliases adicionales — validados contra Libro de Ventas 21-06-2026 +
    # AGROVECA PERU VENTAS 22.06.2026 vs precios piso CL/PE (módulo Productos, 2026-06-24)
    'ALGAP': 'AV ALGAP 30', 'ALGAP 30': 'AV ALGAP 30',
    'AV MAX FULVIC': 'AV MAX FULVIC 45%', 'VECA MAX FULVIC': 'AV MAX FULVIC 45%',
    'VECA FULVIC': 'AV MAX FULVIC 45%',
    'AV AMINSUGAR': 'AV AMIN SUGAR',
    'AV BIOVECA FOLIAR': 'BIOVECA FOLIAR', 'AV BIOVECA RAIZ': 'BIOVECA RAIZ',
    'AV BIOVECA RAÍZ': 'BIOVECA RAIZ', 'BIOVECA RAÍZ': 'BIOVECA RAIZ',
    'AV BIOVECA NEMA OFF': 'BIOVECA NEMA OFF', 'AV BIOVECA NEMAOFF': 'BIOVECA NEMA OFF',
    'BIOVECA NEMAOFF': 'BIOVECA NEMA OFF', 'AV BIOVECA PRADERAS': 'BIOVECA PRADERAS',
    'AV BIOVECA INVERNAL': 'BIOVECA INVERNAL',
    'AV PLUS NP MIX': 'AV PLUS NP-MIX',
    'VECA CALCIO BORO': 'AV PLUS CALCIO BORO', 'CALCIO BORO': 'AV PLUS CALCIO BORO',
    'VECA PLUC HIERRO': 'AV PLUS HIERRO',
    'VECA BLANCE': 'AV BALANCE',
    'AV MAX GREEN GUARDIAN': 'GREEN GUARDIAN MAX',
    # Específicos de Perú (CONCEPTO de ventas sin marca o con typos)
    'VECASIL': 'AV SILFORTE', 'AV SILIFORTE': 'AV SILFORTE', 'VECA SILFORTE': 'AV SILFORTE',
    'VECA ZINC': 'AV PLUS ZINC',
    'VECA MANGANESO': 'AV PLUS ZINC MANGANESO', 'VECA PLUS MANGANESO': 'AV PLUS ZINC MANGANESO',
    'HUMIC ROOT': 'AV HUMIC ROOT', 'VECA POTASIO': 'AV PLUS POTASIO',
    'BIOSOLARIS': 'AV BIOSOLARIS', 'MAGNESIO': 'AV PLUS MAGNESIO',
    'AV MAGNESIO': 'AV PLUS MAGNESIO', 'VECA MAGNESIO': 'AV PLUS MAGNESIO',
    # Aliases adicionales — detectados al construir compute_productos() contra
    # Libro de Ventas 21-06-2026 (Chile), validados 1:1 contra precios piso CHile:
    # mismo producto, nombre con "PLUS"/orden de palabras distinto al de piso.
    'AV PLUS MANGANESO': 'AV PLUS ZINC MANGANESO',
    'VECA MICROMIX': 'AV PLUS MICRO MIX', 'AV MICROMIX': 'AV PLUS MICRO MIX',
    'AV NUTRI MIX': 'AV PLUS NUTRI MIX', 'NUTRI MIX': 'AV PLUS NUTRI MIX',
    'VECA NUTRI MIX': 'AV PLUS NUTRI MIX',
    'AV PLUS AMIN': 'AV AMIN',
    'PK DEFENDER MAX': 'PK-DEFEND MAX',
    'AV PLUS ROOT MAX': 'AV ROOT MAX',
    'AV CALCIO': 'AV PLUS CALCIO', 'AV BORO': 'AV PLUS BORO',
    'AV SOLARIS': 'AV BIOSOLARIS', 'BIORAIZ': 'BIOVECA RAIZ', 'RAIZ': 'BIOVECA RAIZ',
    'BIOFOLIAR': 'BIOVECA FOLIAR', 'BALANCE': 'AV BALANCE',
}

# Typos puntuales detectados en CONCEPTO de ventas Perú — se corrigen ANTES
# de pasar por NAME_NORMALIZE (sustitución de substring, no es alias de marca).
TYPO_FIX_PE = {
    'MANGNESIO': 'MAGNESIO',     # falta una 'A'
    'SILIFORTE': 'SILFORTE',     # letra de más
    'NUTRIMIX':  'NUTRI MIX',    # falta espacio (AV PLUS NUTRIMIX → AV PLUS NUTRI MIX)
}

_UNIT_MAP = {
    'L': 'L', 'LT': 'L', 'LTS': 'L', 'LTRS': 'L', 'LITRO': 'L', 'LITROS': 'L',
    'GR': 'GR', 'GRS': 'GR', 'G': 'GR', 'GRAMO': 'GR', 'GRAMOS': 'GR',
    'KG': 'KG', 'KGS': 'KG', 'KILO': 'KG', 'KILOS': 'KG',
    'ML': 'ML', 'CC': 'ML',
}


def strip_accents(s):
    """Quita tildes/diacríticos (RAÍZ -> RAIZ) — solo para claves de matching,
    nunca para texto que se muestra al usuario."""
    return ''.join(c for c in unicodedata.normalize('NFD', str(s)) if unicodedata.category(c) != 'Mn')


def normalize_fmt(raw):
    """Canonicaliza un formato/unidad: '20LT'/'20 LTS' -> '20 L', '250GRS' -> '250 GR'."""
    s = str(raw).strip().upper().replace(',', '.')
    s = re.sub(r'\(.*?\)', '', s).strip()  # quita anotaciones tipo '(VECA)'
    if not s or s == 'NAN':
        return ''
    m = re.match(r'^(\d+(?:\.\d+)?)\s*([A-ZÀ-Ü]+)\.?$', s)
    if not m:
        return s
    num, unit = m.groups()
    if num.endswith('.0'):
        num = num[:-2]
    return f"{num} {_UNIT_MAP.get(unit, unit)}"


def normalize_prod_name(name):
    """Normaliza nombre de producto para DISPLAY: corrige marca VECA->AV y
    aliases conocidos, pero conserva tildes/formato original cuando no hay
    alias aplicable (no se usa para matching — ver _normalize_key)."""
    upper = str(name).strip().upper()
    upper = re.sub(r'\(.*?\)', '', upper).strip()  # notas de empaque/lote/moneda
    upper = NAME_NORMALIZE.get(upper, upper)
    upper = upper.replace('VECA ', 'AV ').replace('VECASIL', 'AV SIL').replace('AV-', 'AV ')
    upper = re.sub(r'\s{2,}', ' ', upper).strip()
    upper = NAME_NORMALIZE.get(upper, upper)
    return upper


def _normalize_key(name):
    """Clave de matching contra tablas de precio piso: normalize_prod_name + sin tildes."""
    return strip_accents(normalize_prod_name(name))


_FMT_UNIT_RE = r'(?:LTS?|LITROS?|GRS?|GRAMOS?|KGS?|KILOS?|ML|CC)'


def parse_producto_formato(name):
    """Extrae (producto_base, formato) tolerando separador '-', 'X' o ninguno,
    y ruido posterior a la unidad (ej. '(VECA)', '(PRECIO EN USD)', códigos).
    A diferencia de la versión original, NO exige que el formato sea lo
    último en el string — basta con que aparezca tras el nombre."""
    s = str(name).strip()
    if not s or s.upper() == 'NAN':
        return '', ''
    patterns = [
        r'^(.+?)\s*-\s*(\d[\d.,]*\s*[A-Za-zÀ-ÿ%]+)\b',     # separador '-'
        r'^(.+?)\s+X\s*(\d[\d.,]*\s*[A-Za-zÀ-ÿ%]+)\b',      # separador 'X'/'x'
        r'^(.+?)\s+(\d[\d.,]*\s*' + _FMT_UNIT_RE + r')\b',  # sin separador (unidad reconocida)
    ]
    for pat in patterns:
        m = re.match(pat, s, re.IGNORECASE)
        if m:
            prod = m.group(1).strip().upper()
            fmt  = normalize_fmt(m.group(2))
            if prod and fmt:
                return prod, fmt
    return s.upper(), ''


def buscar_piso_chile(piso, prod_base_raw, fmt):
    """Busca entrada de piso Chile por (producto, formato) ya normalizados.
    piso: dict devuelto por load_piso_chile."""
    return piso.get((_normalize_key(prod_base_raw), fmt))


def load_piso_chile(piso_path):
    """Carga tabla de precios piso Chile: dict (prod_norm, fmt) -> {pp, costo_unidad,
    margen_calc, margen_propuesto, clasif}.

    Nota: el FORMATO '5 L (VECA)' es una fila duplicada de '5 L' con valores
    de costo/precio IDÉNTICOS (verificado 2026-06-24 contra el archivo real —
    no es una variante de costo distinta), por lo que se trata como el mismo
    formato tras limpiar la anotación '(VECA)'.
    """
    df = pd.read_excel(piso_path, sheet_name='Pricing Piso Chile', header=4)
    df.columns = [str(c).strip() for c in df.columns]
    pp_col    = [c for c in df.columns if 'PRECIO PISO' in c.upper() and 'CALCULADO' in c.upper()][0]
    prop_col  = [c for c in df.columns if 'NUEVO PRECIO' in c.upper() or 'PROPUESTO' in c.upper()]
    prop_col  = prop_col[0] if prop_col else None
    costo_col = [c for c in df.columns if 'COSTO' in c.upper() and 'UNIDAD' in c.upper()][0]
    marg_col  = [c for c in df.columns if 'MARGEN' in c.upper() and 'CALC' in c.upper()][0]
    margp_col = [c for c in df.columns if 'MARGEN' in c.upper() and 'PROPUESTO' in c.upper()]
    margp_col = margp_col[0] if margp_col else None
    clasif_col = [c for c in df.columns if 'CLASIF' in c.upper()]
    clasif_col = clasif_col[0] if clasif_col else None

    piso = {}
    for _, row in df.iterrows():
        prod = str(row.get('PRODUCTO', '')).strip().upper()
        fmt_raw = str(row.get('FORMATO', '')).strip().upper()
        if not prod or prod == 'NAN':
            continue
        fmt = normalize_fmt(fmt_raw)
        if not fmt:
            continue

        pp_val = pd.to_numeric(row.get(prop_col, np.nan) if prop_col else np.nan, errors='coerce')
        if pd.isna(pp_val):
            pp_val = pd.to_numeric(row[pp_col], errors='coerce')
        costo_u = pd.to_numeric(row.get(costo_col, np.nan), errors='coerce')
        marg_calc = pd.to_numeric(row.get(marg_col, np.nan), errors='coerce')
        marg_prop = pd.to_numeric(row.get(margp_col, np.nan), errors='coerce') if margp_col else np.nan
        clasif = str(row.get(clasif_col, '')).strip() if clasif_col else ''

        piso[(_normalize_key(prod), fmt)] = {
            'pp':               round(float(pp_val)) if pd.notna(pp_val) else None,
            'costo_unidad':     round(float(costo_u), 2) if pd.notna(costo_u) else None,
            'margen_calc':      round(float(marg_calc), 4) if pd.notna(marg_calc) else None,
            'margen_propuesto': round(float(marg_prop), 4) if pd.notna(marg_prop) else None,
            'clasif':           clasif,
        }
    return piso


def compute_iec_chile(df_ventas, piso):
    """
    IEC PONDERADO (Fase 7, 2026-07-28).
    Fórmula maestra: IEC = Σ venta_neta_elegible / Σ valor_piso_teorico
    donde valor_piso_teorico_i = Cantidad_i × precio_piso_unitario_i.
    Aggregable a cualquier nivel: vendedor, país, grupo, YTD.
    NO usar promedio simple de IEC individuales — siempre sumar numeradores y denominadores.

    DIFERENCIA CON FÓRMULA ANTERIOR:
    - Antes: iec = ventas_sobre_piso / total_elegible  (binario: en piso o no)
    - Ahora: iec = Σ venta_neta / Σ (qty × piso)  (ponderado por magnitud real)
    Ejemplo: venta $120 con piso $100 → antes: si venta > piso, $120/$120 = 100%; ahora: $120/$100 = 120%.
    """
    df = df_ventas.copy()
    df['total_n'] = pd.to_numeric(df['Total'],     errors='coerce').fillna(0)
    df['cant_n']  = pd.to_numeric(df['Cantidad'],  errors='coerce').fillna(0)
    df['pv_n']    = pd.to_numeric(df['Precio Uni'],errors='coerce')
    df['prod_base'], df['fmt_parsed'] = zip(*df['Producto'].apply(parse_producto_formato))

    def get_pp(row):
        entry = buscar_piso_chile(piso, row['prod_base'], row['fmt_parsed'])
        return entry['pp'] if entry else None

    df['pp'] = df.apply(get_pp, axis=1)

    # valor_piso_teorico = Cantidad × precio_piso  (denominador IEC ponderado)
    # FIX 2026-07-31: excluir filas con total_n=0 (muestras/demos) del VPT
    # para alinear criterio con build_tx_cl() que exige total>0 para elegibilidad.
    # Filas total=0 con cant_n>0 inflan el denominador sin aporte a VNE → IEC artificialmente bajo.
    df['valor_piso_teorico'] = df.apply(
        lambda r: r['cant_n'] * r['pp'] if pd.notna(r['pp']) and r['cant_n'] > 0 and r['total_n'] > 0 else 0, axis=1
    )
    # venta_neta_elegible = total de la transacción cuando tiene precio piso definido
    df['venta_neta_elegible'] = df.apply(
        lambda r: r['total_n'] if pd.notna(r['pp']) and r['total_n'] > 0 else 0, axis=1
    )
    # Bajo piso: precio unitario de venta < precio piso
    df['cumple']   = df.apply(lambda r: 1 if pd.notna(r['pp']) and pd.notna(r['pv_n']) and r['pv_n'] >= r['pp'] else 0, axis=1)
    df['bp']       = df.apply(lambda r: r['total_n'] if r['valor_piso_teorico'] > 0 and r['cumple'] == 0 else 0, axis=1)
    # IEC por línea (ratio, no %)
    df['iec_linea'] = df.apply(
        lambda r: round(r['total_n'] / r['valor_piso_teorico'], 4)
        if r['valor_piso_teorico'] > 0 else None, axis=1
    )

    # IEC ponderado por vendedor: Σ venta_neta_elegible / Σ valor_piso_teorico
    rtc_map_inv = {
        'PABLO LARATRO': 'laratro', 'FRANCISCO VELASQUEZ': 'velasquez',
        'JORGE CAROCA': 'caroca', 'RODRIGO ENCINA': 'encina',
        'IVAN VEVERKA': 'veverka', 'VALENTINA MUÑOZ': 'munoz',
    }
    iec_vend = {}
    for vend, key in rtc_map_inv.items():
        sub  = df[df['Vendedor'] == vend]
        num  = sub['venta_neta_elegible'].sum()
        den  = sub['valor_piso_teorico'].sum()
        iec_vend[key] = round(num / den, 4) if den > 0 else None

    # IEC ponderado total (país)
    total_num  = df['venta_neta_elegible'].sum()
    total_den  = df['valor_piso_teorico'].sum()
    iec_total  = round(total_num / total_den, 4) if total_den > 0 else None

    # IEC ponderado por cliente (RUT)
    iec_cliente = {}
    g = df.groupby('Rut').agg(
        vne =('venta_neta_elegible', 'sum'),
        vpt =('valor_piso_teorico',  'sum'),
        bp  =('bp',   'sum'),
        tx_elig=('venta_neta_elegible', lambda x: (x > 0).sum()),
        tx_cumple=('cumple', 'sum'),
    )
    for rut, row in g.iterrows():
        iec_cliente[str(rut)] = {
            'pct':            round(row['vne'] / row['vpt'], 4) if row['vpt'] > 0 else None,
            'tx_elegible':    int(row['tx_elig']),
            'tx_cumple':      int(row['tx_cumple']),
            'monto_elegible': int(row['vne']),
            'monto_bajo_piso':int(row['bp']),
            # raw para agregación YTD/grupo sin perder identidad matemática
            'venta_neta_elegible':  int(row['vne']),
            'valor_piso_teorico':   int(row['vpt']),
        }

    # IEC ponderado mensual por vendedor y total — para drill-down por mes en AVBOARD
    # Fórmula identica: cada mes = Σ vne / Σ vpt del mes.
    # Retorna dict: key → lista de 12 valores (uno por mes, None si sin datos).
    iec_mensual = {}
    for vend, key in rtc_map_inv.items():
        vals = []
        for mes in MESES_FULL:
            sub_m = df[(df['MES'] == mes) & (df['Vendedor'] == vend)]
            n_m   = sub_m['venta_neta_elegible'].sum()
            d_m   = sub_m['valor_piso_teorico'].sum()
            vals.append(round(n_m / d_m, 4) if d_m > 0 else None)
        iec_mensual[key] = vals
    # Total país por mes
    total_mes = []
    for mes in MESES_FULL:
        sub_m = df[df['MES'] == mes]
        n_m   = sub_m['venta_neta_elegible'].sum()
        d_m   = sub_m['valor_piso_teorico'].sum()
        total_mes.append(round(n_m / d_m, 4) if d_m > 0 else None)
    iec_mensual['total'] = total_mes

    return {
        'df':           df,
        'por_vendedor': iec_vend,
        'total':        iec_total,
        'bp_total':     int(df['bp'].sum()),
        'por_cliente':  iec_cliente,
        # raw del país para drill-down, YTD y Grupo (Σnum/Σden — nunca promedio)
        'venta_neta_elegible_total': int(total_num),
        'valor_piso_teorico_total':  int(total_den),
        # desglose mensual (Fase 7): país y por vendedor
        'iec_mensual': iec_mensual,
    }


def compute_iec_peru(tx_pe):
    """
    IEC PONDERADO (Fase 7) para Perú — desde la lista TX_PE (build_tx_pe).
    Fórmula idéntica a compute_iec_chile:
        IEC = Σ venta_neta_elegible / Σ valor_piso_teorico
        vne  = total de la transacción cuando elegible=True
        vpt  = qty × precio_piso de la misma transacción

    Retorna dict compatible con el bloque iec: { total, aguirre, ... }
    que consume render_avboard_data_js().

    REEMPLAZA el hardcode anterior:
        # DEPRECATED — hardcode eliminado 2026-07-31
        atalaya: 0.867, valladares: 0.167, ...
    """
    _vend_map = {
        'OSCAR INFANTE':       'infante',
        'NICOLL NAVARRO':      'navarro',
        'OMAR ATALAYA':        'atalaya',
        'ANTONIO GONZALES':    'gonzales',
        'LISBETH AGUIRRE':     'aguirre',
        'LIZBETH AGUIRRE':     'aguirre',  # alias DQ-002
        'PATRICIA VALLADARES': 'valladares',
        'SUSAN DIAZ':          'diaz',
        'SUSAN DÍAZ':          'diaz',
    }

    agg = {}  # key → {'vne': float, 'vpt': float}
    total_vne, total_vpt = 0.0, 0.0

    for tx in tx_pe:
        if not tx.get('elegible'):
            continue
        vne = float(tx.get('total', 0) or 0)
        qty = float(tx.get('qty',   0) or 0)
        pp  = float(tx.get('pp',    0) or 0)
        vpt = qty * pp
        total_vne += vne
        total_vpt += vpt

        key = _vend_map.get(str(tx.get('vendedor', '')).strip().upper())
        if key:
            if key not in agg:
                agg[key] = {'vne': 0.0, 'vpt': 0.0}
            agg[key]['vne'] += vne
            agg[key]['vpt'] += vpt

    def _ratio(num, den):
        return round(num / den, 4) if den > 0 else None

    result = {
        'total':       _ratio(total_vne, total_vpt),
        'aguirre':     _ratio(agg.get('aguirre',     {}).get('vne', 0), agg.get('aguirre',     {}).get('vpt', 0)),
        'infante':     _ratio(agg.get('infante',     {}).get('vne', 0), agg.get('infante',     {}).get('vpt', 0)),
        'atalaya':     _ratio(agg.get('atalaya',     {}).get('vne', 0), agg.get('atalaya',     {}).get('vpt', 0)),
        'valladares':  _ratio(agg.get('valladares',  {}).get('vne', 0), agg.get('valladares',  {}).get('vpt', 0)),
        'gonzales':    _ratio(agg.get('gonzales',    {}).get('vne', 0), agg.get('gonzales',    {}).get('vpt', 0)),
        'navarro':     _ratio(agg.get('navarro',     {}).get('vne', 0), agg.get('navarro',     {}).get('vpt', 0)),
        'diaz':        _ratio(agg.get('diaz',        {}).get('vne', 0), agg.get('diaz',        {}).get('vpt', 0)),
        # Raw para agregación Grupo (Σnum/Σden — no promedio)
        'vne_total':   round(total_vne, 2),
        'vpt_total':   round(total_vpt, 2),
        'impacto_potencial_usd': 4000,   # placeholder hasta tener cálculo real
    }
    # Limpiar claves con vne=0 y vpt=0 (vendedor sin elegibles) → None
    for key in ['aguirre', 'infante', 'atalaya', 'valladares', 'gonzales', 'navarro', 'diaz']:
        if result[key] is None and agg.get(key, {}).get('vne', 0) == 0:
            result[key] = None
    return result


def iec_factor(pct):
    """Convierte IEC % a factor de comisión."""
    if pct is None:
        return None
    if pct < 0.70:  return 0.20
    if pct < 0.85:  return 0.70
    if pct < 0.92:  return 0.80
    if pct < 0.95:  return 0.90
    return 1.05


# ══════════════════════════════════════════════════════════════════════════════
#  3.5 MÓDULO PRODUCTOS — rentabilidad real por país × producto × formato
# ══════════════════════════════════════════════════════════════════════════════
# Reemplaza la antigua tabla estática RENTABILIDAD. Cruza ventas reales contra
# precios piso (costo fábrica/unidad) para calcular margen real por SKU y
# detectar productos que destruyen margen, están subvaluados o no tienen
# costo cargado en la tabla piso (SIN_COSTO). Perú es best-effort: el piso
# perú está en 4 tramos fijos (1/20/200/1000 L) pero las ventas registran
# litros reales por transacción — se asigna al tramo más cercano en escala
# logarítmica (nearest_tier_pe). Validado 2026-06-24 contra Libro de Ventas
# 21-06-2026 + AGROVECA PERU VENTAS 22.06.2026: reconciliación exacta
# (sum(productos.ventas) == sum(Total/DOLARES) en ambos países, diff=0).

def _es_no_clasificable(prod_raw):
    """Detecta filas que no son SKUs reales: bolsones genéricos ('PRODUCTOS
    VARIOS', 'Prodructos de 1 litro') o servicios de laboratorio ('ANÁLISIS
    FOLIAR', 'ANALSIS SUELO', con o sin tilde/typo). Estas filas se excluyen
    de los cálculos de margen (no son productos con costo/piso asociado) pero
    se conservan en el detalle para trazabilidad."""
    s = strip_accents(str(prod_raw).strip().upper())
    partes = s.split(' ')
    primera = partes[0] if partes else ''
    segunda = partes[1] if len(partes) > 1 else ''
    if primera in ('PRODUCTOS', 'PRODUCTO', 'PRODRUCTOS', 'PRODRUCTO') and segunda in ('VARIOS', 'VARIO', 'DE'):
        return True
    return primera in ('ANALISIS', 'ANALSIS', 'ANALSIIS')


def compute_productos_chile(df_ventas, piso):
    """Agrega ventas Chile por (producto, formato) y cruza contra piso para
    calcular costo/margen real. 'Cantidad' en Libro de Ventas está en unidad
    base (litro/kg) — igual que 'COSTO $/UNIDAD' de piso — por lo que
    costo_total = costo_unidad × cantidad sin conversión de envases
    (verificado 2026-06-24: Cantidad × Precio Uni == Total exactamente)."""
    df = df_ventas.copy()
    df['total_n'] = pd.to_numeric(df['Total'], errors='coerce').fillna(0)
    df['cant_n']  = pd.to_numeric(df['Cantidad'], errors='coerce').fillna(0)
    df['prod_base'], df['fmt_parsed'] = zip(*df['Producto'].apply(parse_producto_formato))

    grupos = {}
    for _, row in df.iterrows():
        prod_raw, fmt = row['prod_base'], row['fmt_parsed']
        total, cant = float(row['total_n']), float(row['cant_n'])
        if total == 0 and cant == 0:
            continue
        if _es_no_clasificable(prod_raw):
            forzado = 'NO_CLASIFICABLE'
        elif not fmt:
            forzado = 'FORMATO_NO_IDENTIFICADO'
        else:
            forzado = None
        prod_disp = normalize_prod_name(prod_raw)
        key = (prod_disp, fmt or '?')
        g = grupos.setdefault(key, {'producto': prod_disp, 'formato': fmt or '?',
                                     'ventas': 0.0, 'cantidad': 0.0, 'forzado': forzado})
        g['ventas'] += total
        g['cantidad'] += cant
        if forzado:
            g['forzado'] = forzado

    productos = []
    for (prod_disp, fmt), g in grupos.items():
        entry = buscar_piso_chile(piso, prod_disp, fmt) if fmt != '?' else None
        ventas = round(g['ventas'])
        cantidad = round(g['cantidad'], 2)
        precio_prom = round(ventas / cantidad, 2) if cantidad else None
        if g['forzado']:
            estado = g['forzado']
            costo_unidad = entry['costo_unidad'] if entry else None
        elif entry and entry.get('costo_unidad') is not None:
            estado = 'OK'
            costo_unidad = entry['costo_unidad']
        else:
            estado = 'SIN_COSTO'
            costo_unidad = None
        costo_total  = round(costo_unidad * cantidad) if costo_unidad is not None else None
        margen_total = round(ventas - costo_total) if costo_total is not None else None
        margen_pct   = round(margen_total / ventas, 4) if (margen_total is not None and ventas) else None
        productos.append({
            'pais': 'CL', 'producto': prod_disp, 'formato': fmt,
            'ventas': ventas, 'cantidad': cantidad, 'precio_uni_prom': precio_prom,
            'costo_unidad': costo_unidad, 'costo_total': costo_total,
            'margen_total': margen_total, 'margen_pct': margen_pct,
            'piso': entry['pp'] if entry else None, 'clasif': entry['clasif'] if entry else None,
            'estado': estado,
        })
    return productos


_QTY_UNIT_PE_RE = re.compile(
    r'^(\d+(?:[.,]\d+)?)\s*(?:LTS?|TLS|[ÑN]TS|LITROS?)?\.?\s*(.+)$', re.IGNORECASE
)


def _clean_pe_name(raw):
    """Limpia ruido de envase/lote del CONCEPTO de venta Perú: '(CILINDRO...)
    LOTE:...', '(BIDON...) LOTE...', 'BOTELLA1LT' pegado al nombre."""
    s = str(raw).strip().upper()
    s = re.sub(r'\(.*', '', s)             # corta desde el primer paréntesis (envase/lote)
    s = re.sub(r'\bLOTE\b.*$', '', s)       # 'LOTE: PE L3 25' suelto sin paréntesis
    s = re.sub(r'\bBOTELLA\w*\b', '', s)    # 'BOTELLA1LT' pegado al nombre
    s = re.sub(r'\s{2,}', ' ', s).strip()
    return s


def parse_concepto_pe(concepto):
    """Extrae (qty_litros, nombre_producto_crudo) de un CONCEPTO de venta
    Perú. Retorna None si el CONCEPTO mezcla 2+ productos (',' o '/') — esos
    casos no se pueden atribuir a un solo SKU sin asumir datos y se reportan
    aparte como 'multi-producto' (best-effort, no es un error)."""
    s = str(concepto).strip()
    if not s or s.upper() == 'NAN':
        return None
    if ',' in s or '/' in s:
        return None
    m = _QTY_UNIT_PE_RE.match(s)
    if not m:
        return None
    qty = float(m.group(1).replace(',', '.'))
    name_raw = _clean_pe_name(m.group(2))
    for typo, fix in TYPO_FIX_PE.items():
        name_raw = name_raw.replace(typo, fix)
    return qty, name_raw


def nearest_tier_pe(qty, tiers):
    """Tramo de piso Perú (1/20/200/1000 L) más cercano a qty, en escala
    logarítmica — los tramos son múltiplos de ~10-20x entre sí, por lo que
    distancia log es más representativa que distancia lineal."""
    if not tiers or qty is None or qty <= 0:
        return None
    return min(tiers, key=lambda t: abs(math.log10(qty) - math.log10(t)))


def load_piso_peru(piso_path):
    """Carga tabla de precios piso Perú: {'entries': {(prod_norm, tramo_L): {...}},
    'tiers': {prod_norm: [tramos disponibles]}}. Misma nomenclatura "AV ..."
    que Chile (validado 2026-06-24)."""
    df = pd.read_excel(piso_path, sheet_name='Pricing Piso Peru', header=4)
    df.columns = [str(c).strip() for c in df.columns]
    pp_col    = [c for c in df.columns if 'PRECIO PISO' in c.upper() and 'CALCULADO' in c.upper()][0]
    prop_col  = [c for c in df.columns if 'NUEVO PRECIO' in c.upper() or 'PROPUESTO' in c.upper()]
    prop_col  = prop_col[0] if prop_col else None
    costo_col = [c for c in df.columns if 'COSTO' in c.upper() and 'UNIDAD' in c.upper()][0]
    marg_col  = [c for c in df.columns if 'MARGEN' in c.upper() and 'CALC' in c.upper()][0]
    margp_col = [c for c in df.columns if 'MARGEN' in c.upper() and 'PROPUESTO' in c.upper()]
    margp_col = margp_col[0] if margp_col else None
    clasif_col = [c for c in df.columns if 'CLASIF' in c.upper()]
    clasif_col = clasif_col[0] if clasif_col else None

    entries, tiers_by_prod = {}, {}
    for _, row in df.iterrows():
        prod = str(row.get('PRODUCTO', '')).strip().upper()
        fmt  = str(row.get('FORMATO', '')).strip().upper()
        if not prod or prod == 'NAN':
            continue
        m = re.match(r'^(\d+)', fmt)
        if not m:
            continue
        tier = int(m.group(1))
        prod_n = _normalize_key(prod)

        pp_val = pd.to_numeric(row.get(prop_col, np.nan) if prop_col else np.nan, errors='coerce')
        if pd.isna(pp_val):
            pp_val = pd.to_numeric(row[pp_col], errors='coerce')
        costo_u = pd.to_numeric(row.get(costo_col, np.nan), errors='coerce')
        marg_calc = pd.to_numeric(row.get(marg_col, np.nan), errors='coerce')
        marg_prop = pd.to_numeric(row.get(margp_col, np.nan), errors='coerce') if margp_col else np.nan
        clasif = str(row.get(clasif_col, '')).strip() if clasif_col else ''

        entries[(prod_n, tier)] = {
            'pp':               round(float(pp_val), 2) if pd.notna(pp_val) else None,
            'costo_unidad':     round(float(costo_u), 4) if pd.notna(costo_u) else None,
            'margen_calc':      round(float(marg_calc), 4) if pd.notna(marg_calc) else None,
            'margen_propuesto': round(float(marg_prop), 4) if pd.notna(marg_prop) else None,
            'clasif':           clasif,
        }
        tiers_by_prod.setdefault(prod_n, []).append(tier)
    return {'entries': entries, 'tiers': tiers_by_prod}


def extract_peru_ventas_sku(path):
    """Extrae el detalle de ventas Perú por transacción (CONCEPTO) desde la
    hoja 'VENTAS ACUMULADAS 2026' — usado solo por el módulo Productos
    (extract_peru_ventas() sigue siendo la fuente de los KPIs agregados por
    vendedor/mes que consume el resto del pipeline)."""
    df = pd.read_excel(path, sheet_name='VENTAS ACUMULADAS 2026', header=5)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[df['PERIODO'].notna()].copy()
    df['DOLARES'] = pd.to_numeric(df['DOLARES'], errors='coerce').fillna(0)
    return df


def buscar_piso_peru(piso_pe, prod_raw, qty):
    """Busca entrada de piso Perú por nombre normalizado + tramo más cercano
    (escala log) al qty real de la transacción."""
    prod_n = _normalize_key(prod_raw)
    tiers = piso_pe['tiers'].get(prod_n)
    if not tiers:
        return None, prod_n
    tier = nearest_tier_pe(qty, tiers)
    return piso_pe['entries'].get((prod_n, tier)), prod_n


def compute_productos_peru(df_sku, piso_pe):
    """Agrega ventas Perú por (producto, tramo_asignado) y cruza contra
    piso. Retorna (productos, multi_producto) — multi_producto resume el
    monto en CONCEPTOs que mezclan 2+ productos y no se pudieron atribuir."""
    productos_map = {}
    multi_total, multi_n = 0.0, 0
    for _, row in df_sku.iterrows():
        monto = float(row['DOLARES'])
        parsed = parse_concepto_pe(row.get('CONCEPTO', ''))
        if parsed is None:
            multi_total += monto
            multi_n += 1
            continue
        qty, name_raw = parsed
        entry, prod_n = buscar_piso_peru(piso_pe, name_raw, qty)
        tier = nearest_tier_pe(qty, piso_pe['tiers'].get(prod_n)) if entry else None
        fmt_label = f"{tier} L (tier)" if tier else '?'
        key = (prod_n, fmt_label)
        g = productos_map.setdefault(key, {'producto': normalize_prod_name(name_raw), 'formato': fmt_label,
                                            'ventas': 0.0, 'cantidad': 0.0, 'entry': entry})
        g['ventas'] += monto
        g['cantidad'] += qty

    productos = []
    for (prod_n, fmt_label), g in productos_map.items():
        entry = g['entry']
        ventas = round(g['ventas'], 2)
        cantidad = round(g['cantidad'], 2)
        precio_prom = round(ventas / cantidad, 4) if cantidad else None
        if entry and entry.get('costo_unidad') is not None:
            estado, costo_unidad = 'OK', entry['costo_unidad']
        else:
            estado, costo_unidad = 'SIN_COSTO', None
        costo_total  = round(costo_unidad * cantidad, 2) if costo_unidad is not None else None
        margen_total = round(ventas - costo_total, 2) if costo_total is not None else None
        margen_pct   = round(margen_total / ventas, 4) if (margen_total is not None and ventas) else None
        productos.append({
            'pais': 'PE', 'producto': g['producto'], 'formato': fmt_label,
            'ventas': ventas, 'cantidad': cantidad, 'precio_uni_prom': precio_prom,
            'costo_unidad': costo_unidad, 'costo_total': costo_total,
            'margen_total': margen_total, 'margen_pct': margen_pct,
            'piso': entry['pp'] if entry else None, 'clasif': entry['clasif'] if entry else None,
            'estado': estado,
        })
    return productos, {'monto': round(multi_total, 2), 'n': multi_n}


def compute_productos(cl_df, piso_cl, pe_df_sku, piso_pe):
    """Construye el módulo PRODUCTOS: rentabilidad real por país×producto×
    formato. Detecta productos que destruyen margen (alertas_nivel1, margen
    negativo), subvaluados/en zona de riesgo (alertas_nivel2, margen 0-10%)
    y sin costo cargado en piso (SIN_COSTO, oportunidad de completar datos).
    Perú es best-effort (ver nearest_tier_pe) — validar con Javier antes de
    tomar decisiones de pricing basadas solo en los números de Perú."""
    prod_cl = compute_productos_chile(cl_df, piso_cl)
    prod_pe, multi_pe = compute_productos_peru(pe_df_sku, piso_pe)
    productos = prod_cl + prod_pe

    con_costo = [p for p in productos if p['estado'] == 'OK' and p['margen_pct'] is not None]
    nivel1 = sorted([p for p in con_costo if p['margen_pct'] < 0], key=lambda p: p['margen_total'])
    nivel2 = sorted([p for p in con_costo if 0 <= p['margen_pct'] < 0.10], key=lambda p: p['margen_pct'])

    def _sku_label(p):
        return f"{p['producto']} {p['formato']}".strip()

    alertas_nivel1 = [
        {'pais': p['pais'], 'sku': _sku_label(p), 'margen': p['margen_pct'], 'margen_total': p['margen_total'],
         'accion': 'REVISAR_O_DESCONTINUAR'}
        for p in nivel1
    ]
    alertas_nivel2 = [
        {'pais': p['pais'], 'sku': _sku_label(p), 'margen': p['margen_pct']}
        for p in nivel2
    ]

    impacto_clp = int(round(sum(
        (p['margen_total'] if p['pais'] == 'CL' else p['margen_total'] * TC_CLP_USD)
        for p in nivel1
    ))) if nivel1 else 0

    def _bajo_piso(plist):
        return sum(1 for p in plist if p['estado'] == 'OK' and p['piso'] and p['precio_uni_prom'] is not None
                    and p['precio_uni_prom'] < p['piso'])

    resumen = {
        'alertas_nivel1':       alertas_nivel1,
        'alertas_nivel2':       alertas_nivel2,
        'impacto_clp':          impacto_clp,
        'skus_bajo_piso_chile': _bajo_piso(prod_cl),
        'skus_bajo_piso_peru':  _bajo_piso(prod_pe),
        'skus_sin_costo_chile': sum(1 for p in prod_cl if p['estado'] == 'SIN_COSTO'),
        'skus_sin_costo_peru':  sum(1 for p in prod_pe if p['estado'] == 'SIN_COSTO'),
        'multi_producto_peru':  multi_pe,
    }
    return {'detalle': productos, 'resumen': resumen}


# ══════════════════════════════════════════════════════════════════════════════
#  4. GENERACIÓN avboard_data.js
# ══════════════════════════════════════════════════════════════════════════════

def _fn(v):
    """Formatea float o None como literal JS (number o null)."""
    return f"{v:.4f}" if v is not None else 'null'


def _j(v, indent=0):
    """Serializa a JSON compacto."""
    return _jdumps(v)


def _arr(lst, per_row=12, indent=6):
    """Formatea array de números en columnas."""
    pad = ' ' * indent
    chunks = [lst[i:i+per_row] for i in range(0, len(lst), per_row)]
    return '[\n' + ',\n'.join(pad + str(c) for c in chunks) + '\n' + ' '*(indent-2) + ']'


def render_avboard_data_js(cl_v, pe_v, cl_cxc, iec_cl, cortes, pe_cxc, productos_data, iec_pe=None, insights=None):
    """Genera el contenido completo de avboard_data.js.

    iec_pe: resultado de compute_iec_peru(tx_pe) — dict con {total, aguirre, ...}
            Si None, el IEC Perú se omite (queda null). Siempre debe pasarse.
    """

    version = datetime.now().strftime('%Y-%m-%d')
    tc = TC_CLP_USD

    chile_ytd_usd = round(cl_v['ytd_5m'] / tc)
    peru_ytd_usd  = pe_v['ytd_5m']
    ytd_usd = chile_ytd_usd + peru_ytd_usd
    ytd_clp = cl_v['ytd_5m'] + round(pe_v['ytd_5m'] * tc)

    # IEC Grupo: Σnum/Σden de todos los países con datos de precio piso.
    # Perú: sin datos de precio piso reales → se excluye del numerador/denominador.
    # Grupo IEC = Chile IEC (únicamente) hasta que Perú disponga de piso por transacción.
    _g_vne = iec_cl['venta_neta_elegible_total']
    _g_vpt = iec_cl['valor_piso_teorico_total']
    grupo_iec_total = round(_g_vne / _g_vpt, 4) if _g_vpt > 0 else None
    grupo_iec_val   = f"{grupo_iec_total:.4f}" if grupo_iec_total is not None else 'null'

    def js_rtc_mensual(rtc_dict):
        lines = []
        for k, arr in sorted(rtc_dict.items()):
            vals = ', '.join(str(v) for v in arr)
            lines.append(f'      {k}: [{vals}]')
        return '{\n' + ',\n'.join(lines) + '\n    }'

    def js_rtc_t1(rtc_dict):
        lines = [f'      {k}: {v}' for k, v in sorted(rtc_dict.items())]
        return '{\n' + ',\n'.join(lines) + '\n    }'

    def js_por_rtc_cxc(por_rtc):
        lines = []
        for k, v in sorted(por_rtc.items(), key=lambda x: -x[1]['total']):
            lines.append(
                f"      {k}: {{\n"
                f"        total:   {v['total']},\n"
                f"        pct:     {v['pct']},\n"
                f"        vencida: {v['vencida']},\n"
                f"        t90:     {v['t90']},\n"
                f"        riesgo: '{v['riesgo']}'\n"
                f"      }}"
            )
        return '{\n' + ',\n'.join(lines) + '\n    }'

    def js_cuentas_criticas(lst):
        items = []
        for c in lst:
            parts = [f"        cliente: {_jdumps(c['cliente'])}"]
            if 'rtc' in c:  parts.append(f"        rtc: {_jdumps(c['rtc'])}")
            if 'dias' in c: parts.append(f"        dias: {_jdumps(c['dias'])}")
            parts.append(f"        monto: {c['monto']}")
            parts.append(f"        estado: {_jdumps(c['estado'])}")
            parts.append(f"        alerta: {_jdumps(c.get('alerta'))}")
            if 'nota' in c: parts.append(f"        nota: {_jdumps(c['nota'])}")
            items.append('      {\n' + ',\n'.join(parts) + '\n      }')
        return '[\n' + ',\n'.join(items) + '\n    ]'

    def js_iec(iec_cl):
        vends = {
            'total': iec_cl['total'],
            'velasquez': iec_cl['por_vendedor'].get('velasquez'),
            'laratro':   iec_cl['por_vendedor'].get('laratro'),
            'caroca':    iec_cl['por_vendedor'].get('caroca'),
            'encina':    iec_cl['por_vendedor'].get('encina'),
            'veverka':   iec_cl['por_vendedor'].get('veverka'),
            'munoz':     iec_cl['por_vendedor'].get('munoz'),
        }
        lines = []
        for k, v in vends.items():
            val = f"{v:.3f}" if v is not None else 'null'
            lines.append(f"      {k}: {val}")
        lines.append(f"      impacto_potencial_clp: {iec_cl['bp_total']}")
        # Raw para Grupo (Σnum/Σden — nunca promedio)
        lines.append(f"      vne_total: {iec_cl['venta_neta_elegible_total']}")
        lines.append(f"      vpt_total: {iec_cl['valor_piso_teorico_total']}")
        # Mensual por vendedor y total (Fase 7) — 12 valores, None→null
        mens = iec_cl.get('iec_mensual', {})
        def arr12(lst):
            return '[' + ', '.join(f"{v:.4f}" if v is not None else 'null' for v in (lst or [None]*12)) + ']'
        lines.append(f"      iec_mensual: {{\n"
                     f"        total:     {arr12(mens.get('total'))},\n"
                     f"        velasquez: {arr12(mens.get('velasquez'))},\n"
                     f"        laratro:   {arr12(mens.get('laratro'))},\n"
                     f"        caroca:    {arr12(mens.get('caroca'))},\n"
                     f"        encina:    {arr12(mens.get('encina'))},\n"
                     f"        veverka:   {arr12(mens.get('veverka'))},\n"
                     f"        munoz:     {arr12(mens.get('munoz'))}\n"
                     f"      }}")
        return '{\n' + ',\n'.join(lines) + '\n    }'

    def js_por_vendedor_pe(por_vend):
        lines = []
        for k, v in sorted(por_vend.items()):
            lines.append(
                f"      {k}: {{\n"
                f"        nombre: {_jdumps(v['nombre'])},\n"
                f"        ytd:    {v['ytd']},\n"
                f"        mayo:   {v['mayo']}\n"
                f"      }}"
            )
        return '{\n' + ',\n'.join(lines) + '\n    }'

    def js_productos(lst):
        items = []
        for p in lst:
            items.append(
                "    { "
                f"pais:{_jdumps(p['pais'])}, producto:{_jdumps(p['producto'])}, formato:{_jdumps(p['formato'])}, "
                f"ventas:{_jdumps(p['ventas'])}, cantidad:{_jdumps(p['cantidad'])}, precio_uni_prom:{_jdumps(p['precio_uni_prom'])}, "
                f"costo_unidad:{_jdumps(p['costo_unidad'])}, costo_total:{_jdumps(p['costo_total'])}, "
                f"margen_total:{_jdumps(p['margen_total'])}, margen_pct:{_jdumps(p['margen_pct'])}, "
                f"piso:{_jdumps(p['piso'])}, clasif:{_jdumps(p['clasif'])}, estado:{_jdumps(p['estado'])}"
                " }"
            )
        return '[\n' + ',\n'.join(items) + '\n  ]'

    def js_rentabilidad(r):
        def alert_line(a):
            parts = [f"pais:{_jdumps(a['pais'])}", f"sku:{_jdumps(a['sku'])}", f"margen:{a['margen']}"]
            if 'accion' in a: parts.append(f"accion:{_jdumps(a['accion'])}")
            return '{ ' + ', '.join(parts) + ' }'
        n1 = ', '.join(alert_line(a) for a in r['alertas_nivel1'])
        n2 = ', '.join(alert_line(a) for a in r['alertas_nivel2'])
        return (
            f"    alertas_nivel1: [{n1}],\n"
            f"    alertas_nivel2: [{n2}],\n"
            f"    impacto_clp:    {r['impacto_clp']},\n"
            f"    skus_bajo_piso_chile: {r['skus_bajo_piso_chile']},\n"
            f"    skus_bajo_piso_peru:   {r['skus_bajo_piso_peru']},\n"
            f"    skus_sin_costo_chile: {r['skus_sin_costo_chile']},\n"
            f"    skus_sin_costo_peru:   {r['skus_sin_costo_peru']}"
        )

    # Mensual arrays (12 months)
    cl_mensual = cl_v['mensual'] + [0] * (12 - len(cl_v['mensual']))
    pe_mensual = pe_v['mensual'] + [0] * (12 - len(pe_v['mensual']))
    ppto_cl    = PPTO_MENSUAL_CL
    ppto_pe    = PPTO_MENSUAL_PE

    cl_mensual_str = ', '.join(str(v) for v in cl_mensual)
    pe_mensual_str = ', '.join(str(v) for v in pe_mensual)
    ppto_cl_str    = ', '.join(str(v) for v in ppto_cl)
    ppto_pe_str    = ', '.join(str(v) for v in ppto_pe)

    # RTC ppto mensual (static)
    def js_rtc_ppto(d):
        lines = []
        for k, arr in sorted(d.items()):
            lines.append('      ' + k + ': [' + ', '.join(str(v) for v in arr) + ']')
        return '{\n' + ',\n'.join(lines) + '\n    }'

    # Peru ppto anual por vendedor
    ppto_pe_anual_str = ',\n      '.join(f"{k}: {v}" for k, v in PPTO_RTC_ANUAL_PE.items())

    js = f"""/**
 * ╔══════════════════════════════════════════════════════════════════╗
 * ║  AVBOARD · Capa de datos global · Agroveca Grupo LATAM 2026     ║
 * ║  Archivo: avboard_data.js                                        ║
 * ║  Propósito: fuente única de verdad para todos los dashboards     ║
 * ║                                                                  ║
 * ║  GENERADO AUTOMÁTICAMENTE — scripts/update_avboard.py            ║
 * ║  NO EDITAR MANUALMENTE                                           ║
 * ╚══════════════════════════════════════════════════════════════════╝
 *
 * Cortes:
 *   Chile ventas → {cortes['chile_ventas']}
 *   Chile CxC    → {cortes['chile_cxc']} (2 entidades)
 *   Perú ventas  → {cortes['peru_ventas']}
 *   Perú CxC     → {cortes['peru_cxc']}
 *
 * Actualizado: {version}
 */

var AVBOARD = (function() {{

  var meta = {{
    version:      '{version}',
    tc_clp_usd:   {tc},
    meta_mn:      0.25,
    cortes: {{
      chile_ventas: '{cortes['chile_ventas']}',
      chile_cxc:    '{cortes['chile_cxc']}',
      peru_ventas:  '{cortes['peru_ventas']}',
      peru_cxc:     '{cortes['peru_cxc']}'
    }},
    meses: {_jdumps(MESES)}
  }};

  var grupo = {{
    ytd_usd:      {ytd_usd},
    ytd_clp:      {ytd_clp},
    chile_ytd_usd: {chile_ytd_usd},
    peru_ytd_usd:  {peru_ytd_usd},
    rtc_activos:  12,
    mn_chile:     0.179,
    mn_peru:      null,
    // IEC Grupo ponderado (Fase 7): Σvne/Σvpt across countries con datos de piso.
    // Peru excluido hasta tener precio_piso por transacción. Nota: valor < 1.0 = bajo piso.
    iec_grupo: {grupo_iec_val},
    iec_grupo_nota: 'Chile solamente — Perú sin precio piso por transacción',
    iec_grupo_vne: {_g_vne},
    iec_grupo_vpt: {_g_vpt}
  }};

  var chile_ventas = {{
    ytd_5m:          {cl_v['ytd_5m']},
    ytd_4m:          {cl_v['ytd_4m']},
    mayo_parcial:    {cl_v['mayo_parcial']},
    ppto_anual:      {PPTO_ANUAL_CL},
    ppto_4m:         {cl_v['ppto_4m']},
    ppto_5m:         {cl_v['ppto_5m']},
    cumplimiento_4m: {cl_v['cumplimiento_4m']},
    cumplimiento_5m: {cl_v['cumplimiento_5m']},
    cumplimiento_t1: {round(sum(cl_v['mensual'][:3]) / sum(PPTO_MENSUAL_CL[:3]), 4)},
    mensual_real:  [{cl_mensual_str}],
    mensual_ppto:  [{ppto_cl_str}],
    rtc_real_t1:   {js_rtc_t1(cl_v['rtc_t1'])},
    rtc_ppto_t1: {{
      caroca:    {sum(PPTO_RTC_CL['caroca'][:3])},
      laratro:   {sum(PPTO_RTC_CL['laratro'][:3])},
      encina:    {sum(PPTO_RTC_CL['encina'][:3])},
      velasquez: {sum(PPTO_RTC_CL['velasquez'][:3])},
      veverka:   {sum(PPTO_RTC_CL['veverka'][:3])}
    }},
    rtc_mensual_real: {js_rtc_mensual(cl_v['rtc_mensual'])},
    rtc_mensual_ppto: {js_rtc_ppto(PPTO_RTC_CL)},
    iec: {js_iec(iec_cl)},
    mn_real:  0.179,
    mn_meta:  0.250
  }};

  var chile_cxc = {{
    corte:    '{cortes['chile_cxc']}',
    entidades: 2,
    total:    {cl_cxc['total']},
    vencida:  {cl_cxc['vencida']},
    al_dia:   {cl_cxc['al_dia']},
    por_entidad: {{
      agrocomercial: {{
        nombre: 'Agrocomercial',
        total:  {cl_cxc['por_entidad']['agrocomercial']['total']},
        tramos: {{
          t90:   {cl_cxc['por_entidad']['agrocomercial']['tramos']['t90']},
          t6190: {cl_cxc['por_entidad']['agrocomercial']['tramos']['t6190']},
          t3160: {cl_cxc['por_entidad']['agrocomercial']['tramos']['t3160']},
          t030:  {cl_cxc['por_entidad']['agrocomercial']['tramos']['t030']}
        }}
      }},
      agroveca_chile: {{
        nombre: 'Agroveca Chile',
        total:  {cl_cxc['por_entidad']['agroveca_chile']['total']},
        tramos: {{
          t90:   {cl_cxc['por_entidad']['agroveca_chile']['tramos']['t90']},
          t6190: {cl_cxc['por_entidad']['agroveca_chile']['tramos']['t6190']},
          t3160: {cl_cxc['por_entidad']['agroveca_chile']['tramos']['t3160']},
          t030:  {cl_cxc['por_entidad']['agroveca_chile']['tramos']['t030']}
        }}
      }}
    }},
    tramos: {{
      t90:   {cl_cxc['tramos']['t90']},
      t6190: {cl_cxc['tramos']['t6190']},
      t3160: {cl_cxc['tramos']['t3160']},
      t030:  {cl_cxc['tramos']['t030']}
    }},
    tramos_pct: {{
      t90:   {cl_cxc['tramos_pct']['t90']},
      t6190: {cl_cxc['tramos_pct']['t6190']},
      t3160: {cl_cxc['tramos_pct']['t3160']},
      t030:  {cl_cxc['tramos_pct']['t030']}
    }},
    por_rtc: {js_por_rtc_cxc(cl_cxc['por_rtc'])},
    cuentas_criticas: {js_cuentas_criticas(cl_cxc['cuentas_criticas'])}
  }};

  var peru_ventas = {{
    ytd_5m:       {pe_v['ytd_5m']},
    ytd_4m:       {pe_v['ytd_4m']},
    mayo_parcial: {pe_v['mayo_parcial']},
    ppto_anual:   {PPTO_ANUAL_PE},
    ppto_4m:      {pe_v['ppto_4m']},
    ppto_5m:      {pe_v['ppto_5m']},
    ppto_mes_label: '{pe_v['ppto_mes_label']}',
    cumplimiento_4m: {pe_v['cumplimiento_4m']},
    cumplimiento_5m: {pe_v['cumplimiento_5m']},
    mensual_real: [{pe_mensual_str}],
    mensual_ppto: [{ppto_pe_str}],
    por_vendedor: {js_por_vendedor_pe(pe_v['por_vendedor'])},
    rtc_ppto_anual: {{
      {ppto_pe_anual_str}
    }},
    rtc_mensual_ppto: {js_rtc_ppto(PPTO_RTC_MENSUAL_PE)},
    rtc_mensual_real: {js_rtc_mensual(pe_v['rtc_mensual'])},
    iec: {{
      total:      {_fn(iec_pe.get('total'))      if iec_pe else 'null'},
      aguirre:    {_fn(iec_pe.get('aguirre'))    if iec_pe else 'null'},
      infante:    {_fn(iec_pe.get('infante'))    if iec_pe else 'null'},
      atalaya:    {_fn(iec_pe.get('atalaya'))    if iec_pe else 'null'},
      valladares: {_fn(iec_pe.get('valladares')) if iec_pe else 'null'},
      gonzales:   {_fn(iec_pe.get('gonzales'))   if iec_pe else 'null'},
      navarro:    {_fn(iec_pe.get('navarro'))    if iec_pe else 'null'},
      diaz:       {_fn(iec_pe.get('diaz'))       if iec_pe else 'null'},
      vne_total:  {iec_pe.get('vne_total', 0)   if iec_pe else 0},
      vpt_total:  {iec_pe.get('vpt_total', 0)   if iec_pe else 0},
      impacto_potencial_usd: {iec_pe.get('impacto_potencial_usd', 4000) if iec_pe else 4000}
    }},
    mn_real:  null,
    mn_meta:  0.250
  }};

  var peru_cxc = {_jdumps({k: v for k, v in pe_cxc.items() if k != 'documentos'}, indent=2).replace(chr(10), chr(10)+'  ')};

  var productos = {js_productos(productos_data['detalle'])};

  var rentabilidad = {{
{js_rentabilidad(productos_data['resumen'])}
  }};

  return {{
    meta:         meta,
    grupo:        grupo,
    chile:  {{ ventas: chile_ventas, cxc: chile_cxc }},
    peru:   {{ ventas: peru_ventas,  cxc: peru_cxc  }},
    productos: productos,
    rentabilidad: rentabilidad,
    tc:       function() {{ return meta.tc_clp_usd; }},
    clpToUsd: function(clp) {{ return Math.round(clp / meta.tc_clp_usd); }},
    usdToClp: function(usd) {{ return usd * meta.tc_clp_usd; }},
    fmt_clp:  function(v) {{ return v.toLocaleString('es-CL'); }},
    fmt_usd:  function(v) {{ return v.toLocaleString('en-US', {{minimumFractionDigits:0, maximumFractionDigits:0}}); }},
    fmt_pct:  function(v) {{ return (v * 100).toFixed(1) + '%'; }},
    chile_ytd:     function() {{ return chile_ventas.ytd_5m; }},
    peru_ytd:      function() {{ return peru_ventas.ytd_5m; }},
    grupo_ytd_usd: function() {{ return grupo.ytd_usd; }},
    cxc_chile_t90: function() {{ return chile_cxc.tramos.t90; }},
    cxc_peru_t90:  function() {{ return peru_cxc.tramos.t90; }},
    cxc_alerta_loma_larga: function() {{
      return chile_cxc.cuentas_criticas.find(function(c) {{
        return c.cliente.indexOf('LOMA LARGA') >= 0;
      }});
    }}
  }};
}})();

(function verificarIntegridad() {{
  var ok = true; var errores = [];
  var sumaChile = AVBOARD.chile.ventas.mensual_real.reduce(function(a,b){{return a+b;}}, 0);
  if (sumaChile !== AVBOARD.chile.ventas.ytd_5m)
    errores.push('Chile mensual suma ' + sumaChile + ' ≠ ytd ' + AVBOARD.chile.ventas.ytd_5m);
  var sumaPeruR = AVBOARD.peru.ventas.mensual_real.reduce(function(a,b){{return a+b;}}, 0);
  if (Math.abs(sumaPeruR - AVBOARD.peru.ventas.ytd_5m) > 5)
    errores.push('Perú mensual suma ' + sumaPeruR + ' ≠ ytd ' + AVBOARD.peru.ventas.ytd_5m);
  var sumaCxcCH = AVBOARD.chile.cxc.por_entidad.agrocomercial.total +
                  AVBOARD.chile.cxc.por_entidad.agroveca_chile.total;
  if (sumaCxcCH !== AVBOARD.chile.cxc.total)
    errores.push('CxC entidades ' + sumaCxcCH + ' ≠ total ' + AVBOARD.chile.cxc.total);
  var t = AVBOARD.chile.cxc.tramos;
  var sumaTrCH = t.t90 + t.t6190 + t.t3160 + t.t030;
  if (sumaTrCH !== AVBOARD.chile.cxc.total)
    errores.push('CxC tramos ' + sumaTrCH + ' ≠ total ' + AVBOARD.chile.cxc.total);
  if (ok && errores.length === 0) {{
    console.log('[AVBOARD] ✅ Integridad OK · ' + AVBOARD.meta.version);
  }} else {{
    console.warn('[AVBOARD] ⚠ Errores:');
    errores.forEach(function(e){{ console.warn('  · ' + e); }});
  }}
}})();
"""
    # ── AI Insights ───────────────────────────────────────────────────────────
    import json as _jins
    _ins_js = _jins.dumps(insights, ensure_ascii=False) if insights else 'null'
    js += f'\nAVBOARD.insights = {_ins_js};\n'
    return js


# ══════════════════════════════════════════════════════════════════════════════
#  5. GENERACIÓN avboard_clientes.js
# ══════════════════════════════════════════════════════════════════════════════

def build_clientes(cl_v, pe_v, iec_cl, cl_cxc):
    """Construye arrays CLIENTES_CL y CLIENTES_PE."""

    df_cl = cl_v['df'].copy()
    df_cl['total_n'] = pd.to_numeric(df_cl['Total'], errors='coerce').fillna(0)
    df_cl = df_cl[df_cl['total_n'] > 0]

    # ── CLIENTES CHILE ──────────────────────────────────────────────────────
    clientes_cl = []

    # Group by RUT
    grupos = df_cl.groupby('Rut')
    total_cl_ytd = cl_v['ytd_5m'] or 1  # for pct

    for rut, sub in grupos:
        rut_str = str(rut).strip()
        nombre  = sub['Razón Social'].iloc[0] if len(sub) > 0 else rut_str
        vend    = sub['Vendedor'].mode().iloc[0] if len(sub) > 0 else ''
        region  = sub.get('Región', pd.Series([''])).iloc[0] if 'Región' in sub.columns else ''

        # Ventas mensuales
        mes_v = {}
        for m in MESES_FULL:
            v = int(sub[sub['MES'] == m]['total_n'].sum())
            if v > 0:
                mes_v[m] = v

        ytd    = int(sub['total_n'].sum())
        n_tx   = len(sub)
        n_docs = sub['Folio'].nunique() if 'Folio' in sub.columns else n_tx
        meses_act = len(mes_v)
        frec   = round(meses_act / 5, 2)  # 5 meses en el año
        ticket = round(ytd / n_docs) if n_docs > 0 else 0

        # Tendencia (mes actual vs mes anterior)
        last_val  = mes_v.get(MESES_FULL[4], 0)  # Mayo
        prev_val  = mes_v.get(MESES_FULL[3], 0)  # Abril
        if prev_val > 0 and last_val > 0:
            tend_pct  = round((last_val - prev_val) / prev_val * 100, 1)
            tend_pct  = max(-500, min(500, tend_pct))  # cap ±500%
            tendencia = 'creciente' if tend_pct > 0 else 'decreciente'
        else:
            tend_pct  = 0
            tendencia = 'estable'

        # IEC
        iec_data = iec_cl['por_cliente'].get(rut_str, {})
        iec_pct  = iec_data.get('pct')
        factor   = iec_factor(iec_pct)

        # CxC
        cxc_data = cl_cxc['cxc_por_cliente'].get(rut_str, {
            'saldo': 0, 'max_mora': 0, 'tramo': '', 'estado': 'Sin deuda'
        })

        # Productos top 5
        prod_g = sub.groupby('Producto')['total_n'].sum().sort_values(ascending=False)
        top_prods = []
        for prod, monto in prod_g.head(5).items():
            top_prods.append({
                'prod': str(prod),
                'monto': int(monto),
                'pct': round(int(monto) / ytd * 100, 1) if ytd > 0 else 0
            })
        n_skus = prod_g[prod_g > 0].count()

        # Score (0-100)
        score_iec  = (factor * 30) if factor is not None else 15  # neutro si sin IEC
        score_frec = frec * 25
        # CxC score
        max_mora = cxc_data.get('max_mora', 0)
        if max_mora > 90:   score_cxc = 0
        elif max_mora > 60: score_cxc = 10
        elif max_mora > 30: score_cxc = 15
        elif max_mora > 0:  score_cxc = 20
        else:               score_cxc = 25
        # Diversif
        score_div = min(10, n_skus * 1.5)
        # Volumen (vs mediana)
        score_vol = min(10, ytd / 5000000)  # 10 pts si >= 5M CLP
        score = round(score_iec + score_frec + score_cxc + score_div + score_vol)
        score = max(0, min(100, score))

        # Insights & alertas
        insights  = []
        alertas   = []
        oportunidades = []
        if iec_pct is not None and iec_pct < 0.70:
            alertas.append(f"IEC crítico {iec_pct*100:.1f}% — pricing sistemáticamente bajo piso")
        if cxc_data.get('max_mora', 0) > 90:
            alertas.append(f"CxC crítico +90d · CLP {cxc_data['saldo']:,}")
        if n_skus == 1:
            alertas.append("Mono-producto — riesgo de concentración")
        if tend_pct > 50:
            insights.append(f"Crecimiento {tend_pct:.1f}% vs mes anterior")
        if n_skus >= 5:
            oportunidades.append("Cliente diversificado — potencial premium")

        clientes_cl.append({
            'id':      f"cl_{rut_str.replace('.','').replace('-','')}",
            'nombre':  str(nombre).strip(),
            'rut':     rut_str,
            'pais':    'CL',
            'region':  str(region).strip(),
            'vendedor': str(vend).strip(),
            'moneda':  'CLP',
            'ventas':  {
                'ytd':          ytd,
                'mensual':      mes_v,
                'n_tx':         n_tx,
                'n_docs':       n_docs,
                'meses_activos': meses_act,
                'frecuencia':   frec,
                'ticket_prom':  ticket,
                'tendencia':    tendencia,
                'tendencia_pct': tend_pct,
            },
            'iec': {
                'pct':             iec_pct,
                'factor':          factor,
                'tx_elegible':     iec_data.get('tx_elegible', 0),
                'tx_cumple':       iec_data.get('tx_cumple', 0),
                'monto_elegible':  iec_data.get('monto_elegible', 0),
                'monto_bajo_piso': iec_data.get('monto_bajo_piso', 0),
            },
            'cxc': {
                'saldo':    cxc_data.get('saldo', 0),
                'max_mora': cxc_data.get('max_mora', 0),
                'tramo':    cxc_data.get('tramo', ''),
                'estado':   cxc_data.get('estado', 'Sin deuda'),
            },
            'productos': {
                'top':    top_prods,
                'n_skus': int(n_skus),
                'es_mono': n_skus == 1,
            },
            'score':         score,
            'insights':      insights,
            'alertas':       alertas,
            'oportunidades': oportunidades,
            'ranking':       0,   # filled below
            'pct_venta':     round(ytd / total_cl_ytd * 100, 2),
        })

    # Ranking by score desc, then ytd desc
    clientes_cl.sort(key=lambda c: (-c['score'], -c['ventas']['ytd']))
    for i, c in enumerate(clientes_cl):
        c['ranking'] = i + 1

    # ── CLIENTES PERÚ ───────────────────────────────────────────────────────
    # Peru uses VENTAS ACUMULADAS sheet
    clientes_pe = []
    # Minimal: keep existing structure with updated YTD values
    # (Peru client IEC not computed — no precio piso per-client)
    # We'll use a simplified version based on available data
    # This section can be expanded when Peru client data is available
    # For now, return empty list — the script preserves existing CLIENTES_PE
    # by extracting it from the current avboard_clientes.js

    return clientes_cl, clientes_pe


def render_avboard_clientes_js(clientes_cl, clientes_pe_raw, cortes, n_tx_cl):
    """Genera el contenido de avboard_clientes.js."""
    version = datetime.now().strftime('%Y-%m-%d')
    n_cl = len(clientes_cl)
    ytd_cl = sum(c['ventas']['ytd'] for c in clientes_cl)

    # CXC sin match (clients with CxC but no sales) - extracted from existing file
    cxc_sin_match = '[]'

    header = f"""/* avboard_clientes.js
 * Motor de datos consolidado · Panel Clientes AV Latam 2026
 * Generado: {version} | Corte: Chile {cortes['chile_ventas']} · CxC {cortes['chile_cxc']}
 * GENERADO AUTOMÁTICAMENTE — scripts/update_avboard.py
 *
 * Clientes Chile: {n_cl} · YTD CLP {ytd_cl:,}
 * Score: IEC×30 + Frecuencia×25 + CxC×25 + Diversif×10 + Volumen×10
 */

'use strict';
"""

    cl_json  = _jdumps(clientes_cl,  separators=(',', ':'))
    pe_json  = clientes_pe_raw  # keep existing

    return header + f"""
const CLIENTES_CL = {cl_json};

const CLIENTES_PE = {pe_json};

const CXC_SIN_MATCH_CL = {cxc_sin_match};
"""


# ══════════════════════════════════════════════════════════════════════════════
#  6. ACTUALIZACIÓN Panel_IEC (TX_CL)
# ══════════════════════════════════════════════════════════════════════════════

def build_tx_cl(df_ventas, piso):
    """Construye el array TX_CL completo desde el DataFrame de ventas Chile."""
    df = df_ventas.copy()
    df['total_n'] = pd.to_numeric(df['Total'], errors='coerce').fillna(0)
    df['pv_n']    = pd.to_numeric(df['Precio Uni'], errors='coerce')
    df['cant_n']  = pd.to_numeric(df.get('Cantidad', pd.Series(dtype=float)), errors='coerce')
    df['fecha_dt'] = pd.to_datetime(df['Fecha'], format='%d/%m/%Y', errors='coerce')
    df['prod_base'], df['fmt_parsed'] = zip(*df['Producto'].apply(parse_producto_formato))

    tx_list = []
    for _, row in df.iterrows():
        total = int(row['total_n'])
        pv    = float(row['pv_n']) if pd.notna(row['pv_n']) else None
        entry = buscar_piso_chile(piso, row['prod_base'], row['fmt_parsed'])
        pp    = float(entry['pp']) if entry and entry['pp'] is not None else None
        mes   = str(row.get('MES', '')).strip().upper()

        # Cantidad: usar columna directa si existe; derivar de Total/Precio Uni como fallback.
        # El SIC necesita qty para calcular IEC ponderado (qty × pp = valor_piso_teórico).
        # Verificado 2026-06-24: Cantidad × Precio Uni == Total exactamente en el Libro de Ventas.
        raw_cant = row['cant_n'] if pd.notna(row['cant_n']) else None
        if raw_cant is not None and raw_cant > 0:
            qty = round(float(raw_cant), 4)
        elif pv is not None and pv > 0 and total > 0:
            qty = round(total / pv, 4)
        else:
            qty = None

        if pp is not None and total > 0:
            elegible = True
            cumple   = pv is not None and pv >= pp
            sp       = total if cumple else 0
            bp       = total if not cumple else 0
        else:
            elegible = False
            cumple   = None
            sp       = 0
            bp       = 0

        fecha_str = row['fecha_dt'].strftime('%Y-%m-%d') if pd.notna(row['fecha_dt']) else ''
        tx_list.append({
            'mes':          mes,
            'fecha':        fecha_str,
            'folio':        str(row.get('Folio', '')),
            'doc':          str(row.get('Documento', ''))[:30],
            'cliente':      str(row.get('Razón Social', '')).strip(),
            'vendedor':     str(row.get('Vendedor', '')).strip(),
            'producto':     normalize_prod_name(row['prod_base']),
            'formato':      row['fmt_parsed'] if row['fmt_parsed'] else '?',
            'producto_orig': str(row.get('Producto', '')).strip(),
            'total':        total,
            'pv':           pv,
            'pp':           pp,
            'qty':          qty,
            'elegible':     elegible,
            'sp':           sp,
            'bp':           bp,
            'cumple':       cumple,
        })
    return tx_list


_MESES_ES = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
             'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
_MESES_LABEL = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
                'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

def _build_mes_options(meses_presentes, indent='      '):
    """Construye las <option> de mes para los selectores del Panel IEC.
    Solo incluye meses que tienen datos en TX_CL/TX_PE."""
    lines = []
    for cod, lbl in zip(_MESES_ES, _MESES_LABEL):
        if cod in meses_presentes:
            lines.append(f'{indent}<option value="{cod}">{lbl}</option>')
    return '\n'.join(lines)

def _sync_mes_selectors(content, meses_presentes):
    """Actualiza los <select id="fMes"> e <select id="ic-mes"> con los meses
    reales presentes en TX_CL. Evita que el selector quede desactualizado."""

    # ── fMes (filtro principal tabla tx) ──────────────────────────────────
    opts_fmes = _build_mes_options(meses_presentes, indent='      ')
    content = re.sub(
        r'(<select id="fMes"[^>]*>)\s*<option value="TODOS">.*?</option>.*?(</select>)',
        lambda m: (
            m.group(1) + '\n      <option value="TODOS">Todos (YTD)</option>\n'
            + opts_fmes + '\n    ' + m.group(2)
        ),
        content, flags=re.DOTALL
    )

    # ── ic-mes (selector IEC comparativo) ────────────────────────────────
    opts_icmes = _build_mes_options(meses_presentes, indent='          ')
    content = re.sub(
        r'(<select id="ic-mes"[^>]*>)\s*<option value="TODOS">.*?</option>.*?(</select>)',
        lambda m: (
            m.group(1) + '\n          <option value="TODOS">YTD 2026</option>\n'
            + opts_icmes + '\n        ' + m.group(2)
        ),
        content, flags=re.DOTALL
    )

    return content


def update_panel_iec(tx_cl, corte_date):
    """Reemplaza TX_CL en Panel_IEC, actualiza fechas de corte y sincroniza
    selectores de mes con los meses reales presentes en los datos."""
    with open(PANEL_IEC, 'r', encoding='utf-8') as f:
        content = f.read()

    tx_json = _jdumps(tx_cl, separators=(',', ':'))
    new_tx  = f'const TX_CL = {tx_json};'

    content = re.sub(r'const TX_CL = \[.*?\];', new_tx, content, flags=re.DOTALL)

    # Update corte dates
    content = re.sub(r'corte \d{2}/\d{2}/\d{4}', f'corte {corte_date}', content)
    content = re.sub(r'Datos corte \d{2}/\d{2}/\d{4}', f'Datos corte {corte_date}', content)

    # Sync month selectors from real data
    meses_en_datos = {t.get('mes','').upper() for t in tx_cl if t.get('mes')}
    meses_ordenados = [m for m in _MESES_ES if m in meses_en_datos]
    content = _sync_mes_selectors(content, meses_ordenados)

    with open(PANEL_IEC, 'w', encoding='utf-8') as f:
        f.write(content)

    return len(tx_cl)


def build_tx_pe(df_sku, piso_pe):
    """Construye el array TX_PE completo desde la hoja VENTAS ACUMULADAS 2026.

    FIX DQ-001: Parsea FECHA EMISION con format='%%d/%%m/%%Y' (DD/MM/YYYY)
    eliminando el bug histórico donde se leía como MM/DD/YYYY produciendo
    fechas invertidas (40 registros) o NaN (48 registros).

    FIX DQ-002: TX_PE ahora se regenera en cada corrida del pipeline,
    igual que TX_CL, en lugar de ser una constante estática.
    """
    _vend_map = {
        'OSCAR INFANTE':      'infante',
        'NICOLL NAVARRO':     'navarro',
        'OMAR ATALAYA':       'atalaya',
        'ANTONIO GONZALES':   'gonzales',
        'LISBETH AGUIRRE':    'aguirre',
        'LIZBETH AGUIRRE':    'aguirre',
        'PATRICIA VALLADARES':'valladares',
        'SUSAN DIAZ':         'diaz',
        'SUSAN DÍAZ':         'diaz',
    }

    df = df_sku.copy()
    required = {'PERIODO', 'FECHA EMISION', 'FACTURA', 'DENOMINACION O RAZON SOCIAL',
                'CONCEPTO', 'DOLARES', 'VENDEDOR'}
    missing = required - set(df.columns)
    if missing:
        print(f"   ⚠ build_tx_pe: columnas faltantes en df_sku: {missing} — TX_PE no actualizado")
        return []

    df = df[df['PERIODO'].notna()].copy()

    # ── Parseo de fecha con formato explícito DD/MM/YYYY (corrige DQ-001) ──
    df['fecha_dt'] = pd.to_datetime(
        df['FECHA EMISION'],
        format='%d/%m/%Y',
        dayfirst=True,
        errors='coerce'
    )

    # ── Extraer mes desde PERIODO: "01 ENERO 2026" → "ENERO" ──
    def _extract_mes(periodo):
        parts = str(periodo).strip().split()
        return parts[1].upper() if len(parts) >= 2 else ''
    df['mes_str'] = df['PERIODO'].apply(_extract_mes)

    df['DOLARES'] = pd.to_numeric(df['DOLARES'], errors='coerce').fillna(0)

    tx_list = []
    n_fecha_ok, n_fecha_nan = 0, 0

    for _, row in df.iterrows():
        total = float(row['DOLARES'])
        if total <= 0:
            continue

        fecha_dt = row['fecha_dt']
        if pd.notna(fecha_dt):
            fecha_str = fecha_dt.strftime('%Y-%m-%d')
            n_fecha_ok += 1
        else:
            fecha_str = ''
            n_fecha_nan += 1

        folio    = str(int(row['FACTURA'])) if pd.notna(row.get('FACTURA')) else ''
        cliente  = str(row.get('DENOMINACION O RAZON SOCIAL', '')).strip()[:80]
        vendedor = str(row.get('VENDEDOR', '')).strip()
        concepto = str(row.get('CONCEPTO', '')).strip()
        mes      = row['mes_str']

        # Decisiones GG nivel transacción — leídas desde gg_decisions.json (Master Data Tier 0)
        for _gg in _GG_DECISIONS:
            if (_gg.get('tipo') == 'reasignacion_transaccion'
                    and _gg.get('criterio', {}).get('campo') == 'folio'
                    and folio == _gg['criterio'].get('valor', '')):
                vendedor = _gg['accion'].get('vendor_display_override', vendedor)

        parsed = parse_concepto_pe(concepto)
        if parsed:
            qty, name_raw = parsed
            entry, prod_n = buscar_piso_peru(piso_pe, name_raw, qty)
            tier  = nearest_tier_pe(qty, piso_pe['tiers'].get(prod_n)) if entry else None
            fmt   = (f"{int(tier)} L" if isinstance(tier, (int, float)) else str(tier)) if tier else '?'
            prod  = normalize_prod_name(name_raw)
            pp    = float(entry['pp']) if entry and entry.get('pp') is not None else None
            pv    = round(total / qty, 4) if qty > 0 else None

            if pp is not None and total > 0:
                elegible = True
                cumple   = pv is not None and pv >= pp
                sp       = round(total if cumple else 0, 2)
                bp       = round(total if not cumple else 0, 2)
            else:
                elegible = False
                cumple   = None
                sp, bp   = 0, 0
        else:
            qty, prod, fmt, pp, pv = None, '?', '?', None, None
            elegible, cumple, sp, bp = False, None, 0, 0

        tx_list.append({
            'mes':      mes,
            'fecha':    fecha_str,
            'folio':    folio,
            'cliente':  cliente,
            'vendedor': vendedor,
            'concepto': concepto,
            'producto': prod,
            'formato':  fmt,
            'qty':      qty,
            'total':    round(total, 2),
            'pv':       pv,
            'pp':       pp,
            'elegible': elegible,
            'sp':       sp,
            'bp':       bp,
            'cumple':   cumple,
        })

    pct_ok = round(n_fecha_ok / max(len(tx_list), 1) * 100, 1)
    print(f"   → TX_PE: {len(tx_list)} registros · fechas OK: {n_fecha_ok} ({pct_ok}%) · NaN: {n_fecha_nan}")
    return tx_list


def update_panel_iec_pe(tx_pe, corte_date):
    """Reemplaza TX_PE en Panel_IEC y actualiza la fecha de corte Perú."""
    with open(PANEL_IEC, 'r', encoding='utf-8') as f:
        content = f.read()

    tx_json = _jdumps(tx_pe, separators=(',', ':'))
    new_tx  = f'const TX_PE = {tx_json};'
    content = re.sub(r'const TX_PE = \[.*?\];', new_tx, content, flags=re.DOTALL)

    with open(PANEL_IEC, 'w', encoding='utf-8') as f:
        f.write(content)

    return len(tx_pe)


def write_sic_tx_pe(tx_pe, ppto_rmp, corte_date):
    """Genera apps/sic_av/sic_tx_pe.js desde el mismo pipeline que avboard_data.js.

    C-04: Una corrida válida genera juntos avboard_data.js + sic_tx_pe.js.

    Args:
        tx_pe       — lista de dicts producida por build_tx_pe()
        ppto_rmp    — dict {vendedor_key: [12 valores mensuales]} (PPTO_RTC_MENSUAL_PE)
        corte_date  — string 'DD/MM/YYYY'
    """
    from datetime import datetime as _dt
    import json as _json

    SIC_TX_PE_PATH = REPO / 'apps' / 'sic_av' / 'sic_tx_pe.js'
    now_str = _dt.now().strftime('%Y-%m-%d %H:%M')

    # Serializar transacciones
    tx_json = _jdumps(tx_pe, separators=(',', ':'))

    # Serializar ppto mensual — orden canónico desde catálogo (V-08)
    if _VC_AVAILABLE:
        VEND_ORDER, _src_vo = _vc.get_vendor_order_safe("PE")
        if _src_vo == "legacy":
            print("   ⚠ write_sic_tx_pe: VEND_ORDER cargado desde LEGACY")
    else:
        VEND_ORDER = ['aguirre', 'atalaya', 'diaz', 'gonzales', 'infante', 'martha', 'valladares']
    ppto_lines = []
    for k in VEND_ORDER:
        vals = ppto_rmp.get(k, [0]*12)
        vals_str = '[' + ','.join(str(round(v, 1)) for v in vals) + ']'
        ppto_lines.append(f'      {k}: {vals_str}')
    ppto_block = ',\n'.join(ppto_lines)

    content = (
        f'/**\n'
        f' * SIC-AV — sic_tx_pe.js  (datos históricos Perú)\n'
        f' * Auto-generado — {now_str} | NO EDITAR MANUALMENTE\n'
        f' * Fuente: update_avboard.py | corte {corte_date}\n'
        f' */\n'
        f'(function(global){{\n'
        f'  "use strict";\n'
        f'  global.SIC_TX_PE = {tx_json};\n'
        f'  global.SIC_PPTO_PE = {{ rtc_mensual_ppto: {{\n'
        f'{ppto_block}\n'
        f'    }} }};\n'
        f'}})(typeof window !== "undefined" ? window : globalThis);\n'
    )

    with open(SIC_TX_PE_PATH, 'w', encoding='utf-8') as f:
        f.write(content)

    print(f"   → sic_tx_pe.js escrito · {len(tx_pe)} tx · {SIC_TX_PE_PATH}")
    return len(tx_pe)


# ══════════════════════════════════════════════════════════════════════════════
#  6.5 SINCRONIZAR CACHE-BUSTING (?v=) EN TODOS LOS PANELES HTML
# ══════════════════════════════════════════════════════════════════════════════

def sync_iec_hardcoded(iec_cl: float, iec_pe: float) -> int:
    """
    Actualiza los valores IEC hardcodeados en paneles que NO los leen dinámicamente
    desde avboard_data.js. Mantiene coherencia entre Panel_IEC_Auditoria (fuente
    de verdad) y el resto de vistas.

    Paneles afectados:
        Panel_Jefes_Index.html, Panel_Jefes_Chile_2026.html,
        Panel_Jefes_Peru_2026.html, Executive_Board_View_AV_Latam_2026.html,
        Panel_Rentabilidad_AV_2026.html, dashboard.html

    El IEC oficial es VNE/VPT (compute_iec_chile / compute_iec_peru).
    """
    cl_str = f"{iec_cl*100:.1f}%"
    pe_str = f"{iec_pe*100:.1f}%"

    # Color según valor: ≥100% → green, 85-99% → amber, <85% → red
    def _color(v):
        if v >= 1.0:   return 'var(--green)'
        if v >= 0.85:  return 'var(--amber)'
        return 'var(--red)'
    cl_color = _color(iec_cl)
    pe_color = _color(iec_pe)

    panels = [
        REPO / 'Panel_Jefes_Index.html',
        REPO / 'Panel_Jefes_Chile_2026.html',
        REPO / 'Panel_Jefes_Peru_2026.html',
        REPO / 'Executive_Board_View_AV_Latam_2026.html',
        REPO / 'Panel_Rentabilidad_AV_2026.html',
        REPO / 'dashboard.html',
    ]

    # Regex: reemplaza cualquier valor porcentual de IEC que el script haya escrito
    # antes. Busca el patrón canónico que estas páginas usan: NN.N% donde NN.N es
    # un número de 2–5 dígitos antes del punto. Solo modifica líneas que contienen
    # "IEC" cerca del número.
    IEC_NUM = re.compile(r'(\d{2,3}\.\d)%')

    n_updated = 0
    for p in panels:
        if not p.exists():
            continue
        txt = p.read_text(encoding='utf-8')
        changed = False

        # Estrategia: reemplazar los valores actuales (que el script escribió en la
        # corrida anterior) con los nuevos. Los valores anteriores son los que
        # corresponden a CL e IEC_PE en el texto. Usamos el patrón específico de
        # cada panel para no tocar valores no relacionados.

        # ── Patrón universal: líneas con IEC ──────────────────────────────────
        lines = txt.split('\n')
        new_lines = []
        i = 0
        while i < len(lines):
            line = lines[i]
            # Detectar bloque IEC Chile: línea menciona IEC y tiene un % numérico
            if ('IEC' in line or 'iec' in line) and IEC_NUM.search(line):
                # Reemplazar el primer número % en esta línea con el valor correcto
                # Determinar si es Chile o Peru por contexto
                ctx = line + (lines[i+1] if i+1 < len(lines) else '') + (lines[i-1] if i > 0 else '')
                if 'Chile' in ctx or 'CL' in ctx or 'chile' in ctx or '🇨🇱' in ctx:
                    new_line = IEC_NUM.sub(cl_str, line, count=1)
                    # Actualizar color inline si presente
                    new_line = re.sub(r'color:var\(--(?:red|amber|green)\)', f'color:{cl_color}', new_line, count=1)
                    changed = changed or (new_line != line)
                    line = new_line
                elif 'Per' in ctx or 'PE' in ctx or 'peru' in ctx or '🇵🇪' in ctx:
                    new_line = IEC_NUM.sub(pe_str, line, count=1)
                    new_line = re.sub(r'color:var\(--(?:red|amber|green)\)', f'color:{pe_color}', new_line, count=1)
                    changed = changed or (new_line != line)
                    line = new_line
            new_lines.append(line)
            i += 1

        if changed:
            p.write_text('\n'.join(new_lines), encoding='utf-8')
            n_updated += 1

    return n_updated


def sync_cxc_panel(cl_cxc: dict, pe_cxc: dict | None = None) -> bool:
    """
    Actualiza Panel_CxC_AV_Latam_2026.html en cada pipeline run:
    1. Regenera el array chileData (fallback HTML) desde cl_cxc.cuentas_criticas
    2. Actualiza el texto estático del alerta banner como fallback (JS lo sobreescribe
       al cargar, pero este fallback debe ser coherente con los datos actuales)

    Returns True si el archivo fue modificado.
    """
    cxc_file = REPO / 'Panel_CxC_AV_Latam_2026.html'
    if not cxc_file.exists():
        return False

    txt = cxc_file.read_text(encoding='utf-8')
    original = txt

    # ── 1. Regenerar chileData ────────────────────────────────────────────────
    criticos = cl_cxc.get('cuentas_criticas', [])
    corte    = cl_cxc.get('corte', '—')
    rows = []
    for c in criticos:
        dias   = c.get('dias') or 0
        tramo  = '+90d' if dias > 90 else '61-90d' if dias > 60 else '31-60d' if dias > 30 else '0-30d'
        estado = c.get('estado', 'CRÍTICO')
        est_s  = 'Crítico' if estado == 'CRÍTICO' else 'Alerta' if estado == 'ALERTA' else 'Monitoreo'
        rtc    = (c.get('rtc', '') or '').replace("'", r"\'")
        cli    = (c.get('cliente', '') or '').replace("'", r"\'")
        monto  = c.get('monto', 0)
        rows.append(
            f"  {{rtc:'{rtc}',cliente:'{cli}',folio:'',emision:'',vencimiento:'',"
            f"dias:{dias},tramo:'{tramo}',monto:{monto},estado:'{est_s}',condicion:'Crédito'}}"
        )
    body = (',\n'.join(rows)) if rows else '  // Sin cuentas críticas'
    new_data = f"const chileData = [\n  // Corte {corte} · generado por update_avboard.py\n{body}\n]"
    txt = re.sub(r'const chileData\s*=\s*\[.*?\]', new_data, txt, flags=re.DOTALL)

    # ── 2. Actualizar alerta banner (fallback) ────────────────────────────────
    t90   = cl_cxc.get('tramos', {}).get('t90', 0)
    total = cl_cxc.get('total', 0)
    pe_t90 = (pe_cxc or {}).get('tramos', {}).get('t90', 0)
    n_crit = len([c for c in criticos if (c.get('estado') or '') == 'CRÍTICO'])
    top    = criticos[0] if criticos else None

    if t90 > 0:
        icon = '🚨'
        alert = f'<strong>ALERTA — MORA +90 DÍAS ACTIVA.</strong> Chile +90d: <strong>CLP {t90/1e6:.1f}M CRÍTICO</strong>'
        if top:
            alert += f' — {top.get("cliente","?")} <strong style="color:var(--red)">CLP {top.get("monto",0)/1e6:.1f}M ({top.get("dias","?")}d)</strong>'
        if n_crit > 1:
            alert += f' y {n_crit-1} cuentas más.'
    else:
        icon  = '✅'
        alert = '<strong>✅ SIN MORA CRÍTICA (+90d) en Agrocomercial.</strong>'
    alert += f' Cartera total Chile (2 entidades): <strong>CLP {total/1e6:.1f}M</strong>.'
    if pe_t90 > 0:
        alert += f' Perú +90d: <strong>USD {pe_t90/1000:.1f}K vencidos</strong>.'

    txt = re.sub(
        r'(<div class="alerta-icon" id="av-cxc-alerta-icon">)[^<]*(</div>)',
        rf'\g<1>{icon}\g<2>',
        txt,
    )
    txt = re.sub(
        r'(<div class="alerta-text" id="av-cxc-alerta-text">).*?(</div>)',
        f'\\g<1>\n      {alert}\n    \\g<2>',
        txt,
        flags=re.DOTALL,
    )

    # ── 3. Regenerar peruData (si pe_cxc tiene documentos) ───────────────────
    pe_docs = (pe_cxc or {}).get('all_documentos') or (pe_cxc or {}).get('documentos')
    if pe_docs is not None:
        corte_pe = (pe_cxc or {}).get('corte', '—')

        def _pe_row(d: dict) -> str:
            vend  = (d.get('vendedor') or 'Sin asignar').replace("'", r"\'")
            cli   = (d.get('cliente')  or '').replace("'", r"\'")
            folio = str(d.get('folio', '')      or '').replace("'", r"\'")
            emis  = str(d.get('emision', '')    or '').replace("'", r"\'")
            vcto  = str(d.get('vencimiento', '') or '').replace("'", r"\'")
            dias  = d.get('dias', 0)
            tramo = d.get('tramo', '')
            monto = d.get('monto', 0)
            est   = d.get('estado', 'Normal')
            return (f"  {{vendedor:'{vend}',cliente:'{cli}',folio:'{folio}',"
                    f"emision:'{emis}',vencimiento:'{vcto}',"
                    f"dias:{dias},tramo:'{tramo}',monto:{monto},estado:'{est}'}}")

        pe_rows  = [_pe_row(d) for d in pe_docs]
        pe_body  = ',\n'.join(pe_rows) if pe_rows else '  // Sin documentos'
        new_peru = (f"const peruData = [\n"
                    f"  // Corte {corte_pe} · generado por update_avboard.py\n"
                    f"{pe_body}\n"
                    f"]")
        txt = re.sub(r'const peruData\s*=\s*\[.*?\]', new_peru, txt, flags=re.DOTALL)

    if txt != original:
        cxc_file.write_text(txt, encoding='utf-8')
        return True
    return False


def sync_cache_busting():
    """
    Actualiza el query param ?v= de TODOS los <script src="avboard_data.js...">
    y <script src="avboard_clientes.js..."> en cada panel HTML del repo, usando
    CACHE_V (fecha de esta corrida).

    Por qué existe: cada corrida regenera avboard_data.js / avboard_clientes.js,
    pero si el ?v= del script tag no cambia, el navegador sigue sirviendo la
    copia cacheada de esos archivos y el panel queda "congelado" aunque el dato
    en GitHub ya esté actualizado — los paneles dejan de coincidir entre sí
    según el caché de cada uno. Bug detectado y corregido a mano el 2026-06-23;
    desde ahora se sincroniza automáticamente en cada corrida.

    Solo reemplaza el query string del script tag — NO toca diseño ni estructura.
    """
    patterns = [
        (re.compile(r'(src=["\'])avboard_data\.js(?:\?v=[\w\d]+)?(["\'])'), 'avboard_data.js'),
        (re.compile(r'(src=["\'])avboard_clientes\.js(?:\?v=[\w\d]+)?(["\'])'), 'avboard_clientes.js'),
    ]
    changed = []
    for path in sorted(REPO.glob('*.html')):
        try:
            content = path.read_text(encoding='utf-8')
        except Exception:
            continue
        new_content = content
        for pattern, fname in patterns:
            new_content = pattern.sub(r'\1' + fname + '?v=' + CACHE_V + r'\2', new_content)
        if new_content != content:
            path.write_text(new_content, encoding='utf-8')
            changed.append(path.name)
    return changed


# ══════════════════════════════════════════════════════════════════════════════
#  7. LOGS
# ══════════════════════════════════════════════════════════════════════════════

def write_logs(summary):
    """Actualiza los 3 archivos de log. NUNCA sobreescribe, solo agrega."""
    LOGS.mkdir(exist_ok=True)

    # ── update_log.txt ───────────────────────────────────────────────────────
    log_entry = f"""
{'='*70}
CORTE: {NOW}
ARCHIVOS PROCESADOS:
{chr(10).join('  - ' + f for f in summary['archivos_procesados'])}
ACCIONES:
{chr(10).join('  · ' + a for a in summary['acciones'])}
VALIDACIÓN JS: {'✅ OK' if summary['js_ok'] else '❌ ERRORES'}
{'='*70}
"""
    with open(LOGS / 'update_log.txt', 'a', encoding='utf-8') as f:
        f.write(log_entry)

    # ── resumen_actualizacion.md ─────────────────────────────────────────────
    resumen = f"""
## Actualización {NOW} — Corte {summary['cortes']['chile_ventas']}

**Chile ventas:** CLP {summary['chile_ytd']:,} YTD · Cumpl 4m: {summary['cumpl_4m']:.1%}
**Perú ventas:** USD {summary['peru_ytd']:,} YTD · Cumpl 5m: {summary['cumpl_5m']:.1%}
**CxC Chile:** CLP {summary['cxc_total']:,} total · +90d: CLP {summary['cxc_t90']:,}
**IEC Chile:** {summary['iec_total']:.1%} global

**Alertas CxC:**
{chr(10).join('- ' + c['cliente'] + f" CLP {c['monto']:,} ({c['dias']}d)" for c in summary['cuentas_criticas'] if c.get('estado') == 'CRÍTICO')}

**Módulo Productos (rentabilidad real por SKU):**
- {len(summary['productos']['alertas_nivel1'])} SKU(s) con margen NEGATIVO (destruyen margen) ·
  impacto estimado CLP {summary['productos']['impacto_clp']:,}
- {len(summary['productos']['alertas_nivel2'])} SKU(s) en zona de riesgo (margen 0-10%, subvaluados)
- Sin costo cargado en tabla piso: {summary['productos']['skus_sin_costo_chile']} SKU(s) Chile ·
  {summary['productos']['skus_sin_costo_peru']} SKU(s) Perú (no se puede calcular margen real — completar piso)
- Bajo precio piso propuesto: {summary['productos']['skus_bajo_piso_chile']} SKU(s) Chile ·
  {summary['productos']['skus_bajo_piso_peru']} SKU(s) Perú
{chr(10).join('  - REVISAR: ' + a['sku'] + f" ({a['pais']}) margen {a['margen']:.1%}" for a in summary['productos']['alertas_nivel1'][:10])}

**Decisión sugerida:** priorizar revisión de precio/costo en los SKU con margen
negativo listados arriba; completar costo en tabla piso para los SKU sin costo
cargado (hoy no se puede saber si son rentables). Perú es best-effort —
validar con Javier antes de tomar decisiones de pricing basadas solo en esos
números (ver nota en update_avboard.py / compute_productos).

---
"""
    with open(LOGS / 'resumen_actualizacion.md', 'a', encoding='utf-8') as f:
        f.write(resumen)

    # ── alertas.md ───────────────────────────────────────────────────────────
    crit_list = [c for c in summary['cuentas_criticas'] if c.get('estado') == 'CRÍTICO']
    prod_alertas1 = summary.get('productos', {}).get('alertas_nivel1', [])
    if prod_alertas1:
        prod_alerta_md = (
            f"\n## ⚠ Productos que destruyen margen — {NOW}\n\n"
            "| SKU | País | Margen | Acción |\n"
            "|-----|------|--------|--------|\n"
        )
        for a in prod_alertas1:
            prod_alerta_md += f"| {a['sku']} | {a['pais']} | {a['margen']:.1%} | {a['accion']} |\n"
        prod_alerta_md += (
            f"\nImpacto estimado: CLP {summary['productos']['impacto_clp']:,}. "
            f"SKUs sin costo cargado (no evaluables): "
            f"{summary['productos']['skus_sin_costo_chile']} CL / "
            f"{summary['productos']['skus_sin_costo_peru']} PE.\n---\n"
        )
        with open(LOGS / 'alertas.md', 'a', encoding='utf-8') as f:
            f.write(prod_alerta_md)

    if crit_list:
        alertas = f"""
## ⚠ Alertas Activas — {NOW}

| Cliente | Monto CLP | Días | RTC |
|---------|-----------|------|-----|
"""
        for c in crit_list:
            alertas += f"| {c['cliente']} | {c['monto']:,} | {c.get('dias','?')}d | {c.get('rtc','?')} |\n"

        alertas += f"\nCxC +90d total: CLP {summary['cxc_t90']:,}\n---\n"
        with open(LOGS / 'alertas.md', 'a', encoding='utf-8') as f:
            f.write(alertas)


# ══════════════════════════════════════════════════════════════════════════════
#  8. VALIDACIÓN JS
# ══════════════════════════════════════════════════════════════════════════════

def validate_js(path):
    """Valida sintaxis JS con node --check."""
    try:
        result = subprocess.run(
            ['node', '--check', str(path)],
            capture_output=True, text=True, timeout=10
        )
        return result.returncode == 0, result.stderr.strip()
    except FileNotFoundError:
        return True, '(node not found — skipped)'
    except Exception as e:
        return False, str(e)


# ══════════════════════════════════════════════════════════════════════════════
#  9. EXTRAER CLIENTES_PE EXISTENTE
# ══════════════════════════════════════════════════════════════════════════════

def extract_existing_clientes_pe():
    """Lee CLIENTES_PE del archivo actual para preservarlo."""
    try:
        with open(AVBOARD_CLI, 'r', encoding='utf-8') as f:
            content = f.read()
        m = re.search(r'const CLIENTES_PE = (\[.*?\]);', content, re.DOTALL)
        if m:
            return m.group(1)
    except Exception:
        pass
    return '[]'


# ══════════════════════════════════════════════════════════════════════════════
#  CxC PERÚ — PARSER CANÓNICO (T1b)
#  Lee el vigente desde raw_registry.json, parsea Excel, normaliza vendedores
#  via vendors.json, aplica fusiones GG, retorna objeto estructurado.
# ══════════════════════════════════════════════════════════════════════════════

_vendors_pe_lookup_cache = None

def _load_vendors_pe_lookup() -> dict:
    """Builds UPPERCASE_ALIAS → vendor_id lookup for PE vendors from vendors.json."""
    global _vendors_pe_lookup_cache
    if _vendors_pe_lookup_cache is not None:
        return _vendors_pe_lookup_cache
    vendors_path = ROOT / "pipeline" / "vendors.json"
    lookup: dict = {}
    if vendors_path.exists():
        data = json.loads(vendors_path.read_text(encoding="utf-8"))
        for v in data.get("vendors", []):
            if v.get("pais") == "PE":
                vid = v["vendor_id"]
                for alias in v.get("aliases", []):
                    lookup[alias.strip().upper()] = vid
                lookup[vid.upper()] = vid  # self-map
    _vendors_pe_lookup_cache = lookup
    return lookup


def _normalize_rtc_pe(name: str):
    """Maps raw VENDEDOR string → canonical vendor_id.
    Returns None for empty (supra), 'sin_asignar' for unmapped names."""
    if not name or not str(name).strip() or str(name).strip().upper() == 'NAN':
        return None  # caller handles as supra
    normalized = str(name).strip().upper()
    lookup = _load_vendors_pe_lookup()
    return lookup.get(normalized, "sin_asignar")


def _apply_cxc_gg_fusions(vendor_id: str) -> str:
    """Applies GG fusion decisions to a CxC vendor_id at parse time.
    Reads gg_decisions.json (same SSOT as ventas pipeline)."""
    for dec in _GG_DECISIONS:
        if dec.get("tipo") == "fusion_vendedor":
            src = dec.get("criterio", {}).get("vendor_id")
            tgt = dec.get("accion", {}).get("fused_into")
            if vendor_id == src and tgt:
                return tgt
    return vendor_id


def parse_cxc_pe(path: Path, fecha_corte: str) -> dict:
    """
    Parsea CxC Perú desde Excel.
    Sheet: "AGROVECA" · header row 6 (index 5)
    Columns: CODIGO, NOMBRE, TD, SER., NUMERO, FECHA, VENCIM, VENCIM(mes), SALDO, GLOSA, VENDEDOR

    Retorna objeto estructurado compatible con peru_cxc JS shape + clave 'documentos'
    (lista completa para regenerar peruData en Panel_CxC).

    Si falla la lectura, hace fallback a extract_peru_cxc_static() con advertencia.
    """
    try:
        df = pd.read_excel(path, sheet_name="AGROVECA", header=5)
    except Exception as e:
        print(f"   ❌ parse_cxc_pe: no se pudo leer {path.name}: {e}")
        print(f"   ⚠ Usando datos estáticos como fallback")
        return extract_peru_cxc_static()

    # Parse fecha_corte DD/MM/YYYY → date
    try:
        fc_date = datetime.strptime(fecha_corte, "%d/%m/%Y").date()
    except Exception:
        from datetime import date as _date
        fc_date = _date.today()

    # ── Filtrar registros válidos ─────────────────────────────────────────────
    df = df[df['SALDO'].notna() & (pd.to_numeric(df['SALDO'], errors='coerce') > 0)].copy()
    df = df[df['CODIGO'].notna()].copy()
    df = df[~df['CODIGO'].astype(str).str.strip().str.startswith('INGRESOS')].copy()
    df = df[df['CODIGO'].astype(str).str.strip().str.upper() != 'NAN'].copy()
    df = df.reset_index(drop=True)

    documentos  = []
    supra_docs  = []
    por_vendedor: dict = {}
    supra_total = 0.0

    for _, row in df.iterrows():
        saldo = float(row.get('SALDO', 0) or 0)
        if saldo <= 0:
            continue

        nombre_cliente = str(row.get('NOMBRE', '') or '').strip()
        numero         = str(row.get('NUMERO', '') or '').strip()
        vendedor_raw   = str(row.get('VENDEDOR', '') or '').strip()

        # Normalizar fechas
        try:
            fecha_emision = pd.to_datetime(row.get('FECHA')).date()
            emision_str   = fecha_emision.strftime('%d/%m/%Y')
        except Exception:
            emision_str   = ''

        try:
            fecha_vencim  = pd.to_datetime(row.get('VENCIM')).date()
            vencim_str    = fecha_vencim.strftime('%d/%m/%Y')
            dias_mora     = (fc_date - fecha_vencim).days
        except Exception:
            vencim_str    = ''
            dias_mora     = 0

        # Clasificar tramo
        if dias_mora > 90:
            tramo, estado = '+90d', 'Crítico'
        elif dias_mora > 60:
            tramo, estado = '61-90d', 'Riesgo'
        elif dias_mora > 30:
            tramo, estado = '31-60d', 'Alerta'
        elif dias_mora > 0:
            tramo, estado = '0-30d', 'Normal'
        else:
            tramo, estado = 'Al día', 'Al día'

        # Normalizar vendedor → vendor_id canónico
        vid_raw = _normalize_rtc_pe(vendedor_raw)
        if vid_raw is not None:
            vid = _apply_cxc_gg_fusions(vid_raw)  # aplica fusiones GG (p.ej. navarro→aguirre)
        else:
            vid = None  # supra

        doc = {
            "vendedor":    vendedor_raw if vendedor_raw else "Sin asignar",
            "cliente":     nombre_cliente,
            "folio":       numero,
            "emision":     emision_str,
            "vencimiento": vencim_str,
            "dias":        dias_mora,
            "tramo":       tramo,
            "monto":       round(saldo, 2),
            "estado":      estado,
        }

        if vid is None:
            # Supra: VENDEDOR vacío — cartera no asignada a vendedor
            supra_total += saldo
            supra_docs.append(doc)
        else:
            documentos.append(doc)
            if vid not in por_vendedor:
                por_vendedor[vid] = {"total": 0.0, "vencida": 0.0, "t90": 0.0}
            por_vendedor[vid]["total"]   += saldo
            por_vendedor[vid]["vencida"] += saldo if dias_mora > 0 else 0.0
            por_vendedor[vid]["t90"]     += saldo if dias_mora > 90 else 0.0

    # ── Totales ───────────────────────────────────────────────────────────────
    total         = sum(v["total"]   for v in por_vendedor.values())
    total_vencida = sum(v["vencida"] for v in por_vendedor.values())

    # ── Tramos desde documentos (excl. supra) ────────────────────────────────
    t_acc = {"no_vencida": 0.0, "t030": 0.0, "t3160": 0.0, "t6190": 0.0, "t90": 0.0}
    for d in documentos:
        m, dd = d["monto"], d["dias"]
        if dd > 90:   t_acc["t90"]       += m
        elif dd > 60: t_acc["t6190"]     += m
        elif dd > 30: t_acc["t3160"]     += m
        elif dd > 0:  t_acc["t030"]      += m
        else:         t_acc["no_vencida"] += m

    tramos     = {k: round(v) for k, v in t_acc.items()}
    tramos_pct = {k: round(v / total, 3) if total > 0 else 0 for k, v in t_acc.items()}

    # ── Por vendedor: calcular pct y riesgo ───────────────────────────────────
    por_vendedor_out = {}
    for vid, agg in por_vendedor.items():
        t  = round(agg["total"])
        v  = round(agg["vencida"])
        t9 = round(agg["t90"])
        por_vendedor_out[vid] = {
            "total":   t,
            "pct":     round(agg["total"] / total, 4) if total > 0 else 0,
            "vencida": v,
            "t90":     t9,
            "riesgo":  "CRÍTICO" if t9 > 0 else ("RIESGO" if v > t * 0.3 else "NORMAL"),
        }

    # ── Cuentas críticas: docs con dias > 60, top 15 por monto ───────────────
    cuentas_criticas = sorted(
        [d for d in documentos if d["dias"] > 60],
        key=lambda x: -x["monto"]
    )[:15]

    # ── Todos los docs ordenados para tabla (dias desc, monto desc) ───────────
    all_docs_sorted = sorted(
        documentos + supra_docs,
        key=lambda x: (-x["dias"], -x["monto"])
    )

    return {
        "corte":            fecha_corte,
        "total":            round(total),
        "supra":            round(supra_total),
        "n_documentos":     len(documentos),
        "tramos":           tramos,
        "tramos_pct":       tramos_pct,
        "vencida":          round(total_vencida),
        "por_vendedor":     por_vendedor_out,
        "cuentas_criticas": cuentas_criticas,
        "documentos":       documentos,          # excl. supra (alineado con tramos/por_vendedor)
        "all_documentos":   all_docs_sorted,     # incl. supra (para tabla Panel_CxC)
    }


def extract_peru_cxc_static():
    """Lee el objeto peru_cxc del avboard_data.js actual para preservarlo."""
    # Default static value (unchanged between cortes unless new Peru CxC arrives)
    return {
        "corte": "10/05/2026",
        "total": 117964,
        "supra": 196841,
        "tramos": {"no_vencida": 79300, "t030": 10534, "t3160": 3149, "t6190": 1360, "t90": 23621},
        "tramos_pct": {"no_vencida": 0.672, "t030": 0.089, "t3160": 0.027, "t6190": 0.012, "t90": 0.200},
        "vencida": 38664,
        "por_vendedor": {
            "infante": {"total": 28153, "pct": 0.239, "vencida": 15881, "t90": 4598, "riesgo": "CRÍTICO"},
            "geldres": {"total": 10874, "pct": 0.092, "vencida": 10874, "t90": 10874, "riesgo": "CRÍTICO"},
            "atalaya": {"total": 15343, "pct": 0.130, "vencida": 15343, "t90": 15343, "riesgo": "CRÍTICO"},
            "aguirre_navarro": {"total": 58942, "pct": 0.499, "vencida": 1432, "t90": 0, "riesgo": "RIESGO"},
            "gonzales_valladares": {"total": 1600, "pct": 0.014, "vencida": 800, "t90": 0, "riesgo": "NORMAL"},
            "pradenas_sin_asignar": {"total": 7030, "pct": 0.060, "vencida": 4830, "t90": 0, "riesgo": "NORMAL"},
        },
        "cuentas_criticas": [
            {"cliente": "PAODISA S.A.", "vendedor": "J. Geldres", "dias": "468-648d", "monto": 10874.40, "estado": "CRÍTICO", "nota": "4 facturas 2024 · proceso legal pendiente"},
            {"cliente": "AGROFER MJ E.I.R.L.", "vendedor": "O. Atalaya", "dias": 211, "monto": 9493.00, "estado": "CRÍTICO", "nota": "acuerdo de pago urgente"},
            {"cliente": "LUNA QUINTANILLA BRYAN ALEXANDER", "vendedor": "O. Infante", "dias": "97-143d", "monto": 2349.05, "estado": "CRÍTICO", "nota": "2 folios"},
            {"cliente": "EPIC FARMS S.A.C.", "vendedor": "A. Gonzalez", "dias": 107, "monto": 600.00, "estado": "CRÍTICO", "nota": "escaló desde 61-90d"},
        ]
    }


# ══════════════════════════════════════════════════════════════════════════════
#  SCHEMA VALIDATION — V-04: errores explícitos si columnas críticas faltan
# ══════════════════════════════════════════════════════════════════════════════

_SCHEMA_DEFS_PATH = REPO / "pipeline" / "schema_defs.json"
_schema_defs_cache = None


def _load_schema_defs():
    global _schema_defs_cache
    if _schema_defs_cache is not None:
        return _schema_defs_cache
    if _SCHEMA_DEFS_PATH.exists():
        with open(_SCHEMA_DEFS_PATH, encoding="utf-8") as f:
            _schema_defs_cache = json.load(f).get("types", {})
    else:
        _schema_defs_cache = {}
    return _schema_defs_cache


def validate_schema_before_parse(filepath: Path, tipo: str) -> bool:
    """
    Valida que el archivo Excel tenga las sheets y columnas requeridas para el tipo.
    Retorna True si pasa. Si falla, imprime ERROR_SCHEMA detallado y retorna False.
    NO sustituye columnas faltantes por 0 silenciosamente (V-04).
    """
    schema = _load_schema_defs().get(tipo)
    if not schema:
        return True  # sin schema definido → pasar sin validar

    import openpyxl as _oxl
    try:
        wb = _oxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception as e:
        print(f"   ❌ ERROR_SCHEMA [{tipo}] No se pudo abrir {filepath.name}: {e}")
        return False

    sheet_names = wb.sheetnames

    # Validar sheet requerida
    required_sheet = schema.get("sheet")
    if required_sheet:
        match = next((s for s in sheet_names if s.upper() == required_sheet.upper()), None)
        if not match:
            print(f"   ❌ ERROR_SCHEMA [{tipo}] {filepath.name}")
            print(f"      Sheet requerida: '{required_sheet}'")
            print(f"      Sheets encontradas: {sheet_names}")
            wb.close()
            return False

        # Validar columnas requeridas (leer header row)
        required_cols = schema.get("required_columns", [])
        if required_cols and schema.get("header_row") != "dynamic":
            ws = wb[match]
            hrow = schema.get("header_row", 0)
            headers = []
            for row in ws.iter_rows(min_row=hrow + 1, max_row=hrow + 1, values_only=True):
                headers = [str(c).strip() if c else "" for c in row]
                break

            aliases = schema.get("column_aliases", {})
            found = set()
            missing = []
            for req in required_cols:
                # Buscar col exacta o alias
                all_names = [req] + aliases.get(req, [])
                if any(n in headers for n in all_names):
                    found.add(req)
                else:
                    missing.append(req)

            if missing:
                print(f"   ❌ ERROR_SCHEMA [{tipo}] {filepath.name}")
                print(f"      Columnas faltantes: {missing}")
                print(f"      Columnas encontradas: {[h for h in headers if h][:20]}")
                print(f"      → Revisar que el archivo tenga el formato correcto.")
                wb.close()
                return False

    # Validar sheets requeridas (LIBRO_BASE)
    required_sheets = schema.get("required_sheets", [])
    for rs in required_sheets:
        if not any(s.upper() == rs.upper() for s in sheet_names):
            print(f"   ❌ ERROR_SCHEMA [{tipo}] {filepath.name}")
            print(f"      Sheet requerida '{rs}' no encontrada en {sheet_names}")
            wb.close()
            return False

    wb.close()
    return True


# ══════════════════════════════════════════════════════════════════════════════
#  10. AI INSIGHTS CEO
# ══════════════════════════════════════════════════════════════════════════════

def generate_ai_insights(cl_v: dict, pe_v: dict, cl_cxc: dict,
                         pe_cxc: dict | None, cortes: dict) -> dict | None:
    """Genera diagnóstico ejecutivo con visión CEO via Anthropic API.
    Retorna dict {peru, chile, grupo, generated_at} o None si no hay API key."""
    import os, json as _json
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        print("   [AI insights] ⚠  ANTHROPIC_API_KEY no encontrada — skipping")
        return None
    try:
        import anthropic as _ant
    except ImportError:
        print("   [AI insights] ⚠  pip install anthropic --break-system-packages  — skipping")
        return None

    # ── Construcción del contexto de datos ────────────────────────────────────
    pe_ytd   = pe_v.get('ytd_5m', 0)
    pe_ppto  = pe_v.get('ppto_5m', 0)
    pe_cumpl = pe_v.get('cumplimiento_5m', 0)
    pe_nM    = len([v for v in pe_v.get('mensual_real', []) if v > 0])
    pe_corte = cortes.get('peru_ventas', '—')
    pe_pv    = pe_v.get('por_vendedor', {})
    rtcR_pe  = pe_v.get('rtc_mensual_real', {})
    pe_lines = []
    for k, v in sorted(pe_pv.items(), key=lambda x: -x[1].get('ytd', 0)):
        ytd = v.get('ytd', 0)
        arr = rtcR_pe.get(k, [])
        mes = arr[pe_nM - 1] if arr and pe_nM <= len(arr) else 0
        pe_lines.append(f"  · {v.get('nombre', k)}: YTD USD {ytd:,.0f} · {pe_corte[:5]} USD {mes:,.0f}")
    pe_cxc_tot = (pe_cxc or {}).get('total', 0)
    pe_cxc_t90 = (pe_cxc or {}).get('tramos', {}).get('t90', 0)
    pe_cxc_cte = (pe_cxc or {}).get('corte', '—')

    cl_ytd   = cl_v.get('ytd_5m', 0)
    cl_ppto  = cl_v.get('ppto_5m', 0)
    cl_cumpl = cl_v.get('cumplimiento_5m', 0)
    cl_nM    = len([v for v in cl_v.get('mensual_real', []) if v > 0])
    cl_corte = cortes.get('chile_ventas', '—')
    rtcR_cl  = cl_v.get('rtc_mensual_real', {})
    cl_lines = []
    for k, arr in sorted(rtcR_cl.items(), key=lambda x: -sum(x[1][:cl_nM] if x[1] else [])):
        ytd_cl = sum((arr or [])[:cl_nM])
        mes_cl = (arr or [])[cl_nM - 1] if arr and cl_nM <= len(arr) else 0
        cl_lines.append(f"  · {k}: YTD CLP {ytd_cl/1e6:.1f}M · {cl_corte[:5]} CLP {mes_cl/1e6:.1f}M")
    cl_cxc_tot = cl_cxc.get('total', 0)
    cl_cxc_t90 = cl_cxc.get('tramos', {}).get('t90', 0)

    prompt = f"""Eres el Gerente General de AV LATAM, empresa de agroinsumos con operaciones en Chile (AGROCOMERCIAL) y Perú (AGROVECA). Analiza los datos comerciales del período más reciente y entrega un diagnóstico ejecutivo accionable.

PERÚ — Corte {pe_corte}:
• YTD USD {pe_ytd:,.0f} vs Ppto USD {pe_ppto:,.0f} → Cumplimiento {pe_cumpl*100:.1f}%  ({pe_nM} meses)
• Por vendedor (YTD · último mes):
{chr(10).join(pe_lines)}
• CxC total USD {pe_cxc_tot:,.0f} · Mora +90d USD {pe_cxc_t90:,.0f} · Corte CxC {pe_cxc_cte}

CHILE — Corte {cl_corte}:
• YTD CLP {cl_ytd/1e6:.1f}M vs Ppto CLP {cl_ppto/1e6:.1f}M → Cumplimiento {cl_cumpl*100:.1f}%  ({cl_nM} meses)
• Por RTC (YTD · último mes):
{chr(10).join(cl_lines[:6])}
• CxC total CLP {cl_cxc_tot/1e6:.1f}M · Mora +90d CLP {cl_cxc_t90/1e6:.1f}M

Responde ÚNICAMENTE con JSON válido, sin texto adicional ni markdown:
{{
  "peru": {{
    "diagnostico": "2-3 párrafos ejecutivos: situación general, performance por vendedor, riesgos de cartera y oportunidades clave",
    "acciones": ["Acción concreta 1", "Acción concreta 2", "Acción concreta 3", "Acción concreta 4"]
  }},
  "chile": {{
    "diagnostico": "2-3 párrafos ejecutivos: situación general, performance por RTC, riesgos de cartera y oportunidades clave",
    "acciones": ["Acción concreta 1", "Acción concreta 2", "Acción concreta 3", "Acción concreta 4"]
  }},
  "grupo": {{
    "diagnostico": "1-2 párrafos: perspectiva consolidada LATAM, prioridades estratégicas del trimestre",
    "acciones": ["Prioridad estratégica 1", "Prioridad estratégica 2", "Prioridad estratégica 3"]
  }}
}}"""

    try:
        client = _ant.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=2500,
            messages=[{'role': 'user', 'content': prompt}]
        )
        raw = msg.content[0].text.strip()
        # Limpiar posible markdown code fence
        if raw.startswith('```'):
            raw = raw.split('\n', 1)[-1].rsplit('```', 1)[0].strip()
        data = _json.loads(raw)
        from datetime import datetime
        data['generated_at'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        data['cortes'] = {'peru': pe_corte, 'chile': cl_corte}
        print(f"   [AI insights] ✅  Diagnóstico generado · {len(raw)} chars · model haiku-4-5")
        return data
    except Exception as e:
        print(f"   [AI insights] ❌  Error: {e}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  11. MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print(f"\n{'='*60}")
    print(f"  AVBOARD UPDATE · {NOW}")
    print(f"{'='*60}\n")

    # 1. Ejecutar inbox_detector (SSOT) → genera raw_registry.json fresco
    print("📂 Ejecutando inbox_detector (SSOT)...")
    try:
        import inbox_detector as _inbox_det
        _inbox_det.run()
    except Exception as _det_err:
        print(f"   ⚠ inbox_detector falló: {_det_err} — intentando con registry existente")

    # 1b. Cargar archivos desde registry (única fuente de verdad — V-02)
    print("📂 Cargando vigentes desde raw_registry.json...")
    files = load_files_from_registry()
    for k, v in files.items():
        print(f"   {k}: {v.name}")

    required = ['chile_ventas', 'peru_ventas', 'cxc_agro', 'cxc_avch', 'piso_chile']
    missing  = [r for r in required if r not in files]
    if missing:
        print(f"\n❌ Faltan archivos: {missing}")
        sys.exit(1)

    # 2. Extraer cortes
    # Lee fecha_corte canónica desde raw_registry.archivos (authoritative SSOT)
    # para evitar fallos con nombres de archivo con caracteres dobles (p.ej. "24..08.2026")
    _rdata = json.loads((REPO / "pipeline" / "raw_registry.json").read_text(encoding="utf-8"))
    _rarch_raw = _rdata.get("archivos", [])
    # archivos puede ser lista o dict según versión del registry
    _rarch_list = _rarch_raw if isinstance(_rarch_raw, list) else list(_rarch_raw.values())
    _reg_fc: dict = {}
    for _ae in _rarch_list:
        if _ae.get("estado") == "VIGENTE":
            _tk = f"{_ae.get('tipo_archivo','')}|{_ae.get('empresa_id','')}"
            _iso = _ae.get("fecha_corte", "")
            if _iso:
                try:
                    _reg_fc[_tk] = datetime.strptime(_iso, "%Y-%m-%d").strftime("%d/%m/%Y")
                except Exception:
                    pass

    def _corte(pipeline_key: str, registry_type_key: str) -> str:
        """Returns fecha_corte: registry SSOT first, filename fallback second."""
        return _reg_fc.get(registry_type_key) or extract_date_from_filename(files[pipeline_key])

    cortes = {
        'chile_ventas': _corte('chile_ventas', 'VENTAS_CL|AGROCOMERCIAL_CL'),
        'chile_cxc':    _corte('cxc_agro',     'CXC_CL_AGROCOMERCIAL|AGROCOMERCIAL_CL'),
        'peru_ventas':  _corte('peru_ventas',   'VENTAS_PE|AGROVECA_PE'),
        'peru_cxc':     _reg_fc.get('CXC_PE|AGROVECA_PE', '—') if 'cxc_pe' in files else '—',
    }
    print(f"\n📅 Cortes detectados: {cortes}")

    # 2.5 Validar schema antes de procesar (V-04)
    print("\n🔍 Validando schemas antes de extraer...")
    _schema_checks = [
        ('chile_ventas',  'VENTAS_CL'),
        ('peru_ventas',   'VENTAS_PE'),
        ('libro_base',    'LIBRO_BASE'),
        ('cxc_agro',      'CXC_CL_AGROCOMERCIAL'),
        ('cxc_avch',      'CXC_CL_AGROVECA'),
    ]
    if 'cxc_pe' in files:
        _schema_checks.append(('cxc_pe', 'CXC_PE'))
    _schema_errors = []
    for _fkey, _tipo in _schema_checks:
        if _fkey in files:
            _ok = validate_schema_before_parse(files[_fkey], _tipo)
            if _ok:
                print(f"   ✅ Schema OK — {_tipo} ({files[_fkey].name})")
            else:
                _schema_errors.append(_fkey)
    if _schema_errors:
        _required_with_error = [e for e in _schema_errors if e in ['chile_ventas', 'peru_ventas']]
        if _required_with_error:
            print(f"\n❌ ERROR_SCHEMA en archivos requeridos: {_required_with_error}")
            print("   Corregir los archivos antes de continuar.")
            sys.exit(1)
        else:
            print(f"   ⚠ Schema inválido en archivos opcionales: {_schema_errors} — continuando con best-effort")

    # 3. Extraer datos
    print("\n📊 Extrayendo datos...")
    print("   Chile ventas...")
    cl_v = extract_chile_ventas(files['chile_ventas'])
    print(f"   → YTD: CLP {cl_v['ytd_5m']:,} | {sum(1 for v in cl_v['mensual'] if v>0)} meses")

    print("   Perú ventas...")
    pe_v = extract_peru_ventas(files['peru_ventas'])
    print(f"   → YTD bruto: USD {pe_v['ytd_5m']:,}")

    # Aplicar decisiones GG de atribución de vendedores Perú
    print("   Aplicando decisiones GG (folio 926 + normalización NICOLL/LIZBETH)...")
    pe_v = _apply_peru_vendor_gg_decisions(pe_v)
    print(f"   → YTD post-GG: USD {pe_v['ytd_5m']:,}")

    print("   CxC Chile (2 entidades)...")
    cl_cxc = extract_cxc_chile(files['cxc_agro'], files['cxc_avch'])
    print(f"   → Total: CLP {cl_cxc['total']:,} | +90d: CLP {cl_cxc['tramos']['t90']:,}")

    # 4. Calcular IEC
    print("\n⚡ Calculando IEC Chile...")
    piso_dict = load_piso_chile(files['piso_chile'])
    iec_cl    = compute_iec_chile(cl_v['df'], piso_dict)
    print(f"   → IEC global: {iec_cl['total']:.1%} | BP: CLP {iec_cl['bp_total']:,}")

    # 4.5 Calcular módulo Productos + IEC Perú (Fase 7)
    print("\n💰 Calculando módulo Productos...")
    if 'piso_peru' in files:
        piso_pe_dict = load_piso_peru(files['piso_peru'])
        pe_sku_df    = extract_peru_ventas_sku(files['peru_ventas'])
    else:
        piso_pe_dict = {'entries': {}, 'tiers': {}}
        pe_sku_df    = pd.DataFrame(columns=['CONCEPTO', 'DOLARES'])
        print("   ⚠ Libro Base no encontrado — módulo Productos Perú omitido (best-effort)")
    productos = compute_productos(cl_v['df'], piso_dict, pe_sku_df, piso_pe_dict)
    n_neg = len(productos['resumen']['alertas_nivel1'])
    n_sin_costo = productos['resumen']['skus_sin_costo_chile'] + productos['resumen']['skus_sin_costo_peru']
    print(f"   → {len(productos['detalle'])} SKUs · {n_neg} con margen negativo · {n_sin_costo} sin costo cargado")

    # 4.6 Calcular IEC Perú (Fase 7) — requiere TX_PE construido desde ventas Perú
    print("\n⚡ Calculando IEC Perú (Fase 7)...")
    if 'peru_ventas' in files and piso_pe_dict.get('entries'):
        _pe_sku_pre = extract_peru_ventas_sku(files['peru_ventas'])
        _tx_pe_pre  = build_tx_pe(_pe_sku_pre, piso_pe_dict)
        iec_pe_computed = compute_iec_peru(_tx_pe_pre)
        print(f"   → IEC Perú total: {iec_pe_computed['total']:.1%}" if iec_pe_computed.get('total') else "   → IEC Perú: null (sin elegibles)")
        for vk in ['aguirre', 'infante', 'atalaya', 'valladares', 'gonzales', 'navarro', 'diaz']:
            vv = iec_pe_computed.get(vk)
            if vv is not None:
                print(f"   → {vk:15}: {vv:.1%}")
    else:
        _tx_pe_pre = []
        iec_pe_computed = None
        print("   ⚠ Sin datos Perú — IEC Perú será null")

    # 5. Parsear CxC Perú y generar avboard_data.js
    print("\n📝 Parseando CxC Perú...")
    if 'cxc_pe' in files:
        pe_cxc = parse_cxc_pe(files['cxc_pe'], cortes['peru_cxc'])
        print(f"   → {pe_cxc['n_documentos']} docs · USD {pe_cxc['total']:,} · supra {pe_cxc['supra']:,} · corte {pe_cxc['corte']}")
        print(f"   → por_vendedor keys: {list(pe_cxc['por_vendedor'].keys())}")
    else:
        pe_cxc = extract_peru_cxc_static()
        print("   ⚠ Sin CxC PE en registry — usando datos estáticos (corte congelado)")

    print("\n🤖 Generando AI insights (visión CEO/GM)...")
    ai_insights = generate_ai_insights(cl_v, pe_v, cl_cxc, pe_cxc, cortes)

    print("\n📝 Generando avboard_data.js...")
    js_data = render_avboard_data_js(cl_v, pe_v, cl_cxc, iec_cl, cortes, pe_cxc, productos,
                                     iec_pe=iec_pe_computed, insights=ai_insights)
    with open(AVBOARD_DATA, 'w', encoding='utf-8') as f:
        f.write(js_data)
    ok, err = validate_js(AVBOARD_DATA)
    print(f"   → {'✅ OK' if ok else '❌ ERROR: ' + err}")
    if not ok:
        print(f"     {err}")
        sys.exit(1)

    # 6. Verificar integridad con node
    result = subprocess.run(['node', '-e', open(AVBOARD_DATA).read()],
                            capture_output=True, text=True, timeout=10)
    if result.returncode != 0:
        print(f"   ❌ Runtime error:\n{result.stderr}")
        sys.exit(1)
    print(f"   {result.stdout.strip()}")

    # 7. Generar avboard_clientes.js
    print("\n👥 Generando avboard_clientes.js...")
    clientes_pe_raw = extract_existing_clientes_pe()
    clientes_cl, _ = build_clientes(cl_v, pe_v, iec_cl, cl_cxc)
    print(f"   → {len(clientes_cl)} clientes Chile procesados")

    js_cli = render_avboard_clientes_js(clientes_cl, clientes_pe_raw, cortes, len(cl_v['df']))
    with open(AVBOARD_CLI, 'w', encoding='utf-8') as f:
        f.write(js_cli)
    ok2, err2 = validate_js(AVBOARD_CLI)
    print(f"   → {'✅ OK' if ok2 else '❌ ERROR: ' + err2}")

    # 8. Actualizar Panel_IEC TX_CL + TX_PE
    print("\n🔬 Actualizando Panel_IEC TX_CL...")
    tx_cl = build_tx_cl(cl_v['df'], piso_dict)
    n_tx  = update_panel_iec(tx_cl, cortes['chile_ventas'])
    print(f"   → {n_tx} transacciones · corte {cortes['chile_ventas']}")

    # 8.1 TX_PE — reusar _tx_pe_pre calculado en paso 4.6 (evita doble extracción)
    print("\n🔬 Actualizando Panel_IEC TX_PE...")
    if _tx_pe_pre:
        n_tx_pe = update_panel_iec_pe(_tx_pe_pre, cortes['peru_ventas'])
        print(f"   → TX_PE actualizado · {n_tx_pe} transacciones · corte {cortes['peru_ventas']}")
    else:
        n_tx_pe = 0
        print("   ⚠ TX_PE no disponible — Panel IEC Perú no actualizado")

    # 8.2 Generar sic_tx_pe.js (C-04): mismo pipeline que avboard_data.js
    print("\n📦 Generando sic_tx_pe.js...")
    if _tx_pe_pre:
        write_sic_tx_pe(_tx_pe_pre, PPTO_RTC_MENSUAL_PE, cortes.get('peru_ventas', ''))
    else:
        print("   ⚠ TX_PE vacío — sic_tx_pe.js no actualizado")

    # 8.5 Sincronizar cache-busting en todos los paneles
    print(f"\n🔄 Sincronizando cache-busting (?v={CACHE_V}) en paneles HTML...")
    panels_synced = sync_cache_busting()
    print(f"   → {len(panels_synced)} paneles actualizados: {', '.join(panels_synced) if panels_synced else '(ninguno — ya estaban al día)'}")

    # 8.6 Sincronizar IEC en paneles hardcodeados
    iec_cl_total = iec_cl.get('total', 0)
    iec_pe_total = (iec_pe_computed or {}).get('total', 0) or 0
    print(f"\n📊 Sincronizando IEC en paneles hardcodeados...")
    _n_iec = sync_iec_hardcoded(iec_cl_total, iec_pe_total)
    print(f"   → IEC CL {iec_cl_total*100:.1f}% · PE {iec_pe_total*100:.1f}% escritos en {_n_iec} paneles")

    # 8.7 Sincronizar Panel_CxC (chileData + alerta banner)
    print(f"\n💳 Sincronizando Panel_CxC (chileData + alerta banner)...")
    _cxc_ok = sync_cxc_panel(cl_cxc, pe_cxc)
    print(f"   → {'Actualizado' if _cxc_ok else 'Sin cambios (ya al día)'}")

    # 9. Logs
    print("\n📋 Actualizando logs...")
    archivos = [v.name for v in files.values()]
    acciones = [
        f"avboard_data.js generado · corte {cortes['chile_ventas']}",
        f"avboard_clientes.js regenerado · {len(clientes_cl)} clientes CL",
        f"Panel_IEC TX_CL · {n_tx} transacciones",
        f"Panel_IEC TX_PE · {n_tx_pe} transacciones",
        f"sic_tx_pe.js generado · {n_tx_pe} tx · corte {cortes.get('peru_ventas','')}",
        f"IEC Chile {iec_cl['total']:.1%} · BP CLP {iec_cl['bp_total']:,}",
        f"IEC Perú {iec_pe_computed['total']:.1%} (Fase 7)" if iec_pe_computed and iec_pe_computed.get('total') else "IEC Perú: sin datos",
        f"CxC Chile CLP {cl_cxc['total']:,} · +90d CLP {cl_cxc['tramos']['t90']:,}",
        f"Módulo Productos: {len(productos['detalle'])} SKUs · {n_neg} margen negativo · {n_sin_costo} sin costo",
        f"Cache-busting ?v={CACHE_V} sincronizado en {len(panels_synced)} paneles HTML",
    ]
    summary = {
        'archivos_procesados': archivos,
        'acciones': acciones,
        'js_ok': ok and ok2,
        'cortes': cortes,
        'chile_ytd': cl_v['ytd_5m'],
        'peru_ytd': pe_v['ytd_5m'],
        'cumpl_4m': cl_v['cumplimiento_4m'],
        'cumpl_5m': pe_v['cumplimiento_5m'],
        'cxc_total': cl_cxc['total'],
        'cxc_t90': cl_cxc['tramos']['t90'],
        'iec_total': iec_cl['total'],
        'cuentas_criticas': cl_cxc['cuentas_criticas'],
        'productos': productos['resumen'],
    }
    write_logs(summary)

    print(f"\n{'='*60}")
    print(f"  ✅ ACTUALIZACIÓN COMPLETA · {datetime.now().strftime('%H:%M:%S')}")
    print(f"  Chile YTD: CLP {cl_v['ytd_5m']:>15,}  Cumpl 4m: {cl_v['cumplimiento_4m']:.1%}")
    print(f"  Perú  YTD: USD {pe_v['ytd_5m']:>15,}  Cumpl 5m: {pe_v['cumplimiento_5m']:.1%}")
    print(f"  CxC +90d:  CLP {cl_cxc['tramos']['t90']:>15,}")
    print(f"  IEC Chile: {iec_cl['total']:.1%}")
    print(f"{'='*60}\n")

    # 10. Certificación automática (bloquea commit si hay errores críticos)
    print("\n🔐 Ejecutando certificación automática (certify_avboard.py)...")
    import subprocess as _sp
    cert_result = _sp.run(
        ['python3', str(REPO / 'scripts' / 'certify_avboard.py')],
        capture_output=False
    )
    if cert_result.returncode != 0:
        print("\n❌ CERTIFICACIÓN FALLÓ — revisar errores arriba antes de hacer push.")
    else:
        print("✅ Certificación PASS — seguro para commit y push.")


if __name__ == '__main__':
    main()
