"""
inbox_detector.py — AV LATAM Pipeline · Stage 1: Detector + Clasificador
===========================================================================
Escanea /inbox, clasifica cada archivo, detecta duplicados y conflictos,
y selecciona el archivo vigente por tipo+empresa. Produce raw_registry.json.

Uso:
    python scripts/inbox_detector.py

Salida:
    pipeline/raw_registry.json

Reglas de clasificación: ver SSOT_AV_LATAM_v1.0.md + VERIFICACION_PRE_SPRINT1_2026-08-13.md
Versión: 1.0.0 — 2026-08-13
"""

import os
import re
import json
import hashlib
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

# Openpyxl — solo para la sonda de estructura (detección por contenido, no por nombre)
try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# CONSTANTES
# ---------------------------------------------------------------------------

PIPELINE_VERSION = "1.0.0"

# Directorio base del repo (el script vive en scripts/)
REPO_ROOT = Path(__file__).parent.parent
INBOX_PATH = REPO_ROOT / "inbox"
OUTPUT_PATH = REPO_ROOT / "pipeline" / "raw_registry.json"


# ---------------------------------------------------------------------------
# CATÁLOGO DE TIPOS
# ---------------------------------------------------------------------------

TIPOS = {
    "VENTAS_CL":             {"empresa_id": "AGROCOMERCIAL_CL", "pais_id": "CL", "moneda": "CLP"},
    "CXC_CL_AGROVECA":       {"empresa_id": "AGROVECA_CL",      "pais_id": "CL", "moneda": "CLP"},
    "CXC_CL_AGROCOMERCIAL":  {"empresa_id": "AGROCOMERCIAL_CL", "pais_id": "CL", "moneda": "CLP"},
    "VENTAS_PE":             {"empresa_id": "AGROVECA_PE",       "pais_id": "PE", "moneda": "USD"},
    "CXC_PE":                {"empresa_id": "AGROVECA_PE",       "pais_id": "PE", "moneda": "USD"},
    "COBRADAS_PE":           {"empresa_id": "AGROVECA_PE",       "pais_id": "PE", "moneda": "USD"},
    "COMISIONES_PE":         {"empresa_id": "AGROVECA_PE",       "pais_id": "PE", "moneda": "USD"},
    "LIBRO_BASE":            {"empresa_id": "GRUPO_AV_LATAM",    "pais_id": "LATAM", "moneda": "MULTI"},
    "PRECIO_PISO_CL":        {"empresa_id": "AGROVECA_CL",       "pais_id": "CL", "moneda": "CLP"},
    "PRECIO_PISO_PE":        {"empresa_id": "AGROVECA_PE",       "pais_id": "PE", "moneda": "USD"},
    "RESUMEN":               {"empresa_id": "AGROCOMERCIAL_CL",  "pais_id": "CL", "moneda": "CLP"},
    "REFERENCIA":            {"empresa_id": "GRUPO_AV_LATAM",    "pais_id": "LATAM", "moneda": "N/A"},
    "IGNORADO":              {"empresa_id": None,                 "pais_id": None, "moneda": None},
    "SIN_CLASIFICAR":        {"empresa_id": None,                 "pais_id": None, "moneda": None},
}

# Año por defecto cuando el nombre de archivo no incluye año
AÑO_DEFAULT = 2026


# ---------------------------------------------------------------------------
# REGLAS DE CLASIFICACIÓN
# ---------------------------------------------------------------------------

def clasificar(filename: str) -> dict:
    """
    Clasifica un archivo del inbox y extrae metadata desde el nombre.
    Retorna un dict con: tipo_archivo, empresa_id, pais_id, moneda,
    fecha_corte (str YYYY-MM-DD o None), variante (str o None),
    es_procesable (bool), razon_no_procesable (str o None)
    """
    fn = filename.strip()
    fn_upper = fn.upper()
    result = {
        "tipo_archivo": "SIN_CLASIFICAR",
        "empresa_id": None,
        "pais_id": None,
        "moneda": None,
        "fecha_corte": None,
        "variante": None,
        "es_procesable": True,
        "razon_no_procesable": None,
    }

    # ------------------------------------------------------------------
    # 1. IGNORADOS (no procesables estructuralmente)
    # ------------------------------------------------------------------
    if fn.startswith("~$"):
        return _set(result, "IGNORADO", False, "Archivo temporal de Office (lock)")

    ext = Path(fn).suffix.lower()
    if ext == ".eml":
        return _set(result, "IGNORADO", False, "Contenedor de email — no procesable")
    if ext in (".rtf", ".txt"):
        return _set(result, "IGNORADO", False, f"Formato {ext} no soportado")
    if ext == ".json":
        return _set(result, "REFERENCIA", True, None)

    if ext != ".xlsx":
        return _set(result, "IGNORADO", False, f"Extensión no soportada: {ext}")

    # ------------------------------------------------------------------
    # 2. LIBRO BASE
    # ------------------------------------------------------------------
    if re.search(r"nuevo\s+libro\s+base\s+av\s+\d{4}", fn, re.IGNORECASE):
        result.update(TIPOS["LIBRO_BASE"])
        result["tipo_archivo"] = "LIBRO_BASE"
        m = re.search(r"(\d{4})", fn)
        if m:
            result["fecha_corte"] = f"{m.group(1)}-12-31"  # año completo
        return result

    if re.search(r"libro_base_av_\d{4}", fn, re.IGNORECASE):
        result.update(TIPOS["LIBRO_BASE"])
        result["tipo_archivo"] = "LIBRO_BASE"
        m = re.search(r"(\d{4})", fn)
        if m:
            result["fecha_corte"] = f"{m.group(1)}-12-31"
        return result

    # ------------------------------------------------------------------
    # 3. CXC PERÚ
    #    "AGROVECA - CUENTAS POR COBRAR AL DD..MM.YYYY.xlsx"
    #    "AGROVECA - CUENTAS POR COBRAR AL- DD..MM.YYYY .xlsx"
    # ------------------------------------------------------------------
    if re.search(r"AGROVECA\s*-\s*CUENTAS POR COBRAR", fn_upper):
        result.update(TIPOS["CXC_PE"])
        result["tipo_archivo"] = "CXC_PE"
        # Extraer fecha: DD..MM.YYYY o DD.MM.YYYY
        m = re.search(r"(\d{1,2})\.\.?(\d{2})\.(\d{4})", fn)
        if m:
            result["fecha_corte"] = _fecha(m.group(3), m.group(2), m.group(1))
        return result

    # ------------------------------------------------------------------
    # 4. VENTAS PERÚ
    # ------------------------------------------------------------------

    # 4a. "AGROVECA PERU - VENTAS AL DD.MM.YYYY.xlsx"
    if re.search(r"AGROVECA PERU\s*-\s*VENTAS\s+AL", fn_upper):
        result.update(TIPOS["VENTAS_PE"])
        result["tipo_archivo"] = "VENTAS_PE"
        m = re.search(r"(\d{1,2})\.(\d{2})\.(\d{4})", fn)
        if m:
            result["fecha_corte"] = _fecha(m.group(3), m.group(2), m.group(1))
        # Marcar duplicado numérico si tiene sufijo " 2"
        if re.search(r"\s+2\.xlsx$", fn, re.IGNORECASE):
            result["variante"] = "COPIA_2"
        return result

    # 4b. "VENTAS FACTURADAS AL DD.MM.YYYY.xlsx"
    if re.search(r"VENTAS FACTURADAS AL", fn_upper):
        result.update(TIPOS["VENTAS_PE"])
        result["tipo_archivo"] = "VENTAS_PE"
        m = re.search(r"(\d{1,2})\.(\d{2})\.(\d{4})", fn)
        if m:
            result["fecha_corte"] = _fecha(m.group(3), m.group(2), m.group(1))
        return result

    # 4c. "Ventas al DD-MM.xlsx" y variantes — archivos tempranos de Perú (formato antiguo)
    if re.match(r"ventas\s+al\s+", fn, re.IGNORECASE) and not re.search(r"libro", fn, re.IGNORECASE):
        result.update(TIPOS["VENTAS_PE"])
        result["tipo_archivo"] = "VENTAS_PE"
        # "Ventas al 21-04.xlsx" → day=21, month=04
        # "Ventas al 012-04..xlsx" → day=12, month=04 (0 espurio)
        m = re.search(r"(\d+)-(\d{2})", fn)
        if m:
            dia = int(m.group(1)) % 100  # elimina ceros espurios
            mes = int(m.group(2))
            result["fecha_corte"] = _fecha(str(AÑO_DEFAULT), f"{mes:02d}", f"{dia:02d}")
        # Marcar como formato antiguo → pierde frente a "AGROVECA PERU - VENTAS" del mismo día
        result["variante"] = "FORMATO_ANTIGUO"
        return result

    # ------------------------------------------------------------------
    # 5. COBRADAS PERÚ
    #    "AGROVECA PERU - REPORTE DE VENTAS COBRADAS YYYY.xlsx"
    # ------------------------------------------------------------------
    if re.search(r"REPORTE DE VENTAS COBRADAS", fn_upper):
        result.update(TIPOS["COBRADAS_PE"])
        result["tipo_archivo"] = "COBRADAS_PE"
        m = re.search(r"(\d{4})", fn)
        if m:
            result["fecha_corte"] = f"{m.group(1)}-12-31"
        return result

    # ------------------------------------------------------------------
    # 6. COMISIONES PERÚ
    #    "AGROVECA PERU - COMISIONES TRABAJADORES YYYY.xlsx"
    # ------------------------------------------------------------------
    if re.search(r"COMISIONES TRABAJADORES", fn_upper):
        result.update(TIPOS["COMISIONES_PE"])
        result["tipo_archivo"] = "COMISIONES_PE"
        m = re.search(r"(\d{4})", fn)
        if m:
            result["fecha_corte"] = f"{m.group(1)}-12-31"
        return result

    # ------------------------------------------------------------------
    # 7. CXC CHILE — AGROVECA (Casa Matriz)
    #    "Cuentas Cobrar Agroveca DD-MM.xlsx"
    # ------------------------------------------------------------------
    if re.search(r"cuentas\s+cobrar\s+agroveca\b", fn, re.IGNORECASE):
        result.update(TIPOS["CXC_CL_AGROVECA"])
        result["tipo_archivo"] = "CXC_CL_AGROVECA"
        m = re.search(r"(\d{2})-(\d{2})", fn)
        if m:
            result["fecha_corte"] = _fecha(str(AÑO_DEFAULT), m.group(2), m.group(1))
        return result

    # ------------------------------------------------------------------
    # 8. CXC CHILE — AGROCOMERCIAL (AGLM)
    #    "Cuentas Cobrar AGrocomercial DD-MM.xlsx"
    #    "Cuentas Cobrar Agrocomercial DD-MM.xlsx"
    #    "Cuentas Cobrar NNNN.xlsx" (1204, 1704, 2904 — familia temprana)
    # ------------------------------------------------------------------
    if re.search(r"cuentas\s+cobrar\s+a[Gg]rocomercial\b", fn, re.IGNORECASE):
        result.update(TIPOS["CXC_CL_AGROCOMERCIAL"])
        result["tipo_archivo"] = "CXC_CL_AGROCOMERCIAL"
        m = re.search(r"(\d{2})-(\d{2})", fn)
        if m:
            result["fecha_corte"] = _fecha(str(AÑO_DEFAULT), m.group(2), m.group(1))
        return result

    # Familia temprana: "Cuentas Cobrar 1204.xlsx" = día 12, mes 04
    if re.match(r"cuentas\s+cobrar\s+(\d{4})\.xlsx", fn, re.IGNORECASE):
        m = re.match(r"cuentas\s+cobrar\s+(\d{2})(\d{2})\.xlsx", fn, re.IGNORECASE)
        if m:
            result.update(TIPOS["CXC_CL_AGROCOMERCIAL"])
            result["tipo_archivo"] = "CXC_CL_AGROCOMERCIAL"
            result["variante"] = "EARLY_AGROCOMERCIAL"
            result["fecha_corte"] = _fecha(str(AÑO_DEFAULT), m.group(2), m.group(1))
        return result

    # ------------------------------------------------------------------
    # 9. VENTAS CHILE — Libro de Ventas (AGLM)
    # ------------------------------------------------------------------

    # 9a. "Libro de Ventas DD-MM-YYYY.xlsx" y variantes
    if re.search(r"libro\s+de\s+ventas", fn, re.IGNORECASE):
        result.update(TIPOS["VENTAS_CL"])
        result["tipo_archivo"] = "VENTAS_CL"
        # Extraer fecha: DD-MM-YYYY con posibles espacios
        m = re.search(r"(\d{1,2})\s*-\s*(\d{2})-(\d{4})", fn)
        if m:
            result["fecha_corte"] = _fecha(m.group(3), m.group(2), m.group(1))
        # Marcar variantes
        if re.search(r"actualizada", fn, re.IGNORECASE):
            result["variante"] = "ACTUALIZADA"
        elif re.search(r"aglm", fn, re.IGNORECASE):
            result["variante"] = "AGLM"
        elif re.search(r"\s+2\.xlsx$", fn, re.IGNORECASE):
            result["variante"] = "COPIA_2"
        return result

    # 9b. "Ventas Julio GRUPO AV LATAM.xlsx"
    if re.search(r"ventas\s+julio\s+grupo", fn, re.IGNORECASE):
        result.update(TIPOS["VENTAS_CL"])
        result["tipo_archivo"] = "VENTAS_CL"
        result["fecha_corte"] = f"{AÑO_DEFAULT}-07-31"
        result["variante"] = "GRUPO_AV_LATAM"
        return result

    # ------------------------------------------------------------------
    # 10. PRECIO PISO
    # ------------------------------------------------------------------
    if re.search(r"precio\s+piso\s+peru|precios\s+piso\s+peru", fn, re.IGNORECASE):
        result.update(TIPOS["PRECIO_PISO_PE"])
        result["tipo_archivo"] = "PRECIO_PISO_PE"
        return result

    if re.search(r"precio[s]?\s+piso\s+ch", fn, re.IGNORECASE):
        result.update(TIPOS["PRECIO_PISO_CL"])
        result["tipo_archivo"] = "PRECIO_PISO_CL"
        return result

    # ------------------------------------------------------------------
    # 11. RESUMEN EJECUTIVO
    # ------------------------------------------------------------------
    if re.search(r"resumen.{0,20}ejecutivo", fn, re.IGNORECASE):
        result.update(TIPOS["RESUMEN"])
        result["tipo_archivo"] = "RESUMEN"
        return result

    # Sin clasificar
    return result


# ---------------------------------------------------------------------------
# SONDA DE ESTRUCTURA — clasificación por contenido, no solo por nombre (V-01)
# ---------------------------------------------------------------------------

# Ruta al catálogo de schemas (pipeline/schema_defs.json)
_SCHEMA_PATH = REPO_ROOT / "pipeline" / "schema_defs.json"
_schema_cache = None


def _load_schema_defs() -> dict:
    global _schema_cache
    if _schema_cache is not None:
        return _schema_cache
    if _SCHEMA_PATH.exists():
        with open(_SCHEMA_PATH, encoding="utf-8") as f:
            _schema_cache = json.load(f).get("types", {})
    else:
        _schema_cache = {}
    return _schema_cache


def classify_by_structure(filepath: Path) -> dict:
    """
    Intenta clasificar un archivo .xlsx leyendo su estructura interna
    (sheet names + columnas de la primera fila) cuando la clasificación
    por nombre devolvió SIN_CLASIFICAR.

    Retorna dict con keys: tipo_archivo (str), confianza ('alta'|'media'|'baja'|None).
    Si no puede clasificar: retorna {'tipo_archivo': None, 'confianza': None}.

    IMPORTANTE: solo se invoca para archivos .xlsx SIN_CLASIFICAR.
    Requiere openpyxl — si no está disponible, retorna sin clasificar.
    """
    if not _HAS_OPENPYXL:
        return {"tipo_archivo": None, "confianza": None}

    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        sheet_names_upper = [s.upper() for s in wb.sheetnames]

        schema_defs = _load_schema_defs()

        # ── VENTAS_CL: sheet 'VENTAS' + columnas MES / Vendedor / Total ──────
        if "VENTAS" in sheet_names_upper:
            ws = wb["VENTAS"] if "VENTAS" in wb.sheetnames else None
            if ws is None:
                # buscar case-insensitive
                for s in wb.sheetnames:
                    if s.upper() == "VENTAS":
                        ws = wb[s]
                        break
            if ws:
                # Leer fila de header (row_index=2 porque header_row=1 en pandas = fila 2 Excel)
                headers = []
                for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                    headers = [str(c).strip().upper() if c else "" for c in row]
                    break
                fp = schema_defs.get("VENTAS_CL", {}).get("structure_fingerprint", [])
                hits = sum(1 for col in fp if col.upper() in headers)
                if hits >= 2:
                    wb.close()
                    return {
                        "tipo_archivo": "VENTAS_CL",
                        "confianza": "alta" if hits >= 3 else "media",
                    }

        # ── VENTAS_PE: sheet 'RESUMEN' + fila con VENDEDOR + ENERO ──────────
        if "RESUMEN" in sheet_names_upper:
            ws = None
            for s in wb.sheetnames:
                if s.upper() == "RESUMEN":
                    ws = wb[s]
                    break
            if ws:
                # Buscar en las primeras 10 filas una que tenga VENDEDOR + un mes
                for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                    cells = [str(c).strip().upper() for c in row if c]
                    has_vend = any("VENDEDOR" in c for c in cells)
                    has_mes  = any("ENERO" in c or "FEBRERO" in c or "MARZO" in c for c in cells)
                    if has_vend and has_mes:
                        wb.close()
                        return {"tipo_archivo": "VENTAS_PE", "confianza": "alta"}

        # ── LIBRO_BASE: sheets 'Pricing Piso Chile' ───────────────────────────
        if any("PRICING PISO CHILE" in s.upper() for s in wb.sheetnames):
            wb.close()
            return {"tipo_archivo": "LIBRO_BASE", "confianza": "alta"}

        wb.close()
    except Exception:
        pass

    return {"tipo_archivo": None, "confianza": None}


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def _set(result: dict, tipo: str, procesable: bool, razon: str) -> dict:
    """Aplica tipo e indica si es procesable."""
    result["tipo_archivo"] = tipo
    result.update(TIPOS.get(tipo, {}))
    result["es_procesable"] = procesable
    result["razon_no_procesable"] = razon
    return result


def _fecha(año: str, mes: str, dia: str) -> str:
    """Construye fecha ISO YYYY-MM-DD con validación básica."""
    try:
        a, m, d = int(año), int(mes), int(dia)
        # Corrección de año de 2 dígitos (no esperada, pero por si acaso)
        if a < 100:
            a += 2000
        # Corrección de día/mes invertidos (no debería pasar, pero validamos)
        if m > 12 and d <= 12:
            m, d = d, m
        return date(a, m, d).isoformat()
    except Exception:
        return None


def md5_file(filepath: Path) -> str:
    """Calcula MD5 del archivo en bloques."""
    h = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# LÓGICA DE VIGENTE
# ---------------------------------------------------------------------------

VARIANTE_PRIORIDAD = {
    "ACTUALIZADA": 10,
    "AGLM": 5,
    "GRUPO_AV_LATAM": 3,
    "EARLY_AGROCOMERCIAL": -3,
    "FORMATO_ANTIGUO": -5,
    "COPIA_2": -10,
    None: 0,
}

# Variantes que resuelven un conflicto sin necesidad de alerta
VARIANTES_RESUELVEN_CONFLICTO = {"ACTUALIZADA", "AGLM", "COPIA_2", "FORMATO_ANTIGUO"}


def seleccionar_vigente(registros: list) -> str:
    """
    Dado un grupo de registros del mismo tipo+empresa, selecciona el filename
    vigente. Criterio: (1) fecha_corte más reciente, (2) variante con mayor
    prioridad en caso de empate de fecha.
    Retorna el filename del vigente.
    """
    def sort_key(r):
        fecha = r["fecha_corte"] or "0000-00-00"
        prio = VARIANTE_PRIORIDAD.get(r.get("variante"), 0)
        return (fecha, prio)

    ordenados = sorted(registros, key=sort_key, reverse=True)
    return ordenados[0]["filename"]


# ---------------------------------------------------------------------------
# PIPELINE PRINCIPAL
# ---------------------------------------------------------------------------

def run():
    ahora = datetime.utcnow().isoformat() + "Z"

    if not INBOX_PATH.exists():
        raise FileNotFoundError(f"Inbox no encontrado: {INBOX_PATH}")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    archivos_raw = sorted(INBOX_PATH.iterdir())

    registros = []
    hash_index = {}   # hash → list of filenames

    print(f"\n{'='*60}")
    print(f"  AV LATAM · Inbox Detector v{PIPELINE_VERSION}")
    print(f"  Inbox: {INBOX_PATH}")
    print(f"  Archivos encontrados: {len(archivos_raw)}")
    print(f"{'='*60}")

    for filepath in archivos_raw:
        if filepath.is_dir():
            continue

        fn = filepath.name
        clasificacion = clasificar(fn)

        registro = {
            "filename": fn,
            "filepath": str(filepath),
            "tipo_archivo": clasificacion["tipo_archivo"],
            "empresa_id": clasificacion.get("empresa_id"),
            "pais_id": clasificacion.get("pais_id"),
            "moneda": clasificacion.get("moneda"),
            "fecha_corte": clasificacion.get("fecha_corte"),
            "variante": clasificacion.get("variante"),
            "es_procesable": clasificacion["es_procesable"],
            "razon_no_procesable": clasificacion.get("razon_no_procesable"),
            "file_size_bytes": filepath.stat().st_size,
            "file_hash": None,
            "estado": "PENDIENTE",
            "fecha_ingesta": ahora,
        }

        if clasificacion["es_procesable"]:
            try:
                registro["file_hash"] = md5_file(filepath)
            except Exception as e:
                registro["es_procesable"] = False
                registro["razon_no_procesable"] = f"Error al leer: {e}"

        if registro["file_hash"]:
            if registro["file_hash"] not in hash_index:
                hash_index[registro["file_hash"]] = []
            hash_index[registro["file_hash"]].append(fn)

        registros.append(registro)

    # ------------------------------------------------------------------
    # DETECCIÓN DE DUPLICADOS (mismo hash exacto)
    # ------------------------------------------------------------------
    duplicados_hash = {h: fns for h, fns in hash_index.items() if len(fns) > 1}

    archivos_duplicados = set()
    for hash_val, fns in duplicados_hash.items():
        # El primero alfabéticamente es el "original" (convención)
        fns_sorted = sorted(fns)
        for fn in fns_sorted[1:]:
            archivos_duplicados.add(fn)

    # ------------------------------------------------------------------
    # AGRUPACIÓN POR tipo+empresa PARA DETECTAR CONFLICTOS Y VIGENTES
    # ------------------------------------------------------------------
    grupos = defaultdict(list)
    for r in registros:
        if r["es_procesable"] and r["tipo_archivo"] not in ("IGNORADO", "SIN_CLASIFICAR",
                                                              "LIBRO_BASE", "REFERENCIA",
                                                              "RESUMEN", "PRECIO_PISO_CL",
                                                              "PRECIO_PISO_PE"):
            clave = (r["tipo_archivo"], r["empresa_id"])
            grupos[clave].append(r)

    # ------------------------------------------------------------------
    # SELECCIÓN DE VIGENTE Y DETECCIÓN DE CONFLICTOS
    # ------------------------------------------------------------------
    vigentes = {}     # (tipo, empresa) → filename vigente
    conflictos = []   # alertas de conflicto

    for (tipo, empresa), grupo in grupos.items():
        if not grupo:
            continue

        # Detectar conflictos: misma fecha_corte, distinto hash, distinta variante
        fecha_hash = defaultdict(set)
        fecha_variante = defaultdict(set)
        for r in grupo:
            if r["fecha_corte"] and r["file_hash"]:
                fecha_hash[r["fecha_corte"]].add(r["file_hash"])
                fecha_variante[r["fecha_corte"]].add(r.get("variante"))

        for fc, hashes in fecha_hash.items():
            variantes = fecha_variante[fc]
            # Un conflicto es genuino solo si no puede resolverse por jerarquía de variantes
            conflicto_resuelto = bool(variantes & VARIANTES_RESUELVEN_CONFLICTO)
            if len(hashes) > 1 and not conflicto_resuelto:
                conflictos.append({
                    "tipo": "CONFLICTO",
                    "clave": f"{tipo}|{empresa}",
                    "fecha_corte": fc,
                    "archivos": [r["filename"] for r in grupo if r["fecha_corte"] == fc],
                    "hashes": list(hashes),
                    "descripcion": f"Misma fecha_corte ({fc}), diferente contenido, sin variante diferenciadora",
                })

        vigente_fn = seleccionar_vigente(grupo)
        vigentes[(tipo, empresa)] = vigente_fn

    # ------------------------------------------------------------------
    # ASIGNAR ESTADO FINAL A CADA REGISTRO
    # ------------------------------------------------------------------
    # LIBRO_BASE y PRECIO_PISO_* se excluyen del grupos loop pero deben aparecer
    # en vigentes para que load_files_from_registry() pueda ubicarlos.
    TIPOS_ESPECIALES = ("LIBRO_BASE", "PRECIO_PISO_CL", "PRECIO_PISO_PE")
    for tipo_esp in TIPOS_ESPECIALES:
        candidatos = [
            r for r in registros
            if r["tipo_archivo"] == tipo_esp and r.get("es_procesable")
        ]
        if candidatos:
            # Seleccionar el más reciente por fecha_corte o mtime
            mejor = max(candidatos, key=lambda r: (r.get("fecha_corte") or "0000-00-00",
                                                    r.get("file_size_bytes", 0)))
            empresa = mejor.get("empresa_id")
            clave = (tipo_esp, empresa)
            if clave not in vigentes:
                vigentes[clave] = mejor["filename"]

    vigentes_set = set(vigentes.values())

    for r in registros:
        tipo = r["tipo_archivo"]
        clave = (tipo, r.get("empresa_id"))

        if not r["es_procesable"] or tipo == "IGNORADO":
            r["estado"] = "IGNORADO"

        elif tipo in ("LIBRO_BASE", "REFERENCIA"):
            r["estado"] = "VIGENTE"

        elif tipo in ("RESUMEN", "PRECIO_PISO_CL", "PRECIO_PISO_PE"):
            r["estado"] = "REFERENCIA"

        elif r["filename"] in archivos_duplicados:
            r["estado"] = "DUPLICADO"

        elif r["filename"] in vigentes_set:
            r["estado"] = "VIGENTE"

        elif tipo == "SIN_CLASIFICAR":
            r["estado"] = "SIN_CLASIFICAR"

        else:
            r["estado"] = "SUPERSEDED"

    # ------------------------------------------------------------------
    # ALERTAS ADICIONALES
    # ------------------------------------------------------------------
    alertas = list(conflictos)

    # Alertas por duplicados de hash
    for hash_val, fns in duplicados_hash.items():
        alertas.append({
            "tipo": "DUPLICADO_EXACTO",
            "archivos": sorted(fns),
            "hash": hash_val,
            "descripcion": f"Archivos con contenido idéntico (hash {hash_val[:8]}…)",
        })

    # ── SONDA DE ESTRUCTURA: segunda oportunidad para SIN_CLASIFICAR ─────────
    # Si el nombre no fue suficiente, intentar clasificar por estructura interna.
    # Esto protege contra archivos renombrados que mantengan la estructura correcta.
    estructura_detectada = []
    for r in registros:
        if r["estado"] != "SIN_CLASIFICAR" or not r["filename"].lower().endswith(".xlsx"):
            continue
        filepath = Path(r["filepath"])
        if not filepath.exists():
            continue
        probe = classify_by_structure(filepath)
        if probe["tipo_archivo"]:
            tipo_detected = probe["tipo_archivo"]
            r["tipo_archivo"]       = tipo_detected
            r["variante"]           = "ESTRUCTURA_DETECTADA"
            r["empresa_id"]         = TIPOS.get(tipo_detected, {}).get("empresa_id")
            r["pais_id"]            = TIPOS.get(tipo_detected, {}).get("pais_id")
            r["moneda"]             = TIPOS.get(tipo_detected, {}).get("moneda")
            # Re-insertar en el grupo correspondiente para selección de vigente
            clave = (tipo_detected, r["empresa_id"])
            grupos[clave].append(r)
            vigente_fn = seleccionar_vigente(grupos[clave])
            vigentes[clave] = vigente_fn
            vigentes_set.add(vigente_fn)
            if r["filename"] in vigentes_set:
                r["estado"] = "VIGENTE"
            else:
                r["estado"] = "SUPERSEDED"
            estructura_detectada.append({
                "filename": r["filename"],
                "tipo_detectado": tipo_detected,
                "confianza": probe["confianza"],
            })
            alertas.append({
                "tipo": "ESTRUCTURA_DETECTADA",
                "archivo": r["filename"],
                "tipo_detectado": tipo_detected,
                "confianza": probe["confianza"],
                "descripcion": (
                    f"Archivo clasificado por estructura (no por nombre) como {tipo_detected}. "
                    f"Confianza: {probe['confianza']}. Considerar renombrar al formato canónico."
                ),
            })

    # Alerta por archivos SIN_CLASIFICAR (los que quedaron sin resolver)
    sin_clasi = [r["filename"] for r in registros if r["estado"] == "SIN_CLASIFICAR"]
    if sin_clasi:
        alertas.append({
            "tipo": "SIN_CLASIFICAR",
            "archivos": sin_clasi,
            "descripcion": "Archivos en inbox que no pudieron clasificarse (ni por nombre ni por estructura)",
        })

    # ------------------------------------------------------------------
    # RESUMEN POR TIPO
    # ------------------------------------------------------------------
    resumen_tipo = defaultdict(lambda: {"total": 0, "vigente": None, "superseded": 0,
                                         "ignorado": 0, "duplicado": 0, "conflicto": 0,
                                         "sin_clasificar": 0})
    for r in registros:
        t = r["tipo_archivo"]
        resumen_tipo[t]["total"] += 1
        estado = r["estado"]
        if estado == "VIGENTE":
            resumen_tipo[t]["vigente"] = r["filename"]
        elif estado == "SUPERSEDED":
            resumen_tipo[t]["superseded"] += 1
        elif estado in ("IGNORADO",):
            resumen_tipo[t]["ignorado"] += 1
        elif estado == "DUPLICADO":
            resumen_tipo[t]["duplicado"] += 1
        elif estado == "SIN_CLASIFICAR":
            resumen_tipo[t]["sin_clasificar"] += 1

    # ------------------------------------------------------------------
    # CONSTRUIR OUTPUT JSON
    # ------------------------------------------------------------------
    conteos = {
        "total": len(registros),
        "procesables": sum(1 for r in registros if r["es_procesable"]),
        "ignorados": sum(1 for r in registros if r["estado"] == "IGNORADO"),
        "vigentes": sum(1 for r in registros if r["estado"] == "VIGENTE"),
        "superseded": sum(1 for r in registros if r["estado"] == "SUPERSEDED"),
        "duplicados_exactos": sum(1 for r in registros if r["estado"] == "DUPLICADO"),
        "sin_clasificar": sum(1 for r in registros if r["estado"] == "SIN_CLASIFICAR"),
    }

    registro_vigentes = {
        f"{tipo}|{empresa}": filename
        for (tipo, empresa), filename in vigentes.items()
    }

    output = {
        "metadata": {
            "generado_en": ahora,
            "pipeline_version": PIPELINE_VERSION,
            "inbox_path": str(INBOX_PATH),
            "conteos": conteos,
        },
        "vigentes": registro_vigentes,
        "alertas": alertas,
        "resumen_por_tipo": {k: dict(v) for k, v in resumen_tipo.items()},
        "archivos": registros,
    }

    # ------------------------------------------------------------------
    # ESCRIBIR JSON
    # ------------------------------------------------------------------
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    # ------------------------------------------------------------------
    # PRINT RESUMEN
    # ------------------------------------------------------------------
    print(f"\n  CONTEOS:")
    for k, v in conteos.items():
        print(f"    {k:<25} {v}")

    print(f"\n  VIGENTES:")
    for clave, fn in sorted(registro_vigentes.items()):
        print(f"    {clave:<40} → {fn}")

    if alertas:
        print(f"\n  ALERTAS ({len(alertas)}):")
        for a in alertas:
            print(f"    [{a['tipo']}] {a.get('descripcion', '')}")
            for fn in a.get("archivos", [])[:3]:
                print(f"      - {fn}")

    print(f"\n  ✓ raw_registry.json escrito en: {OUTPUT_PATH}")
    print(f"{'='*60}\n")

    return output


if __name__ == "__main__":
    run()
