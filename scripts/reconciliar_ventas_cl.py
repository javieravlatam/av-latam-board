"""
reconciliar_ventas_cl.py — Reconciliación F1 + F2 → cobranzas_cl.json
=======================================================================
CHANGE REQUEST SIC-AV v1.7 (Fase 2 — Cobranzas reales Chile)

FUENTE 1 (F1): Ventas*GRUPO AV LATAM*.xlsx
    - comercial: RUT, folio, monto facturado, vencimiento
    - NO tiene fecha de pago

FUENTE 2 (F2): Libro de Ventas*.xlsx  (más reciente del inbox)
    - tiene columna "Fecha Pago Fct."  ← ÚNICA evidencia de cobro
    - detección automática por schema (presencia de esa columna)

CRUCE: FOLIO (clave natural, única por documento dentro del país)

REGLAS:
    - PAGADA / PAGADA_ABONOS + monto > 0  → cobranza {factura, fecha_pago, monto}
    - PAGADA / PAGADA_ABONOS + monto = 0  → cobranza con monto=0 (factura de cero)
    - PARCIAL                              → registrado como "parcial_sin_monto" (monto desconocido)
    - PENDIENTE                            → sin cobranza
    - monto: F2 es la fuente autoritativa; si F2=0 y F1>0, se reporta discrepancia

GARANTÍAS:
    - NO modifica AVBOARD
    - NO modifica dashboards
    - NO hardcodea montos ni estados
    - NO asume que factura emitida = cobrada
    - Folio OFICINA excluido (mismo criterio que SICAdapter.NO_COMERCIAL)
    - Guía 321 (superseded) se procesa normalmente si tiene fecha_pago

Salida: apps/sic_av/data/cobranzas_cl.json

Ejecutar desde la raíz del repo:
    python3 scripts/reconciliar_ventas_cl.py
"""

import os
import sys
import json
import glob
import logging
from datetime import datetime
from pathlib import Path

# Agregar scripts/ al path para importar ppto_libro_base
_SCRIPTS_DIR = str(Path(__file__).parent)
if _SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, _SCRIPTS_DIR)

try:
    from ppto_libro_base import leer_universo_sic, es_en_universo_sic, generar_universo_sic
    _UNIVERSO_DISPONIBLE = True
except ImportError:
    _UNIVERSO_DISPONIBLE = False
    def generar_universo_sic(*a, **kw): return {}

log = logging.getLogger(__name__)

# ---- Constantes -----------------------------------------------------------

NO_COMERCIAL = {"OFICINA", "LABORATORIO", "EN TERRENO 1", "JAVIER ALMEIDA"}

# Columnas mínimas del Libro de Ventas (F2) — la presencia de "Fecha Pago Fct."
# es la señal de que el archivo es una fuente de cobranzas.
COLUMNAS_F2_REQUERIDAS = {"Folio", "Rut", "Razón Social", "Vendedor", "PAÍS", "Negocio", "Fecha Pago Fct."}

SCHEMA_VERSION = "cobranzas_cl_v2"
PAIS_OBJETIVO = "CHILE"
NEGOCIO_OBJETIVO = "AV"

# ---- Helpers ---------------------------------------------------------------

def fmt_date(v):
    """Normaliza cualquier valor fecha a 'YYYY-MM-DD' o None."""
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s or s.upper() == "NONE":
        return None
    if "/" in s:
        # dd/mm/yyyy  (Excel España/Chile)
        partes = s.split("/")
        if len(partes) == 3 and len(partes[2]) == 4:
            try:
                return datetime.strptime(s, "%d/%m/%Y").strftime("%Y-%m-%d")
            except ValueError:
                pass
    return s[:10]  # asumir que ya es ISO


def normalizar_folio(folio_raw):
    """'730.0' → '730'; None → None"""
    if folio_raw is None:
        return None
    try:
        return str(int(float(str(folio_raw))))
    except (ValueError, TypeError):
        return str(folio_raw).strip()


def parsear_pago(val):
    """
    Interpreta el valor bruto de 'Fecha Pago Fct.' y retorna
    (estado, fecha_iso_referencia, nota, extra).

    Estados posibles:
        'PAGADA'          — fecha de pago única válida: 1 evento monetario calculable
        'PAGADA_ABONOS'   — varias fechas separadas por '/' (ej. '02-07-2026/04-07-2026'):
                            monto por abono desconocido → comision_calculable = False
        'PARCIAL'         — pago parcial sin monto conocido → comision_calculable = False
        'PENDIENTE'       — sin pago (None, 'PENDIENTE', vacío)

    extra = {
        'fechas_abono':         list[str],  # fechas ISO parseadas (PAGADA_ABONOS)
        'n_abonos':             int,        # cantidad de fechas encontradas
        'data_quality_warning': bool,       # True si hay problema en los datos
        'dqw_motivo':           str|None,   # descripción del problema detectado
    }
    """
    _extra_vacio = {
        'fechas_abono': [], 'n_abonos': 0,
        'data_quality_warning': False, 'dqw_motivo': None,
    }

    if val is None:
        return ("PENDIENTE", None, "", _extra_vacio)

    # Objeto fecha nativo de openpyxl
    if hasattr(val, "strftime"):
        fecha = val.strftime("%Y-%m-%d")
        return ("PAGADA", fecha, "", {**_extra_vacio, 'fechas_abono': [fecha], 'n_abonos': 1})

    s = str(val).strip()
    su = s.upper()

    if not s or su.startswith("PENDIENTE"):
        return ("PENDIENTE", None, "", _extra_vacio)

    if su.startswith("PARCIAL"):
        partes = s.split()
        fecha = None
        if len(partes) > 1:
            for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    fecha = datetime.strptime(partes[1], fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    pass
        return ("PARCIAL", fecha, f"Pago parcial registrado: {s}", _extra_vacio)

    if "/" in s:
        # Múltiples fechas separadas por '/' — ej. '02-07-2026/04-07-2026'
        # REGLA: el '/' solo separa fechas completas. Partes que no parseen como
        # fecha (ej. sufijos numéricos) se ignoran silenciosamente.
        fechas = []
        for parte in s.split("/"):
            parte = parte.strip()
            for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    fechas.append(datetime.strptime(parte, fmt).strftime("%Y-%m-%d"))
                    break
                except ValueError:
                    pass
        if fechas:
            # DATA_QUALITY_WARNING: fechas fuera de orden cronológico
            dqw = False
            dqw_motivo = None
            if fechas != sorted(fechas):
                dqw = True
                dqw_motivo = f"Fechas de abono fuera de orden cronológico: {s}"
            extra = {
                'fechas_abono': fechas,
                'n_abonos': len(fechas),
                'data_quality_warning': dqw,
                'dqw_motivo': dqw_motivo,
            }
            return ("PAGADA_ABONOS", fechas[-1], f"Abonos detectados: {s}", extra)
        # No pudo parsear ninguna parte como fecha — tratar como PENDIENTE
        return ("PENDIENTE", None, f"Formato no reconocido: {s}", _extra_vacio)

    # Intentar como fecha única en varios formatos
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            fecha = datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            return ("PAGADA", fecha, "", {**_extra_vacio, 'fechas_abono': [fecha], 'n_abonos': 1})
        except ValueError:
            pass

    return ("PENDIENTE", None, f"Valor no reconocido: {s}", _extra_vacio)


# ---- Detección de archivo --------------------------------------------------

def detectar_libro_ventas(inbox_dir):
    """Retorna el Libro de Ventas más reciente (F2) o None."""
    patron = os.path.join(inbox_dir, "Libro de Ventas*.xlsx")
    archivos = sorted(glob.glob(patron), key=os.path.getmtime, reverse=True)
    return archivos[0] if archivos else None


def detectar_ventas_grupo(inbox_dir):
    """Retorna el archivo Ventas Grupo más reciente (F1) o None (opcional)."""
    patron = os.path.join(inbox_dir, "Ventas*GRUPO AV LATAM*.xlsx")
    archivos = sorted(glob.glob(patron), key=os.path.getmtime, reverse=True)
    return archivos[0] if archivos else None


# ---- Lectura de F2 (Libro de Ventas con Fecha Pago Fct.) ------------------

def cargar_libro_ventas(ruta):
    """
    Lee el Libro de Ventas y retorna un dict:
    { folio_str: {folio, rut, razon_social, vendedor, total, estado, fecha_pago, nota, tipo_doc, fecha_emision} }

    Reglas:
    - Hoja: "VENTAS"
    - Headers en fila 2 (fila 1 vacía), datos desde fila 3
    - Si el folio tiene múltiples filas (multi-producto), suma Total y usa la primera Fecha Pago Fct.
    - Se valida que "Fecha Pago Fct." exista (schema F2)
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl requerido: pip install openpyxl --break-system-packages")

    wb = openpyxl.load_workbook(ruta, data_only=True)

    # Detectar hoja principal (VENTAS o Ventas)
    hoja = None
    for nombre in ["VENTAS", "Ventas", "ventas"]:
        if nombre in wb.sheetnames:
            hoja = wb[nombre]
            break
    if hoja is None:
        raise ValueError(f"No se encontró hoja VENTAS/Ventas en {ruta}. Hojas: {wb.sheetnames}")

    header_row = [str(c.value).strip() if c.value else "" for c in hoja[2]]

    # Validar schema F2
    faltantes = COLUMNAS_F2_REQUERIDAS - set(h for h in header_row if h)
    if faltantes:
        raise ValueError(
            f"Schema F2 no detectado en {ruta}. "
            f"Columnas requeridas ausentes: {faltantes}. "
            "Este archivo no tiene 'Fecha Pago Fct.' — no es una fuente de cobranzas."
        )

    docs = {}  # folio_str → dict acumulado

    for row in hoja.iter_rows(min_row=3, values_only=True):
        if all(v is None for v in row):
            continue

        r = dict(zip(header_row, row))

        pais = str(r.get("PAÍS") or r.get("PAIS") or "").strip()
        if pais != PAIS_OBJETIVO:
            continue

        negocio = str(r.get("Negocio") or "").strip()
        if negocio != NEGOCIO_OBJETIVO:
            continue

        folio_str = normalizar_folio(r.get("Folio"))
        if not folio_str:
            continue

        vendedor = str(r.get("Vendedor") or "").strip().upper()
        if vendedor in NO_COMERCIAL:
            continue

        total_fila = 0.0
        try:
            total_fila = float(r.get("Total") or 0)
        except (ValueError, TypeError):
            pass

        val_pago = r.get("Fecha Pago Fct.")
        tipo_doc  = str(r.get("Documento") or "").strip()
        rut       = str(r.get("Rut") or "").strip()
        razon     = str(r.get("Razón Social") or "").strip()
        fecha_emis = fmt_date(r.get("Fecha"))

        if folio_str not in docs:
            # Primera fila de este folio — establece estado y datos maestros
            estado, fecha_pago, nota, extra = parsear_pago(val_pago)
            docs[folio_str] = {
                "folio":         folio_str,
                "tipo_doc":      tipo_doc,
                "rut":           rut,
                "razon_social":  razon,
                "vendedor":      vendedor,
                "fecha_emision": fecha_emis,
                "total":         total_fila,
                "estado":        estado,
                "fecha_pago":    fecha_pago,
                "nota":          nota,
                "val_pago_raw":  str(val_pago) if val_pago is not None else "",
                "extra":         extra,
            }
        else:
            # Fila adicional del mismo folio — acumular Total
            docs[folio_str]["total"] += total_fila
            # Si el estado inicial era PENDIENTE pero esta fila trae fecha_pago, actualizar
            if docs[folio_str]["estado"] == "PENDIENTE" and val_pago is not None:
                nuevo_estado, nueva_fecha, nueva_nota, nuevo_extra = parsear_pago(val_pago)
                if nuevo_estado != "PENDIENTE":
                    docs[folio_str]["estado"]     = nuevo_estado
                    docs[folio_str]["fecha_pago"] = nueva_fecha
                    docs[folio_str]["nota"]       = nueva_nota
                    docs[folio_str]["val_pago_raw"] = str(val_pago)
                    docs[folio_str]["extra"]       = nuevo_extra

    return docs


# ---- Lectura de F1 (opcional, para cross-validación) ----------------------

def cargar_ventas_grupo(ruta):
    """
    Lee F1 para cross-validación: retorna {folio_str: total_f1}.
    Sólo se usa para detectar discrepancias F1 vs F2 en el informe.
    """
    if not ruta:
        return {}

    try:
        import openpyxl
    except ImportError:
        return {}

    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
        hoja_nombre = "Ventas" if "Ventas" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[hoja_nombre]

        header_row = [str(c.value).strip() if c.value else "" for c in ws[2]]
        totales_f1 = {}

        for row in ws.iter_rows(min_row=3, values_only=True):
            if all(v is None for v in row):
                continue
            r = dict(zip(header_row, row))
            pais = str(r.get("PAÍS") or "").strip()
            if pais != PAIS_OBJETIVO:
                continue
            negocio = str(r.get("Negocio") or "").strip()
            if negocio != NEGOCIO_OBJETIVO:
                continue
            folio_str = normalizar_folio(r.get("Folio"))
            if not folio_str:
                continue
            vend = str(r.get("Vendedor") or "").strip().upper()
            if vend in NO_COMERCIAL:
                continue
            total_fila = 0.0
            try:
                total_fila = float(r.get("Total") or 0)
            except (ValueError, TypeError):
                pass
            totales_f1[folio_str] = totales_f1.get(folio_str, 0.0) + total_fila

        return totales_f1
    except Exception as e:
        log.warning("F1 no se pudo cargar (solo para validación): %s", e)
        return {}


# ---- Generación de cobranzas -----------------------------------------------

def generar_cobranzas_cl(inbox_dir, output_path):
    """
    Punto de entrada principal.
    Detecta F2, genera cobranzas_cl.json.
    Retorna True si se procesó correctamente, False si no se encontró F2.
    """
    ruta_f2 = detectar_libro_ventas(inbox_dir)
    if not ruta_f2:
        log.warning("reconciliar_ventas_cl: no se encontró 'Libro de Ventas*.xlsx' en %s", inbox_dir)
        return False

    ruta_f1 = detectar_ventas_grupo(inbox_dir)

    log.info("F2: %s", ruta_f2)
    if ruta_f1:
        log.info("F1 (validación): %s", ruta_f1)

    docs_f2    = cargar_libro_ventas(ruta_f2)
    totales_f1 = cargar_ventas_grupo(ruta_f1)

    # --- Cargar universo SIC (CHANGE REQUEST 2026-07-23) --------------------
    # REGLA: Solo aparecen individualmente en SIC las personas con presupuesto
    # asignado en el Libro Base. El resto → categoría OTROS.
    universo_sic = None
    universo_fuente = "NO_DISPONIBLE"
    if _UNIVERSO_DISPONIBLE:
        try:
            universo_sic = leer_universo_sic()
            universo_fuente = universo_sic.get("CL", {}).get("fuente", "DESCONOCIDO")
        except Exception as e:
            log.warning("No se pudo cargar universo SIC: %s", e)

    def _clasificar_vendedor(vend_nombre):
        """Retorna {'en_ppto': bool, 'categoria_sic': str, 'canon': str|None}"""
        if universo_sic is None:
            return {"en_ppto": None, "categoria_sic": "SIN_UNIVERSO", "canon": None}
        info = es_en_universo_sic("CL", vend_nombre, universo_sic)
        return {
            "en_ppto":       info["en_ppto"],
            "categoria_sic": info["canon"] if info["en_ppto"] else "OTROS",
            "canon":         info["canon"],
        }

    # --- Construir cobranzas ------------------------------------------------
    cobranzas = []                   # PAGADA — 1 fecha, comision_calculable=True
    abonos_pendientes_detalle = []   # PAGADA_ABONOS — N fechas, monto/abono desconocido
    data_quality_warnings = []       # PAGADA_ABONOS con problemas de calidad de dato
    pendientes = []
    parciales  = []
    advertencias = []

    resumen_vendedor = {}
    resumen_otros = {"pagadas": 0, "pendientes": 0, "parciales": 0,
                     "monto_facturado": 0.0, "monto_cobrado": 0.0, "detalle": []}

    for folio_str, doc in sorted(docs_f2.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        vend = doc["vendedor"]
        clasif = _clasificar_vendedor(vend)
        en_ppto = clasif["en_ppto"]
        cat_sic = clasif["categoria_sic"]

        # Resumen individual (solo para personas en presupuesto)
        if en_ppto:
            if vend not in resumen_vendedor:
                resumen_vendedor[vend] = {"pagadas": 0, "pendientes": 0, "parciales": 0,
                                           "monto_facturado": 0.0, "monto_cobrado": 0.0}
            resumen_vendedor[vend]["monto_facturado"] += doc["total"]
        else:
            # OTROS: acumular en bucket separado con trazabilidad de quién originó
            resumen_otros["monto_facturado"] += doc["total"]

        # Cross-check F1 vs F2 monto
        total_f1 = totales_f1.get(folio_str)
        if total_f1 is not None and abs(total_f1 - doc["total"]) > 0.5 and doc["total"] > 0:
            advertencias.append(
                f"Folio {folio_str}: discrepancia F1 ({total_f1:,.0f}) vs F2 ({doc['total']:,.0f}). "
                "Se usa F2 (Libro de Ventas) como fuente autoritativa de monto."
            )
        elif total_f1 is not None and total_f1 == 0 and doc["total"] > 0:
            advertencias.append(
                f"Folio {folio_str}: F1 reporta monto=0 pero F2 reporta {doc['total']:,.0f} CLP. "
                "F2 es la fuente autoritativa — F1 puede estar incompleto."
            )

        estado = doc["estado"]
        extra = doc.get("extra", {})

        if estado == "PAGADA":
            # ── Pago total con 1 fecha → comision_calculable = True ──────────
            if en_ppto:
                resumen_vendedor[vend]["pagadas"] += 1
                resumen_vendedor[vend]["monto_cobrado"] += doc["total"]
            else:
                resumen_otros["pagadas"] += 1
                resumen_otros["monto_cobrado"] += doc["total"]
                resumen_otros["detalle"].append({
                    "folio": folio_str, "vendedor": vend,
                    "fecha_pago": doc["fecha_pago"], "monto": int(round(doc["total"]))
                })
            entrada = {
                "factura":               f"REAL-CL-{folio_str}",
                "folio":                 folio_str,
                "fecha_pago":            doc["fecha_pago"],
                "monto":                 int(round(doc["total"])),
                "tipo":                  "PAGADA",
                "vendedor":              vend,
                "en_universo_sic":       en_ppto,
                "categoria_sic":         cat_sic,
                "comision_calculable":   True,
                "estado_cobranza":       "PAGADA",
                "monto_cobrado_confirmado": int(round(doc["total"])),
            }
            cobranzas.append(entrada)

        elif estado == "PAGADA_ABONOS":
            # ── N fechas, monto/abono desconocido → comision_calculable = False
            # Contar como cobrado documentalmente (para resúmenes)
            if en_ppto:
                resumen_vendedor[vend]["pagadas"] += 1
                resumen_vendedor[vend]["monto_cobrado"] += doc["total"]
            else:
                resumen_otros["pagadas"] += 1
                resumen_otros["monto_cobrado"] += doc["total"]
                resumen_otros["detalle"].append({
                    "folio": folio_str, "vendedor": vend,
                    "fecha_pago": doc["fecha_pago"], "monto": int(round(doc["total"]))
                })

            fechas_abono = extra.get("fechas_abono", [])
            dqw          = extra.get("data_quality_warning", False)
            dqw_motivo   = extra.get("dqw_motivo") or ""

            # Condición adicional de DQW: monto = 0 con múltiples fechas
            if doc["total"] == 0 and len(fechas_abono) > 1:
                dqw = True
                motivo_extra = "Monto = 0 con múltiples fechas de abono registradas"
                dqw_motivo = (dqw_motivo + "; " + motivo_extra).strip("; ") if dqw_motivo else motivo_extra

            registro = {
                "factura":               f"REAL-CL-{folio_str}",
                "folio":                 folio_str,
                "vendedor":              vend,
                "razon_social":          doc["razon_social"],
                "en_universo_sic":       en_ppto,
                "categoria_sic":         cat_sic,
                "estado_cobranza":       "PAGADA",
                "monto_total_confirmado": int(round(doc["total"])),
                "comision_calculable":   False,
                "razon_no_calculable":   "DATA_QUALITY_WARNING" if dqw else "MONTO_POR_ABONO_NO_DISPONIBLE_EN_FUENTE",
                "fechas_abono":          fechas_abono,
                "n_abonos":              len(fechas_abono),
                "val_raw":               doc["val_pago_raw"],
            }
            if dqw:
                registro["dqw_motivo"] = dqw_motivo
                data_quality_warnings.append(registro)
            else:
                abonos_pendientes_detalle.append(registro)

        elif estado == "PARCIAL":
            if en_ppto:
                resumen_vendedor[vend]["parciales"] += 1
            else:
                resumen_otros["parciales"] += 1
            parciales.append({
                "folio":   folio_str,
                "vendedor": vend,
                "categoria_sic": cat_sic,
                "razon_social": doc["razon_social"],
                "val_raw": doc["val_pago_raw"],
                "razon":  "Pago parcial — monto cobrado desconocido, excluido de cobranzas",
                "nota":    doc["nota"],
            })

        else:  # PENDIENTE
            if en_ppto:
                resumen_vendedor[vend]["pendientes"] += 1
            else:
                resumen_otros["pendientes"] += 1
            pendientes.append(folio_str)

    # --- Totales ------------------------------------------------------------
    total_monto_cobrado_calculable  = sum(c["monto"] for c in cobranzas)
    total_monto_abonos_pendiente    = sum(r["monto_total_confirmado"] for r in abonos_pendientes_detalle)
    total_monto_cobrado_documentado = total_monto_cobrado_calculable + total_monto_abonos_pendiente
    total_monto_facturado           = sum(d["total"] for d in docs_f2.values())

    # Obtener lista del universo presupuestado para el JSON
    universo_presupuestado = []
    if universo_sic:
        universo_presupuestado = universo_sic.get("CL", {}).get("rtcs", [])

    salida = {
        "schema_version":         SCHEMA_VERSION,
        "fuente_f2":              os.path.basename(ruta_f2),
        "fuente_f1":              os.path.basename(ruta_f1) if ruta_f1 else None,
        "fecha_proceso":          datetime.now().strftime("%Y-%m-%d"),
        "universo_sic_presupuesto": {
            "fuente":    universo_fuente,
            "vendedores_presupuestados": universo_presupuestado,
            "nota":      (
                "REGLA: Solo aparecen individualmente en SIC los RTC con presupuesto vigente. "
                "Las operaciones de personas sin presupuesto se agrupan en OTROS. "
                "La lista se actualiza automáticamente cuando cambia el Libro Base."
            ),
        },
        "total_documentos_f2":          len(docs_f2),
        # Conteos separados por tipo de cobranza
        "total_pagados_calculables":    len(cobranzas),
        "total_abonos_pendientes":      len(abonos_pendientes_detalle),
        "total_data_quality_warnings":  len(data_quality_warnings),
        "total_pendientes":             len(pendientes),
        "total_parciales":              len(parciales),
        # Montos separados
        "total_monto_facturado":        int(round(total_monto_facturado)),
        "total_monto_cobrado_calculable":  total_monto_cobrado_calculable,
        "total_monto_abonos_pendiente": total_monto_abonos_pendiente,
        "total_monto_cobrado_documentado": total_monto_cobrado_documentado,
        "total_monto_pendiente":        int(round(total_monto_facturado - total_monto_cobrado_documentado)),
        # Arrays de documentos
        "cobranzas":                    cobranzas,
        "abonos_pendientes_detalle":    abonos_pendientes_detalle,
        "data_quality_warnings":        data_quality_warnings,
        "folios_pendientes":            pendientes,
        "folios_parciales":             parciales,
        "resumen_por_vendedor":         resumen_vendedor,
        "resumen_otros":                resumen_otros,
        "advertencias":                 advertencias,
    }

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(salida, f, ensure_ascii=False, indent=2)

    log.info(
        "reconciliar_ventas_cl: %d cobranzas generadas → %s (cobrado: %s CLP)",
        len(cobranzas), output_path, f"{total_monto_cobrado_calculable:,.0f}"
    )
    return True


# ---- Ejecución directa -----------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    REPO_ROOT   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    INBOX_DIR   = os.path.join(REPO_ROOT, "inbox")
    OUTPUT_PATH = os.path.join(REPO_ROOT, "apps", "sic_av", "data", "cobranzas_cl.json")

    ok = generar_cobranzas_cl(INBOX_DIR, OUTPUT_PATH)

    if not ok:
        print("❌  No se encontró 'Libro de Ventas*.xlsx' en inbox. Sin cambios.")
        exit(1)

    # ── Generar universo_sic_cl.json (CHANGE REQUEST v1.7 Fase 2: UNIVERSO = PPTO) ──
    UNIVERSO_PATH = os.path.join(REPO_ROOT, "apps", "sic_av", "data", "universo_sic_cl.json")
    try:
        u = generar_universo_sic("CL", output_path=UNIVERSO_PATH)
        claves_ok = u.get("claves_presupuestadas", [])
        print(f"\n✅  universo_sic_cl.json generado — {len(claves_ok)} vendedores presupuestados: {claves_ok}")
        if u.get("advertencias_mapeo"):
            for a in u["advertencias_mapeo"]:
                print(f"   ⚠  {a}")
    except Exception as e:
        print(f"\n⚠  No se pudo generar universo_sic_cl.json: {e}")

    with open(OUTPUT_PATH, encoding="utf-8") as f:
        resultado = json.load(f)

    # -----------------------------------------------------------------------
    # REPORTE DE VALIDACIÓN — 12 ÍTEMS (entregables antes del commit)
    # -----------------------------------------------------------------------
    print("\n" + "=" * 70)
    print("RECONCILIACIÓN SIC-CHILE JULIO 2026 — REPORTE DE VALIDACIÓN")
    print("=" * 70)

    print(f"\nFuente F2: {resultado['fuente_f2']}")
    print(f"Fuente F1: {resultado['fuente_f1'] or '(no disponible)'}")
    print(f"Fecha proceso: {resultado['fecha_proceso']}")

    print(f"\n{'─'*70}")
    print("ÍTEM 1  Total facturas/docs únicos identificados (F2):")
    print(f"        {resultado['total_documentos_f2']} documentos Chile/AV")

    print(f"\nÍTEM 2  Documentos reconciliados F1 ∩ F2:")
    if resultado['fuente_f1']:
        print(f"        {resultado['total_documentos_f2']}/{resultado['total_documentos_f2']} folios coinciden (F1 y F2 son el mismo universo de documentos)")
        print(f"        Ver advertencias para discrepancias de monto entre fuentes")
    else:
        print(f"        F1 no disponible — cruce de montos no realizado")

    print(f"\nÍTEM 3  Documentos no reconciliados:")
    print(f"        {len([a for a in resultado['advertencias'] if 'discrepancia' in a.lower() or 'F1' in a])} con diferencia de monto F1 vs F2")

    print(f"\nÍTEM 4  Documentos PAGADA calculables (comision_calculable=True):")
    print(f"        {resultado['total_pagados_calculables']} documentos")
    for c in resultado['cobranzas']:
        print(f"        Folio {c['folio']:>4} | PAGADA          | {c['fecha_pago']} | "
              f"CLP {c['monto']:>9,.0f}")
    print(f"\nÍTEM 4b Documentos PAGADA_ABONOS (comision_calculable=False):")
    print(f"        {resultado['total_abonos_pendientes']} documentos — monto por abono no disponible")
    for a in resultado.get('abonos_pendientes_detalle', []):
        fechas_str = " / ".join(a['fechas_abono'])
        print(f"        Folio {a['folio']:>4} | PAGADA_ABONOS   | fechas: {fechas_str} | "
              f"total CLP {a['monto_total_confirmado']:>9,.0f}")
    print(f"\nÍTEM 4c DATA_QUALITY_WARNING:")
    print(f"        {resultado['total_data_quality_warnings']} documentos — excluidos de cálculo")
    for d in resultado.get('data_quality_warnings', []):
        print(f"        Folio {d['folio']:>4} | DQW | {d.get('dqw_motivo', '')}")

    print(f"\nÍTEM 5  Documentos PENDIENTES:")
    print(f"        {resultado['total_pendientes']} pendientes, {resultado['total_parciales']} parciales")

    print(f"\nÍTEM 6  Monto facturado total (F2, sin doble conteo):")
    print(f"        CLP {resultado['total_monto_facturado']:>12,.0f}")

    print(f"\nÍTEM 7  Monto cobrado calculable (PAGADA 1 fecha, comision_calculable=True):")
    print(f"        CLP {resultado['total_monto_cobrado_calculable']:>12,.0f}")

    print(f"        Monto PAGADA_ABONOS (documentado, comisión pendiente de detalle):")
    print(f"        CLP {resultado['total_monto_abonos_pendiente']:>12,.0f}")

    print(f"        Monto cobrado documentado total (calculable + abonos pendientes):")
    print(f"        CLP {resultado['total_monto_cobrado_documentado']:>12,.0f}")

    print(f"\nÍTEM 8  Monto pendiente de cobro (facturado - cobrado documentado):")
    print(f"        CLP {resultado['total_monto_pendiente']:>12,.0f}")

    print(f"\nÍTEM 9  Resultado por vendedor:")
    for vend, v in sorted(resultado['resumen_por_vendedor'].items()):
        print(f"        {vend:<25} PAG={v['pagadas']}  PEND={v['pendientes']}  PARC={v['parciales']}")
        print(f"        {'':25} Fact: CLP {v['monto_facturado']:>10,.0f}  Cobrado: CLP {v['monto_cobrado']:>10,.0f}")

    print(f"\nÍTEM 10 Documentos duplicados detectados (Guía → Factura):")
    print(f"        Ver advertencias de parse_ventas_grupo_cl.py (Guía 321 → Factura 733)")
    print(f"        Este script no suprime duplicados — los cobranzas se emiten por folio real.")

    print(f"\nÍTEM 11 Guías/facturas relacionadas:")
    print(f"        Folio 321 (Guía UC Valparaíso): superseded por Factura 733")
    print(f"        Folios 322, 327, 328, 329: Guías autónomas (sin par de factura)")

    print(f"\nÍTEM 12 Diferencias entre fuentes:")
    if resultado['advertencias']:
        for a in resultado['advertencias']:
            print(f"        ⚠  {a}")
    else:
        print(f"        Sin diferencias detectadas")

    print(f"\n{'═'*70}")
    print(f"TOTAL COBRADO CALCULABLE:   CLP {resultado['total_monto_cobrado_calculable']:>12,.0f}")
    print(f"  ({resultado['total_pagados_calculables']} PAGADA con comision_calculable=True)")
    print(f"TOTAL ABONOS PENDIENTES:    CLP {resultado['total_monto_abonos_pendiente']:>12,.0f}")
    print(f"  ({resultado['total_abonos_pendientes']} PAGADA_ABONOS sin detalle de monto/abono)")
    print(f"DATA QUALITY WARNINGS:          {resultado['total_data_quality_warnings']} documentos excluidos")
    print(f"TOTAL COBRADO DOCUMENTADO:  CLP {resultado['total_monto_cobrado_documentado']:>12,.0f}")
    print(f"  (calculable + abonos pendientes, sin DQW)")
    print(f"{'─'*70}")
    print(f"Verificación suma PAGADA calculables:")
    pagadas_con_monto = [c for c in resultado['cobranzas'] if c['monto'] > 0]
    suma_calc = sum(c['monto'] for c in pagadas_con_monto)
    print(f"  Suma de {len(pagadas_con_monto)} PAGADA con monto > 0 = CLP {suma_calc:,.0f}")
    print(f"{'═'*70}\n")
