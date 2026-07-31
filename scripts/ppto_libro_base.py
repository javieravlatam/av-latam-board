#!/usr/bin/env python3
"""
ppto_libro_base.py
==================
Módulo único de presupuesto AVBOARD — Chile y Perú.

Fuente principal : inbox/nuevo libro base AV 2026.xlsx → hoja "Presupuesto Pais"
Fallback         : constantes LEGACY (históricas hasta jun 2026)

Estructura real de la hoja:
  Fila 0  : "PRESUPUESTO 2026 — CHILE (CLP)"
  Fila 1  : RTC | ENE | FEB | ... | DIC | TOTAL
  Filas 2-8 : RTCs Chile
  Fila 9  : TOTAL Chile
  Fila 12 : "PRESUPUESTO 2026 — PERÚ (USD)"
  Fila 13 : RTC | ENE | FEB | ... | DIC | TOTAL
  Filas 14-20: RTCs Perú
  Fila 21 : TOTAL Perú

Compatible: Python 3.9.6+
"""

from pathlib import Path
from typing import Dict, List, Optional

# ─────────────────────────────────────────────────────────────
# LEGACY CHILE (CLP)
# "Presupuesto histórico utilizado hasta junio 2026"
# ─────────────────────────────────────────────────────────────

PPTO_MENSUAL_CL_LEGACY: List[float] = [
     83_558_032,   # Ene
     41_601_950,   # Feb
     42_000_000,   # Mar
     46_431_368,   # Abr
     51_000_000,   # May
     49_730_000,   # Jun
     62_800_000,   # Jul
     76_500_000,   # Ago
    110_800_000,   # Sep
    112_700_000,   # Oct
     97_500_000,   # Nov
     86_700_000,   # Dic
]
PPTO_ANUAL_CL_LEGACY: float = sum(PPTO_MENSUAL_CL_LEGACY)   # 861,321,350

# ─────────────────────────────────────────────────────────────
# LEGACY PERÚ (USD)
# "Presupuesto histórico utilizado hasta junio 2026"
# ─────────────────────────────────────────────────────────────

PPTO_MENSUAL_PE_LEGACY: List[float] = [
     51_674,   # Ene
     58_489,   # Feb
    103_222,   # Mar
     71_299,   # Abr
     61_946,   # May
     78_710,   # Jun
    100_675,   # Jul
    178_180,   # Ago
    125_564,   # Sep
    165_842,   # Oct
     98_481,   # Nov
     42_952,   # Dic
]
PPTO_ANUAL_PE_LEGACY: float = sum(PPTO_MENSUAL_PE_LEGACY)   # 1,137,034

# ─────────────────────────────────────────────────────────────
# VALORES ESPERADOS DESDE LIBRO_BASE (para validación)
# ─────────────────────────────────────────────────────────────

EXPECTED_CL_ANUAL:  float = 728_110_400.0   # nuevo libro base AV 2026 (actualizado Jul 2026)
EXPECTED_CL_PPTO5M: float = 290_513_800.0   # Ene–May 2026 nuevo libro
EXPECTED_PE_ANUAL:  float = 1_210_600.0
EXPECTED_PE_PPTO5M: float = 350_134.3

# ─────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────

REPO_DIR    = Path(__file__).parent.parent
INBOX_EXCEL = REPO_DIR / "inbox" / "nuevo libro base AV 2026.xlsx"
SHEET_NAME  = "Presupuesto Pais"

MONTH_VARIANTS: List[List[str]] = [
    ["ene", "jan", "enero",      "january"],
    ["feb",        "febrero",    "february"],
    ["mar",        "marzo",      "march"],
    ["abr", "apr", "abril",      "april"],
    ["may",        "mayo"],
    ["jun",        "junio",      "june"],
    ["jul",        "julio",      "july"],
    ["ago", "aug", "agosto",     "august"],
    ["sep",        "septiembre", "september"],
    ["oct",        "octubre",    "october"],
    ["nov",        "noviembre",  "november"],
    ["dic", "dec", "diciembre",  "december"],
]


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def _norm(value) -> str:
    if value is None:
        return ""
    t = str(value).strip().lower()
    for a, b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ü","u")]:
        t = t.replace(a, b)
    return t


def _safe_float(value) -> Optional[float]:
    if value is None:
        return None
    try:
        f = float(value)
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None


def _match_month(header_norm: str) -> Optional[int]:
    for idx, variants in enumerate(MONTH_VARIANTS):
        if any(header_norm == v or header_norm.startswith(v) for v in variants):
            return idx
    return None


def _build_result(mensual: List[float], source: str,
                  warning: Optional[str]) -> dict:
    return {
        "mensual": mensual,
        "ppto_4m": round(sum(mensual[:4]), 2),
        "ppto_5m": round(sum(mensual[:5]), 2),
        "anual":   round(sum(mensual),     2),
        "source":  source,
        "warning": warning,
    }


# ─────────────────────────────────────────────────────────────
# PARSER EXCEL
# ─────────────────────────────────────────────────────────────

def _find_sheet(wb) -> Optional[str]:
    target_norm = _norm(SHEET_NAME)
    for name in wb.sheetnames:
        if _norm(name) == target_norm:
            return name
    for name in wb.sheetnames:
        if "presupuesto" in _norm(name) and "pais" in _norm(name):
            return name
    return None


def _parse_month_headers(row: tuple) -> Dict[int, int]:
    month_map: Dict[int, int] = {}
    for j, cell in enumerate(row):
        n = _norm(cell)
        if n in ("total", "anual", "suma"):
            continue
        m = _match_month(n)
        if m is not None and m not in month_map:
            month_map[m] = j
    return month_map


def _read_total_row(rows: List[tuple], start_idx: int,
                    month_map: Dict[int, int]) -> Optional[List[float]]:
    for row in rows[start_idx:]:
        if not any(c is not None for c in row):
            continue
        is_total = False
        for k in range(min(3, len(row))):
            if "total" in _norm(row[k]):
                is_total = True
                break
        if not is_total:
            continue
        mensual = [0.0] * 12
        has_values = False
        for month_idx, col_idx in month_map.items():
            if col_idx < len(row):
                v = _safe_float(row[col_idx])
                if v is not None:
                    mensual[month_idx] = v
                    has_values = True
        if has_values:
            return mensual
    return None


def _load_presupuesto() -> Optional[Dict[str, List[float]]]:
    try:
        import openpyxl
    except ImportError:
        return None

    if not INBOX_EXCEL.exists():
        return None

    try:
        wb = openpyxl.load_workbook(str(INBOX_EXCEL), data_only=True, read_only=True)
    except Exception:
        return None

    sheet_name = _find_sheet(wb)
    if sheet_name is None:
        wb.close()
        return None

    try:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
    except Exception:
        wb.close()
        return None

    wb.close()

    if len(rows) < 10:
        return None

    result: Dict[str, List[float]] = {}

    # ── Chile ─────────────────────────────────────────────────
    cl_header_row = None
    cl_month_map: Dict[int, int] = {}

    for i, row in enumerate(rows):
        row_text = " ".join(_norm(c) for c in row if c is not None)
        if "chile" in row_text and "presupuesto" in row_text:
            if i + 1 < len(rows):
                cl_month_map = _parse_month_headers(rows[i + 1])
                if len(cl_month_map) >= 6:
                    cl_header_row = i + 1
            break

    if cl_header_row is not None and cl_month_map:
        cl_total = _read_total_row(rows, cl_header_row + 1, cl_month_map)
        if cl_total:
            result["chile"] = cl_total

    # ── Perú ──────────────────────────────────────────────────
    pe_header_row = None
    pe_month_map: Dict[int, int] = {}

    for i, row in enumerate(rows):
        row_text = " ".join(_norm(c) for c in row if c is not None)
        if ("peru" in row_text or "per" in row_text) and "presupuesto" in row_text:
            if i + 1 < len(rows):
                pe_month_map = _parse_month_headers(rows[i + 1])
                if len(pe_month_map) >= 6:
                    pe_header_row = i + 1
            break

    if pe_header_row is not None and pe_month_map:
        pe_total = _read_total_row(rows, pe_header_row + 1, pe_month_map)
        if pe_total:
            result["peru"] = pe_total

    return result if result else None


# ─────────────────────────────────────────────────────────────
# UNIVERSO SIC — RTCs presupuestados (CHANGE REQUEST 2026-07-23)
# ─────────────────────────────────────────────────────────────
# La fuente autoritativa del universo SIC es el PRESUPUESTO VIGENTE.
# Solo aparecen individualmente en SIC las personas con presupuesto asignado.
# Personas sin presupuesto → OTROS (trazabilidad en detalle, no en resumen).
#
# Esta función lee dinámicamente los RTCs desde el Libro Base.
# Cuando cambie el archivo de presupuesto (nuevo vendedor o baja),
# el universo SIC se actualiza automáticamente en el siguiente proceso.
# ─────────────────────────────────────────────────────────────

# Entidades que aparecen como "RTC" en el presupuesto pero NO son personas
# comerciales individuales (canales, cooperativas, etc.).
ENTIDADES_NO_PERSONA = {"CAPEL", "RTC ICA 2 / LIZBETH AGUIRRE", "N/A"}

# Alias de nombres — variantes del mismo vendedor por errores de captura o
# cambio de acento. Se mapean a la forma canónica del presupuesto.
# REGLA: solo agregar un alias cuando hay evidencia de que son la misma
# persona (mismo RTC, distinta ortografía en distintas fuentes).
ALIASES_NOMBRES = {
    # Chile
    "FRANCISCO VELASQUEZ": "FRANCISCO VELÁSQUEZ",   # sin acento en ventas
    # Perú
    "LISBETH AGUIRRE":     "LIZBETH AGUIRRE",        # variante L (GG-002)
    "ANTONIO GONZALES":    "ANTONIO GONZALEZ",        # variante S/Z ortográfica
}


def _norm_rtc(nombre: str) -> str:
    """Normaliza nombre de RTC para comparación: mayúsculas sin acentos."""
    t = str(nombre or "").strip().upper()
    for a, b in [("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("Ü","U")]:
        t = t.replace(a, b)
    return t


def leer_universo_sic(libro_base_path: Optional[str] = None) -> Dict[str, dict]:
    """
    Lee el archivo de presupuesto y retorna el universo SIC por país.

    Retorna:
    {
        "CL": {
            "rtcs": ["FRANCISCO VELÁSQUEZ", "PABLO LARATRO", ...],   # forma canónica en ppto
            "claves_norm": {"FRANCISCO VELASQUEZ": "FRANCISCO VELÁSQUEZ", ...},
                            # mapeo: nombre normalizado → nombre canónico ppto
            "fuente": "LIBRO_BASE" | "FALLBACK",
        },
        "PE": { ... }
    }

    Uso en reconciliar_ventas_cl.py:
        universo = leer_universo_sic()
        nombre_norm = _norm_rtc(vendedor)
        canon = universo["CL"]["claves_norm"].get(nombre_norm)
        en_ppto = canon is not None

    REGLA DE NEGOCIO:
        - Si en_ppto=True  → aparece individualmente en SIC como su clave RTC
        - Si en_ppto=False → clasificar como OTROS (mantener trazabilidad en detalle)
    """
    if libro_base_path is None:
        libro_base_path = str(INBOX_EXCEL)

    rtcs_cl: List[str] = []
    rtcs_pe: List[str] = []
    fuente = "FALLBACK"

    try:
        import openpyxl
        wb = openpyxl.load_workbook(libro_base_path, data_only=True)

        # Preferir hoja "Base presupuesto consolidada" (granular, por producto)
        # Fallback: "Presupuesto Pais" (agregado mensual)
        sheet_base_norm = "base presupuesto  consolidada"
        sheet_pais_norm = "presupuesto pais"

        sheet_base = None
        sheet_pais = None
        for name in wb.sheetnames:
            n_norm = _norm(name)
            if n_norm == sheet_base_norm or ("base" in n_norm and "presupuesto" in n_norm and "consolidad" in n_norm):
                sheet_base = name
            if n_norm == sheet_pais_norm or ("presupuesto" in n_norm and "pais" in n_norm):
                sheet_pais = name

        if sheet_base:
            ws = wb[sheet_base]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(v or "").strip() for v in rows[0]]
                try:
                    col_pais = header.index("PAÍS")
                    col_rtc  = header.index("RTC")
                except ValueError:
                    col_pais, col_rtc = 0, 2  # fallback por posición

                seen_cl: set = set()
                seen_pe: set = set()
                for row in rows[1:]:
                    if not any(v for v in row if v):
                        continue
                    pais_val = str(row[col_pais] or "").strip()
                    rtc_val  = str(row[col_rtc]  or "").strip()
                    if not rtc_val or rtc_val == "RTC":
                        continue
                    rtc_norm = _norm_rtc(rtc_val)
                    if rtc_norm in ENTIDADES_NO_PERSONA or _norm_rtc(rtc_val) in {_norm_rtc(e) for e in ENTIDADES_NO_PERSONA}:
                        continue
                    if _norm(pais_val) in ("chile", "cl") and rtc_norm not in seen_cl:
                        rtcs_cl.append(rtc_val)
                        seen_cl.add(rtc_norm)
                    elif _norm(pais_val) in ("peru", "perú", "pe") and rtc_norm not in seen_pe:
                        rtcs_pe.append(rtc_val)
                        seen_pe.add(rtc_norm)

            fuente = "LIBRO_BASE"

        elif sheet_pais:
            # Fallback: leer del "Presupuesto Pais" (solo totales, sin granularidad)
            ws = wb[sheet_pais]
            rows = list(ws.iter_rows(values_only=True))
            country = None
            for row in rows:
                row_text = " ".join(_norm(c) for c in row if c is not None)
                if "chile" in row_text and "presupuesto" in row_text:
                    country = "CL"
                    continue
                if ("peru" in row_text or "per" in row_text) and "presupuesto" in row_text:
                    country = "PE"
                    continue
                if country and row[0]:
                    rtc_raw = str(row[0]).strip()
                    rtc_norm_v = _norm_rtc(rtc_raw)
                    excl = {_norm_rtc(e) for e in ENTIDADES_NO_PERSONA} | {"RTC", "TOTAL", "TOTAL ANUAL"}
                    if rtc_norm_v in excl or not rtc_raw:
                        continue
                    if country == "CL":
                        rtcs_cl.append(rtc_raw)
                    else:
                        rtcs_pe.append(rtc_raw)
            fuente = "LIBRO_BASE_PAIS"

    except Exception as e:
        # Si falla la lectura, universo vacío → todo irá a OTROS hasta que se corrija
        import warnings
        warnings.warn(f"leer_universo_sic: no se pudo leer {libro_base_path}: {e}")
        fuente = "FALLBACK_ERROR"

    # Si no encontramos nada, advertir (no usar lista hardcodeada —
    # queremos que el error sea visible, no silencioso)
    if not rtcs_cl and fuente != "FALLBACK_ERROR":
        import warnings
        warnings.warn("leer_universo_sic: no se encontraron RTCs Chile en el presupuesto")

    def _build_claves_norm(rtcs: List[str]) -> Dict[str, str]:
        """
        Construye mapa: nombre_normalizado → nombre_canónico_ppto
        Incluye aliases definidos en ALIASES_NOMBRES para cubrir
        variantes ortográficas de la misma persona.
        """
        m: Dict[str, str] = {}
        for rtc in rtcs:
            m[_norm_rtc(rtc)] = rtc
        # Agregar aliases: apuntan al canónico del presupuesto cuando exista
        for alias, canon in ALIASES_NOMBRES.items():
            canon_norm = _norm_rtc(canon)
            if canon_norm in m:      # el canónico está en este universo
                m[_norm_rtc(alias)] = m[canon_norm]  # alias → mismo canónico
        return m

    return {
        "CL": {
            "rtcs":        rtcs_cl,
            "claves_norm": _build_claves_norm(rtcs_cl),
            "fuente":      fuente,
        },
        "PE": {
            "rtcs":        rtcs_pe,
            "claves_norm": _build_claves_norm(rtcs_pe),
            "fuente":      fuente,
        },
    }


# Mapa nombre_canónico_presupuesto → clave SIC (debe coincidir con VENDEDOR_MAP en sic_data_adapter.js).
# REGLA: este dict define SOLO la correspondencia nombre↔clave.
# La decisión de quién es individual vs OTROS la toma el presupuesto vigente, no este dict.
# Cuando un nuevo vendedor se incorpora al presupuesto, se añade aquí una entrada.
NOMBRE_A_CLAVE_SIC: Dict[str, Dict[str, str]] = {
    "CL": {
        "FRANCO RIFFO":        "franco_riffo",
        "FRANCISCO VELÁSQUEZ": "velasquez",
        "RODRIGO ENCINA":      "encina",
        "VALENTINA MUÑOZ":     "munoz",
        "PABLO LARATRO":       "laratro",
        "JORGE CAROCA":        "caroca",
    },
    "PE": {
        "OMAR ATALAYA":          "atalaya",
        "ANTONIO GONZALEZ":      "gonzales",
        "LIZBETH AGUIRRE":       "aguirre",
        "OSCAR INFANTE":         "infante",
        "PATRICIA VALLADARES":   "valladares",
        "SUSAN DIAZ":            "diaz",
        "MARTHA HIDALGO - KAM":  "martha",
        "NICOLL NAVARRO":        "navarro",
    },
}


def _auto_clave_sic(nombre: str) -> str:
    """Fallback: auto-deriva clave SIC desde nombre si no está en NOMBRE_A_CLAVE_SIC."""
    import re
    t = _norm_rtc(nombre)
    return re.sub(r"[^A-Z0-9]+", "_", t).strip("_").lower()


def generar_universo_sic(pais: str, output_path: Optional[str] = None,
                          universo: Optional[Dict[str, dict]] = None) -> dict:
    """
    Genera el archivo universo_sic_CL.json (o PE) desde el presupuesto vigente.

    La lista de personas presupuestadas viene de leer_universo_sic() (Libro Base).
    La conversión nombre→clave SIC usa NOMBRE_A_CLAVE_SIC; si no hay entrada,
    se auto-deriva una clave y se registra en "advertencias_mapeo".

    Parámetros:
        pais        : "CL" | "PE"
        output_path : ruta absoluta donde escribir el JSON (opcional)
        universo    : resultado de leer_universo_sic() (se llama internamente si es None)

    Retorna dict:
    {
        "schema_version": "1.0",
        "pais": "CL",
        "fuente": "LIBRO_BASE",
        "fecha_proceso": "2026-07-23",
        "claves_presupuestadas": ["caroca", "encina", ...],      # SIC keys de ppto vigente
        "nombres_canonicos_por_clave": {"caroca": "JORGE CAROCA", ...},
        "advertencias_mapeo": [...]   # solo si hay claves auto-derivadas
    }
    """
    import json
    from datetime import date as _date

    if universo is None:
        universo = leer_universo_sic()

    pais_up = pais.upper()
    pais_data = universo.get(pais_up, {"rtcs": [], "claves_norm": {}, "fuente": "NO_DISPONIBLE"})
    rtcs = pais_data.get("rtcs", [])
    fuente = pais_data.get("fuente", "DESCONOCIDO")

    # Reverse-lookup: _norm_rtc(nombre) → clave_sic
    mapa_sic = NOMBRE_A_CLAVE_SIC.get(pais_up, {})
    mapa_norm: Dict[str, str] = {_norm_rtc(k): v for k, v in mapa_sic.items()}

    claves_presupuestadas: List[str] = []
    nombres_por_clave: Dict[str, str] = {}
    sin_mapeo: List[str] = []

    for rtc_nombre in rtcs:
        rtc_n = _norm_rtc(rtc_nombre)
        clave_sic = mapa_norm.get(rtc_n)
        if clave_sic is None:
            # Auto-derivar — nunca descartar silenciosamente
            clave_sic = _auto_clave_sic(rtc_nombre)
            sin_mapeo.append(rtc_nombre)
        claves_presupuestadas.append(clave_sic)
        nombres_por_clave[clave_sic] = rtc_nombre  # forma canónica del presupuesto

    result: dict = {
        "schema_version": "1.0",
        "pais": pais_up,
        "fuente": fuente,
        "fecha_proceso": _date.today().isoformat(),
        "claves_presupuestadas": sorted(claves_presupuestadas),
        "nombres_canonicos_por_clave": nombres_por_clave,
    }
    if sin_mapeo:
        result["advertencias_mapeo"] = [
            f"RTC '{n}' no está en NOMBRE_A_CLAVE_SIC['{pais_up}'] — clave auto-derivada: "
            f"'{_auto_clave_sic(n)}' — agregar entrada si es correcto"
            for n in sin_mapeo
        ]

    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

    return result


def es_en_universo_sic(pais: str, nombre_vendedor: str,
                        universo: Optional[Dict[str, dict]] = None) -> dict:
    """
    Verifica si un vendedor está en el universo SIC del presupuesto vigente.

    Retorna:
    {
        "en_ppto":     True | False,
        "canon":       "FRANCISCO VELÁSQUEZ" | None,  # forma canónica del presupuesto
        "categoria":   "velasquez" | "OTROS",         # clave SIC o "OTROS"
        "alias_usado": True | False,
    }

    Ejemplo:
        info = es_en_universo_sic("CL", "FRANCISCO VELASQUEZ")
        # → {"en_ppto": True, "canon": "FRANCISCO VELÁSQUEZ", "categoria": ..., "alias_usado": True}

        info = es_en_universo_sic("CL", "RAYEN BERNAZAR")
        # → {"en_ppto": False, "canon": None, "categoria": "OTROS", "alias_usado": False}
    """
    if universo is None:
        universo = leer_universo_sic()

    pais_key = pais.upper()
    universo_pais = universo.get(pais_key, {"rtcs": [], "claves_norm": {}})
    nombre_norm = _norm_rtc(nombre_vendedor)
    canon = universo_pais["claves_norm"].get(nombre_norm)
    alias_usado = canon is not None and _norm_rtc(canon) != nombre_norm

    return {
        "en_ppto":     canon is not None,
        "canon":       canon,
        "categoria":   "OTROS" if canon is None else canon,
        "alias_usado": alias_usado,
    }


# ─────────────────────────────────────────────────────────────
# FUNCIÓN PRINCIPAL
# ─────────────────────────────────────────────────────────────

def get_ppto_all() -> Dict[str, dict]:
    """
    Retorna presupuesto de Chile y Perú.
    Fallback independiente por país: LIBRO_BASE → LEGACY.

    Retorna:
    {
        "chile": { mensual, ppto_4m, ppto_5m, anual, source, warning },
        "peru":  { mensual, ppto_4m, ppto_5m, anual, source, warning },
    }
    """
    libro = _load_presupuesto()
    resultados: Dict[str, dict] = {}

    cl_data = libro.get("chile") if libro else None
    if cl_data and sum(cl_data) > 0:
        resultados["chile"] = _build_result(cl_data, "LIBRO_BASE", None)
    else:
        msg = (
            f"⚠️  PPTO CHILE: no se pudo leer desde hoja '{SHEET_NAME}'. "
            f"Usando LEGACY: anual={PPTO_ANUAL_CL_LEGACY:,.0f} CLP"
        )
        print(msg)
        resultados["chile"] = _build_result(
            list(PPTO_MENSUAL_CL_LEGACY), "LEGACY", msg
        )

    pe_data = libro.get("peru") if libro else None
    if pe_data and sum(pe_data) > 0:
        resultados["peru"] = _build_result(pe_data, "LIBRO_BASE", None)
    else:
        msg = (
            f"⚠️  PPTO PERÚ: no se pudo leer desde hoja '{SHEET_NAME}'. "
            f"Usando LEGACY: anual={PPTO_ANUAL_PE_LEGACY:,.0f} USD"
        )
        print(msg)
        resultados["peru"] = _build_result(
            list(PPTO_MENSUAL_PE_LEGACY), "LEGACY", msg
        )

    return resultados


# ─────────────────────────────────────────────────────────────
# VALIDACIÓN
# ─────────────────────────────────────────────────────────────

def validar_resultado(ppto: Dict[str, dict]) -> bool:
    TOL = 1.0
    checks = [
        ("Chile anual",   ppto["chile"]["anual"],  EXPECTED_CL_ANUAL),
        ("Chile ppto_5m", ppto["chile"]["ppto_5m"], EXPECTED_CL_PPTO5M),
        ("Peru anual",    ppto["peru"]["anual"],   EXPECTED_PE_ANUAL),
        ("Peru ppto_5m",  ppto["peru"]["ppto_5m"],  EXPECTED_PE_PPTO5M),
    ]
    all_pass = True
    for label, got, expected in checks:
        ok = abs(got - expected) <= TOL
        status = "✅" if ok else "❌"
        if not ok:
            all_pass = False
        print(f"  {status} {label:<18} esperado={expected:>16,.1f}  obtenido={got:>16,.1f}")
    return all_pass


# ─────────────────────────────────────────────────────────────
# TEST STANDALONE
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    MESES = ["Ene","Feb","Mar","Abr","May","Jun",
             "Jul","Ago","Sep","Oct","Nov","Dic"]

    print("=" * 65)
    print("  AVBOARD — Módulo presupuesto por país")
    print("=" * 65)
    print()

    ppto = get_ppto_all()

    for pais, moneda in [("chile", "CLP"), ("peru", "USD")]:
        p = ppto[pais]
        print(f"{'─'*45}")
        print(f"  {pais.upper()} [{p['source']}] — {moneda}")
        print(f"{'─'*45}")
        print(f"  Anual    : {moneda} {p['anual']:>18,.1f}")
        print(f"  Ppto 5m  : {moneda} {p['ppto_5m']:>18,.1f}")
        print(f"  Ppto 4m  : {moneda} {p['ppto_4m']:>18,.1f}")
        print()
        for i, (mes, val) in enumerate(zip(MESES, p["mensual"])):
            marca = " ◄ 5m" if i == 4 else ""
            print(f"    {mes}  {val:>18,.1f}{marca}")
        if p["warning"]:
            print(f"\n  {p['warning']}")
        print()

    print("─" * 45)
    print("  VALIDACIÓN vs valores esperados del Libro Base")
    print("─" * 45)
    ok = validar_resultado(ppto)
    print()
    if ok:
        print("  ✅ PASS TOTAL — valores coinciden con Libro Base real")
    else:
        print("  ❌ FAIL — revisar parser o estructura de la hoja")

    print()
    print("Constantes LEGACY:")
    print(f"  CL anual = {PPTO_ANUAL_CL_LEGACY:>18,.0f} CLP")
    print(f"  PE anual = {PPTO_ANUAL_PE_LEGACY:>18,.0f} USD")
